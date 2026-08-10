import csv
import math
import re
from io import BytesIO, StringIO
from openpyxl import load_workbook
from datetime import date, time, timedelta
from decimal import Decimal
from unittest.mock import Mock, patch
from uuid import uuid4

from django.core.cache import cache
from django.core import mail
from django.core.management import call_command
from django.contrib.auth.models import User
from django.db import connection
from django.test import override_settings
from django.utils import timezone
from django.test.utils import CaptureQueriesContext
from rest_framework.test import APITestCase, APIRequestFactory, force_authenticate
from rest_framework_simplejwt.token_blacklist.models import BlacklistedToken, OutstandingToken
from rest_framework_simplejwt.tokens import RefreshToken

from .chat.services import (
    get_current_booking_context,
    get_current_booking_contexts,
    get_partner_context,
    serialize_booking_context,
)
from .paymongo_money_movement import PayMongoCashOutError
from .views import (
    COUNTED_STRIKE_WALLET_DEDUCTION,
    credit_tutor_wallet,
    dev_add_wallet_funds,
    dev_remove_wallet_funds,
    WEEKDAY_MAP,
)
from .models import (
    Booking,
    Course,
    EmailOTPChallenge,
    GRACE_CUTOFF_HOURS,
    InstitutionRequest,
    PartnerInstitution,
    Payment,
    PaymentMethod,
    Rating,
    Subjects,
    SupportTicket,
    SessionCheckIn,
    Tutor,
    TutorAvailability,
    TutorAvailabilityOverride,
    TutorSubjects,
    UserProfile,
    Transaction,
    TutorApplication,
    TutorDocumentRenewalReview,
    TuteeApplication,
    TuteeDocumentRenewalReview,
    PlatformActivity,
    Wallet,
    WalletTopUp,
    WithdrawalRequest,
    Preference,
)
from .chat.models import ChatRoom, Message


class SuperAdminRedesignApiTests(APITestCase):
    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="Central Philippine University",
            school_email_domain="cpu.edu.ph",
            is_active=True,
            contact_person="Registrar",
        )
        self.inactive_institution = PartnerInstitution.objects.create(
            institution_name="Dormant College",
            school_email_domain="dormant.edu.ph",
            is_active=False,
        )
        self.super_user = User.objects.create_user(
            username="superadmin",
            email="superadmin@studybuddy.test",
            password="password",
        )
        self.super_profile = UserProfile.objects.create(
            user=self.super_user,
            fname="Super",
            mname="",
            lname="Admin",
            role="SuperAdmin",
            profile_completed=True,
            is_domain_exempt=True,
        )
        self.target_user = User.objects.create_user(
            username="target",
            email="target@cpu.edu.ph",
            password="password",
        )
        self.target_profile = UserProfile.objects.create(
            user=self.target_user,
            fname="Target",
            mname="",
            lname="User",
            role="Tutee",
            institution=self.institution,
            profile_completed=True,
        )
        self.domain_user = User.objects.create_user(
            username="external",
            email="external@gmail.com",
            password="password",
        )
        self.domain_profile = UserProfile.objects.create(
            user=self.domain_user,
            fname="External",
            mname="",
            lname="Learner",
            role="Tutee",
            profile_completed=True,
            is_domain_exempt=False,
        )
        self.client.force_authenticate(user=self.super_user)

    def test_pending_actions_aggregates_superadmin_only_items(self):
        InstitutionRequest.objects.create(
            institution_name="West Visayas State University",
            school_email_domain="wvsu.edu.ph",
            contact_person="WVSU Admin",
            contact_email="admin@wvsu.edu.ph",
        )
        response = self.client.get("/api/admin/pending-actions/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 3)
        self.assertEqual(
            {
                item["type"]
                for item in response.data["items"]
            },
            {
                "institution_request",
                "institution_activation",
                "domain_exemption",
            },
        )

    def test_pending_actions_includes_pending_verifications(self):
        application = TuteeApplication.objects.create(
            profile=self.target_profile,
            application_status="pending",
        )
        TuteeApplication.objects.create(
            profile=self.domain_profile,
            application_status="approved",
        )

        response = self.client.get("/api/admin/pending-actions/")

        self.assertEqual(response.status_code, 200)
        verification_items = [
            item for item in response.data["items"]
            if item["type"] == "tutee_application"
        ]
        self.assertEqual(len(verification_items), 1)
        self.assertEqual(verification_items[0]["id"], application.id)
        self.assertIn(self.target_user.email, verification_items[0]["meta"])

    def test_institution_request_approval_creates_active_partner_institution(self):
        request_obj = InstitutionRequest.objects.create(
            institution_name="West Visayas State University",
            school_email_domain="wvsu.edu.ph",
            contact_person="WVSU Admin",
            contact_email="admin@wvsu.edu.ph",
        )

        response = self.client.patch(
            f"/api/admin/institution-requests/{request_obj.id}/",
            {"action": "approve"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["status"], "approved")
        self.assertTrue(
            PartnerInstitution.objects.filter(
                school_email_domain="wvsu.edu.ph",
                is_active=True,
            ).exists()
        )

    def test_superadmin_can_patch_role_institution_and_domain_exemption(self):
        response = self.client.patch(
            f"/api/admin/users/{self.target_profile.id}/",
            {
                "role": "Tutor",
                "institution": self.inactive_institution.id,
                "is_domain_exempt": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.target_profile.refresh_from_db()
        self.assertEqual(self.target_profile.role, "Tutor")
        self.assertEqual(self.target_profile.institution, self.inactive_institution)
        self.assertTrue(self.target_profile.is_domain_exempt)

    def test_superadmin_can_patch_tutor_session_load_limit(self):
        tutor_user = User.objects.create_user(
            username="limit-tutor",
            email="limit-tutor@cpu.edu.ph",
            password="password",
        )
        tutor_profile = UserProfile.objects.create(
            user=tutor_user,
            fname="Limit",
            mname="",
            lname="Tutor",
            role="Tutor",
            institution=self.institution,
            profile_completed=True,
        )
        tutor = Tutor.objects.create(profile=tutor_profile)

        response = self.client.patch(
            f"/api/admin/users/{tutor_profile.id}/",
            {"session_load_limit": 14},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        tutor.refresh_from_db()
        self.assertEqual(tutor.session_load_limit, 14)
        self.assertEqual(response.data["tutor_session_load_limit"], 14)

    @staticmethod
    def _export_workbook(response):
        """Load the exported workbook, so tests assert on sheets and cells rather than raw text."""
        return load_workbook(BytesIO(response.content))

    @staticmethod
    def _sheet_rows(workbook, title):
        return list(workbook[title].values)

    def test_analytics_includes_completion_subject_popularity_and_workbook(self):
        course = Course.objects.create(course_code="MATH", course_name="Mathematics")
        subject = Subjects.objects.create(
            subject_code="ALG101",
            subject_name="College Algebra",
            department="Math",
        )
        tutor_user = User.objects.create_user(
            username="tutor",
            email="tutor@cpu.edu.ph",
            password="password",
        )
        tutor_profile = UserProfile.objects.create(
            user=tutor_user,
            fname="Tutor",
            mname="",
            lname="One",
            role="Tutor",
            institution=self.institution,
            course=course,
            profile_completed=True,
        )
        tutor = Tutor.objects.create(
            profile=tutor_profile,
            hourly_rate=Decimal("500.00"),
            total_sessions=1,
        )
        TutorSubjects.objects.create(tutor=tutor, subject=subject, expertise_level=5)
        availability = TutorAvailability.objects.create(
            tutor=tutor,
            day="Mon",
            time_slot=time(9, 0),
            is_active=True,
        )
        booking = Booking.objects.create(
            student=self.target_profile,
            tutor=tutor,
            availability=availability,
            session_date=date.today(),
            session_mode="Online",
            status="Completed",
            subject=subject,
        )
        payment_method, _ = PaymentMethod.objects.get_or_create(
            code="online", defaults={"method_name": "Online Payment"}
        )
        Payment.objects.create(
            booking=booking,
            method=payment_method,
            amount=Decimal("500.00"),
            payment_status="Paid",
        )

        response = self.client.get("/api/admin/analytics/?period=7d")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["completion_rate"], 100.0)
        self.assertEqual(response.data["subject_popularity"][0]["subject_name"], "College Algebra")
        self.assertEqual(response.data["top_tutors"][0]["earnings"], 450.0)

        export_response = self.client.get("/api/admin/analytics/export/?period=7d")
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn('filename="studybuddy-report-', export_response["Content-Disposition"])
        self.assertTrue(export_response["Content-Disposition"].endswith('.xlsx"'))

        workbook = self._export_workbook(export_response)
        self.assertEqual(workbook.sheetnames, [
            "Report Info", "Summary", "Sessions Over Time",
            "Top Tutors", "Subject Popularity", "Booking Transactions",
        ])

        # Scope metadata lives on its own sheet, so each data sheet starts at A1 with its header.
        info = dict(self._sheet_rows(workbook, "Report Info")[1:])
        self.assertEqual(info["Period"], "7d")
        self.assertEqual(info["Institution filter"], "All institutions")

        # Summary is transposed to label/value rows.
        summary = dict(self._sheet_rows(workbook, "Summary")[1:])
        self.assertEqual(summary["Completed Sessions"], 1)
        self.assertEqual(summary["Gross Revenue"], 500.0)
        self.assertEqual(summary["Commissions"], 50.0)
        self.assertEqual(summary["Tutor Payouts"], 450.0)

        transactions = self._sheet_rows(workbook, "Booking Transactions")
        self.assertEqual(transactions[0], (
            "Session Date", "Institution", "Tutor", "Tutee", "Subject",
            "Amount", "Commission", "Payout",
        ))
        self.assertEqual(transactions[1][1:5], (
            self.institution.institution_name, "Tutor One", "Target User", "College Algebra",
        ))
        self.assertEqual(transactions[1][5:], (500.0, 50.0, 450.0))

        self.assertEqual(
            self._sheet_rows(workbook, "Subject Popularity")[1], ("College Algebra", 1)
        )
        self.assertEqual(self._sheet_rows(workbook, "Top Tutors")[1][:2], ("Tutor One", 1))

    def test_analytics_export_writes_only_the_requested_sections(self):
        self.client.force_authenticate(user=self.super_user)

        response = self.client.get("/api/admin/analytics/export/?period=7d&sections=summary")

        self.assertEqual(response.status_code, 200)
        # A single section names the file after itself, so a folder of exports stays readable.
        self.assertIn('filename="studybuddy-summary-', response["Content-Disposition"])
        # Report Info is metadata rather than a tickable section, so it is always present.
        self.assertEqual(
            self._export_workbook(response).sheetnames, ["Report Info", "Summary"]
        )

    def test_analytics_export_ignores_unknown_sections_and_defaults_to_all(self):
        self.client.force_authenticate(user=self.super_user)

        all_sheets = [
            "Report Info", "Summary", "Sessions Over Time",
            "Top Tutors", "Subject Popularity", "Booking Transactions",
        ]

        # Omitted entirely: the pre-`sections` caller shape must keep working.
        default_response = self.client.get("/api/admin/analytics/export/?period=7d")
        # Only unrecognised slugs: falls back rather than emitting a file with no data sheets.
        bogus_response = self.client.get(
            "/api/admin/analytics/export/?period=7d&sections=nonsense"
        )

        self.assertEqual(self._export_workbook(default_response).sheetnames, all_sheets)
        self.assertEqual(self._export_workbook(bogus_response).sheetnames, all_sheets)
        self.assertIn('filename="studybuddy-report-', bogus_response["Content-Disposition"])

    def test_analytics_export_writes_multiple_sections_in_file_order(self):
        self.client.force_authenticate(user=self.super_user)

        response = self.client.get(
            "/api/admin/analytics/export/?period=7d&sections=transactions,summary"
        )

        self.assertEqual(
            self._export_workbook(response).sheetnames,
            ["Report Info", "Summary", "Booking Transactions"],
        )
        self.assertIn('filename="studybuddy-report-', response["Content-Disposition"])

    def test_analytics_export_marks_empty_sections_rather_than_leaving_them_blank(self):
        """An empty result must be distinguishable from a broken export."""
        self.client.force_authenticate(user=self.super_user)

        response = self.client.get(
            "/api/admin/analytics/export/?period=7d&sections=transactions"
        )
        rows = self._sheet_rows(self._export_workbook(response), "Booking Transactions")

        self.assertEqual(rows[0][0], "Session Date")
        self.assertEqual(rows[1][0], "No data for this period.")

    def test_analytics_export_keeps_formula_like_names_as_text(self):
        """openpyxl writes '='-prefixed strings as live formulas unless the type is pinned."""
        self.client.force_authenticate(user=self.super_user)
        self.institution.institution_name = "=cmd()"
        self.institution.save(update_fields=["institution_name"])

        response = self.client.get(
            f"/api/admin/analytics/export/?period=7d&institution_id={self.institution.pk}"
        )
        cell = self._export_workbook(response)["Report Info"]["B5"]

        self.assertEqual(cell.value, "=cmd()")
        self.assertEqual(cell.data_type, "s")


class SuperAdminUserDetailApiTests(APITestCase):
    def setUp(self):
        self.super_user = User.objects.create_user(
            username="detail-superadmin",
            email="detail-superadmin@studybuddy.test",
            password="password",
        )
        UserProfile.objects.create(
            user=self.super_user,
            fname="Detail",
            mname="",
            lname="Admin",
            role="SuperAdmin",
            profile_completed=True,
        )
        self.tutee_user = User.objects.create_user(
            username="detail-tutee",
            email="detail-tutee@studybuddy.test",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Taylor",
            mname="",
            lname="Tutee",
            role="Tutee",
            profile_completed=True,
        )
        tutor_user = User.objects.create_user(
            username="detail-tutor",
            email="detail-tutor@studybuddy.test",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=tutor_user,
            fname="Robin",
            mname="",
            lname="Tutor",
            role="Tutor",
            profile_completed=True,
        )
        self.tutor = Tutor.objects.create(profile=self.tutor_profile)
        self.subject = Subjects.objects.create(
            subject_code="DETAIL101",
            subject_name="Detail Testing",
            department="Quality",
        )
        self.morning_availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=time(9, 0),
            is_active=True,
        )
        self.afternoon_availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Tue",
            time_slot=time(14, 0),
            is_active=True,
        )
        self.bookings = []
        for offset in range(12):
            self.bookings.append(
                Booking.objects.create(
                    student=self.tutee_profile,
                    tutor=self.tutor,
                    availability=(
                        self.afternoon_availability if offset == 11 else self.morning_availability
                    ),
                    subject=self.subject,
                    session_date=date(2026, 7, 1) + timedelta(days=offset),
                    session_mode="Online",
                    status="Completed",
                )
            )
        self.client.force_authenticate(user=self.super_user)

    def test_bookings_returns_most_recent_page_with_both_party_names(self):
        response = self.client.get(
            f"/api/admin/users/{self.tutee_profile.id}/bookings/"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["total"], 12)
        self.assertEqual(response.data["page"], 1)
        self.assertEqual(response.data["page_size"], 10)
        self.assertEqual(len(response.data["results"]), 10)
        self.assertEqual(response.data["results"][0]["id"], self.bookings[-1].id)
        self.assertEqual(response.data["results"][0]["tutor_name"], "Robin Tutor")
        self.assertEqual(response.data["results"][0]["tutee_name"], "Taylor Tutee")

    def test_bookings_filters_inclusively_and_rejects_invalid_ranges(self):
        url = f"/api/admin/users/{self.tutor_profile.id}/bookings/"
        response = self.client.get(
            url,
            {"date_from": "2026-07-03", "date_to": "2026-07-05"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["total"], 3)
        self.assertEqual(
            [row["date"] for row in response.data["results"]],
            ["2026-07-05", "2026-07-04", "2026-07-03"],
        )

        invalid_date = self.client.get(url, {"date_from": "not-a-date"})
        reversed_range = self.client.get(
            url,
            {"date_from": "2026-07-05", "date_to": "2026-07-03"},
        )
        self.assertEqual(invalid_date.status_code, 400)
        self.assertEqual(reversed_range.status_code, 400)

    def test_bookings_export_matches_the_json_filter(self):
        params = {"date_from": "2026-07-03", "date_to": "2026-07-05"}
        json_response = self.client.get(
            f"/api/admin/users/{self.tutee_profile.id}/bookings/",
            params,
        )
        export_response = self.client.get(
            f"/api/admin/users/{self.tutee_profile.id}/bookings/export/",
            params,
        )
        csv_rows = list(csv.reader(StringIO(export_response.content.decode())))

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_response["Content-Type"], "text/csv")
        self.assertEqual(
            csv_rows[0],
            ["date", "time", "subject", "tutor name", "tutee name", "mode", "status"],
        )
        self.assertEqual(len(csv_rows) - 1, json_response.data["total"])
        self.assertEqual(csv_rows[1][0], json_response.data["results"][0]["date"])

    def test_bookings_endpoints_forbid_non_superadmin_users(self):
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get(
            f"/api/admin/users/{self.tutee_profile.id}/bookings/"
        )
        export_response = self.client.get(
            f"/api/admin/users/{self.tutee_profile.id}/bookings/export/"
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(export_response.status_code, 403)

    def test_availability_returns_all_slots_with_state_and_booking_id(self):
        inactive = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Wed",
            time_slot=time(10, 0),
            is_active=False,
            is_booked=False,
        )
        booked = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Thu",
            time_slot=time(11, 0),
            is_active=True,
            is_booked=True,
        )
        linked_booking = Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=booked,
            subject=self.subject,
            session_date=date(2026, 8, 1),
            session_mode="F2F",
            status="Confirmed",
        )

        response = self.client.get(
            f"/api/admin/users/{self.tutor_profile.id}/availability/"
        )

        self.assertEqual(response.status_code, 200)
        rows_by_id = {row["id"]: row for row in response.data}
        self.assertEqual(len(rows_by_id), 4)
        self.assertEqual(rows_by_id[self.morning_availability.id]["state"], "open")
        self.assertEqual(rows_by_id[inactive.id]["state"], "inactive")
        self.assertEqual(rows_by_id[booked.id]["state"], "booked")
        self.assertEqual(rows_by_id[booked.id]["booking_id"], linked_booking.id)

    def test_availability_rejects_tutees_and_exports_matching_rows(self):
        tutee_response = self.client.get(
            f"/api/admin/users/{self.tutee_profile.id}/availability/"
        )
        json_response = self.client.get(
            f"/api/admin/users/{self.tutor_profile.id}/availability/"
        )
        export_response = self.client.get(
            f"/api/admin/users/{self.tutor_profile.id}/availability/export/"
        )
        csv_rows = list(csv.reader(StringIO(export_response.content.decode())))

        self.assertEqual(tutee_response.status_code, 400)
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_response["Content-Type"], "text/csv")
        self.assertEqual(csv_rows[0], ["day", "time slot", "state", "booking id"])
        self.assertEqual(len(csv_rows) - 1, len(json_response.data))

    def test_tutor_stats_computes_ratings_subjects_cancellations_and_counterparts(self):
        TutorSubjects.objects.create(
            tutor=self.tutor,
            subject=self.subject,
            expertise_level=4,
        )
        for booking, score in zip(self.bookings[:5], [1, 3, 5, 5, 5]):
            Rating.objects.create(
                booking=booking,
                student=self.tutee_profile,
                tutor=self.tutor,
                rating_score=score,
                comment=f"Review {score}",
            )

        alpha_profile = self._create_tutee("Alpha", "Learner")
        beta_profile = self._create_tutee("Beta", "Learner")
        for profile, count in [(alpha_profile, 3), (beta_profile, 3)]:
            for offset in range(count):
                Booking.objects.create(
                    student=profile,
                    tutor=self.tutor,
                    availability=self.morning_availability,
                    subject=self.subject,
                    session_date=date(2026, 9, 1) + timedelta(days=offset),
                    session_mode="Online",
                    status="Cancelled",
                    cancelled_by_role="tutor" if offset == 0 else "tutee",
                )

        response = self.client.get(
            f"/api/admin/users/{self.tutor_profile.id}/stats/"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(sum(response.data["rating_distribution"].values()), 5)
        self.assertEqual(response.data["rating_distribution"]["5"], 3)
        self.assertEqual(response.data["recent_reviews"][0]["reviewer_name"], "Taylor Tutee")
        self.assertEqual(response.data["subjects"], [{"name": "Detail Testing", "expertise_level": 4}])
        self.assertEqual(response.data["cancellations"], {"by_user": 2, "by_counterpart": 4})
        self.assertEqual(
            response.data["top_counterparts"],
            [
                {"name": "Taylor Tutee", "session_count": 12},
                {"name": "Alpha Learner", "session_count": 3},
                {"name": "Beta Learner", "session_count": 3},
            ],
        )

    def test_tutee_stats_omits_tutor_only_rating_keys(self):
        response = self.client.get(
            f"/api/admin/users/{self.tutee_profile.id}/stats/"
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("rating_distribution", response.data)
        self.assertNotIn("recent_reviews", response.data)
        self.assertEqual(response.data["subjects"], [{"name": "Detail Testing"}])
        self.assertEqual(
            response.data["top_counterparts"],
            [{"name": "Robin Tutor", "session_count": 12}],
        )

    def test_stats_export_flattens_computed_values(self):
        Rating.objects.create(
            booking=self.bookings[0],
            student=self.tutee_profile,
            tutor=self.tutor,
            rating_score=5,
            comment="Clear explanation",
        )

        export_response = self.client.get(
            f"/api/admin/users/{self.tutor_profile.id}/stats/export/"
        )
        csv_rows = list(csv.reader(StringIO(export_response.content.decode())))

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_response["Content-Type"], "text/csv")
        self.assertEqual(csv_rows[0], ["stat name", "value"])
        self.assertIn(["rating 5 stars", "1"], csv_rows)
        self.assertIn(["counterpart 1", "Taylor Tutee (12 sessions)"], csv_rows)

    def _create_tutee(self, first_name, last_name):
        username = f"{first_name.lower()}-{last_name.lower()}"
        user = User.objects.create_user(
            username=username,
            email=f"{username}@studybuddy.test",
            password="password",
        )
        return UserProfile.objects.create(
            user=user,
            fname=first_name,
            mname="",
            lname=last_name,
            role="Tutee",
            profile_completed=True,
        )


class GlobalSubjectCatalogTests(APITestCase):
    def setUp(self):
        self.course = Course.objects.create(course_code="BSCS", course_name="Computer Science")
        self.subject = Subjects.objects.create(
            subject_code="CS101",
            subject_name="Introduction to Computing",
            department="Computer Science",
            category="Technology & Computer Science",
        )
        self.super_user = User.objects.create_user(
            username="super",
            email="super@studybuddy.test",
            password="password",
        )
        UserProfile.objects.create(
            user=self.super_user,
            fname="Super",
            mname="",
            lname="Admin",
            role="SuperAdmin",
            profile_completed=True,
            is_domain_exempt=True,
        )
        self.client.force_authenticate(user=self.super_user)

    def test_catalog_is_global_subject_crud_without_institution_fields(self):
        list_response = self.client.get("/api/admin/course-catalog/?institution_id=999")
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual([item["subject_code"] for item in list_response.data], ["CS101"])
        self.assertNotIn("institution", list_response.data[0])
        self.assertNotIn("owning_institution", list_response.data[0])

        create_response = self.client.post(
            "/api/admin/course-catalog/",
            {
                "subject_code": "AI201",
                "subject_name": "Applied Artificial Intelligence",
                "department": "Computer Science",
                "category": "Technology & Computer Science",
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, 201)

        update_response = self.client.patch(
            "/api/admin/course-catalog/AI201/",
            {"subject_name": "Applied AI"},
            format="json",
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.data["subject_name"], "Applied AI")

        delete_response = self.client.delete("/api/admin/course-catalog/AI201/")
        self.assertEqual(delete_response.status_code, 204)
        self.assertFalse(Subjects.objects.filter(subject_code="AI201").exists())

    def test_per_institution_catalog_model_is_removed(self):
        from django.apps import apps

        with self.assertRaises(LookupError):
            apps.get_model('studybuddy', 'InstitutionCourseCatalog')

    def test_catalog_create_and_update_persist_keywords(self):
        create_response = self.client.post(
            "/api/admin/course-catalog/",
            {
                "subject_code": "AI202",
                "subject_name": "Machine Learning",
                "department": "Computer Science",
                "category": "Technology & Computer Science",
                "keywords": "ml, coding, ai",
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, 201)
        self.assertEqual(create_response.data["keywords"], "ml, coding, ai")

        update_response = self.client.patch(
            "/api/admin/course-catalog/AI202/",
            {"keywords": "ml, coding, ai, deep learning"},
            format="json",
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.data["keywords"], "ml, coding, ai, deep learning")

    def test_catalog_keywords_defaults_to_blank(self):
        self.assertEqual(self.subject.keywords, "")

    def test_catalog_create_and_update_persist_description(self):
        create_response = self.client.post(
            "/api/admin/course-catalog/",
            {
                "subject_code": "AI203",
                "subject_name": "Neural Networks",
                "department": "Computer Science",
                "category": "Technology & Computer Science",
                "description": "Layered models trained to recognize patterns in data.",
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, 201)
        self.assertEqual(
            create_response.data["description"],
            "Layered models trained to recognize patterns in data.",
        )

        update_response = self.client.patch(
            "/api/admin/course-catalog/AI203/",
            {"description": "Updated description."},
            format="json",
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.data["description"], "Updated description.")

    def test_catalog_description_defaults_to_blank(self):
        self.assertEqual(self.subject.description, "")


class RecommendTutorsViewTests(APITestCase):
    def setUp(self):
        self.subject = Subjects.objects.create(
            subject_code="MATH101",
            subject_name="College Algebra",
            department="Math",
        )
        self.other_subject = Subjects.objects.create(
            subject_code="SCI101",
            subject_name="Science",
            department="Science",
        )
        self.student_user = User.objects.create_user(
            username="student",
            email="student@example.com",
            password="password",
        )
        self.student_profile = UserProfile.objects.create(
            user=self.student_user,
            fname="Student",
            mname="",
            lname="User",
            role="Tutee",
            year_level=11,
        )
        self.client.force_authenticate(user=self.student_user)
        # The next Monday (add_slots seeds "Mon"), always inside the booking window and never today,
        # so past-time filtering cannot make these assertions flaky.
        today = timezone.localdate()
        self.search_date = today + timedelta(days=((0 - today.weekday()) % 7) or 7)

    def create_tutor(
        self,
        username,
        subject=None,
        hourly_rate=200,
        can_online=True,
        can_f2f=False,
    ):
        from django.core.files.uploadedfile import SimpleUploadedFile

        user = User.objects.create_user(
            username=username,
            email=f"{username}@example.com",
            password="password",
        )
        profile = UserProfile.objects.create(
            user=user,
            fname=username.title(),
            mname="",
            lname="Tutor",
            role="Tutor",
            year_level=12,
        )
        tutor = Tutor.objects.create(
            profile=profile,
            hourly_rate=hourly_rate,
            can_online=can_online,
            can_f2f=can_f2f,
            teaching_level="SHS",
        )
        TutorSubjects.objects.create(
            tutor=tutor,
            subject=subject or self.subject,
            expertise_level=5,
        )
        TutorApplication.objects.create(
            profile=profile,
            school_id=SimpleUploadedFile(
                f"{username}-id.jpg", b"id", content_type="image/jpeg"
            ),
            enrollment_proof=SimpleUploadedFile(
                f"{username}-rf.pdf", b"rf", content_type="application/pdf"
            ),
            application_status="approved",
        )
        return tutor

    def add_slots(self, tutor, slot_times):
        return [
            TutorAvailability.objects.create(
                tutor=tutor,
                day="Mon",
                time_slot=slot_time,
                is_active=True,
            )
            for slot_time in slot_times
        ]

    def recommend(self, **overrides):
        payload = {
            "subject": self.subject.subject_code,
            "preferred_mode": "Online",
            "min_budget": 100,
            "max_budget": 300,
            "date": self.search_date.isoformat(),
            "start_time": "14:00",
            "end_time": "15:00",
        }
        payload.update(overrides)
        return self.client.post("/api/recommend-tutors/", payload, format="json")

    def response_ids(self, response):
        return {item["id"] for item in response.data["tutors"]}

    def test_filters_by_subject_mode_and_budget(self):
        matching = self.create_tutor("matching")
        wrong_subject = self.create_tutor("science", subject=self.other_subject)
        f2f_only = self.create_tutor("f2f", can_online=False, can_f2f=True)
        expensive = self.create_tutor("expensive", hourly_rate=500)

        for tutor in [matching, wrong_subject, f2f_only, expensive]:
            self.add_slots(tutor, [time(14, 0), time(14, 30)])

        response = self.recommend()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {matching.profile_id})

    def test_multislot_search_requires_every_slot(self):
        complete = self.create_tutor("complete")
        incomplete = self.create_tutor("incomplete")
        self.add_slots(complete, [time(14, 0), time(15, 0)])
        self.add_slots(incomplete, [time(14, 0)])

        response = self.recommend(end_time="16:00")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {complete.profile_id})

    def test_falls_back_to_subject_matches_when_exact_availability_is_empty(self):
        matching = self.create_tutor("fallback")
        wrong_subject = self.create_tutor("wrongsubject", subject=self.other_subject)
        f2f_only = self.create_tutor("f2ffallback", can_online=False, can_f2f=True)
        expensive = self.create_tutor("fallbackexpensive", hourly_rate=500)

        response = self.recommend()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {matching.profile_id})

    def test_partially_booked_tutor_is_offered_for_other_times_that_day(self):
        # The requested window is taken but the tutor still has 14:30 free, so they belong in the
        # date-only fallback. A tutor with no slots that weekday at all does not.
        no_slots_that_day = self.create_tutor("fallbacksafe")
        partially_booked = self.create_tutor("fallbackbooked")
        booked_slots = self.add_slots(partially_booked, [time(14, 0), time(14, 30)])
        Booking.objects.create(
            student=self.student_profile,
            tutor=partially_booked,
            availability=booked_slots[0],
            session_date=self.search_date,
            session_mode="Online",
            status="Confirmed",
        )

        response = self.recommend()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["match_stage"], "date_only")
        self.assertEqual(self.response_ids(response), {partially_booked.profile_id})
        self.assertNotIn(no_slots_that_day.profile_id, self.response_ids(response))

    def test_fully_booked_tutor_is_excluded_from_any_time_search(self):
        free = self.create_tutor("stillfree")
        fully_booked = self.create_tutor("solidlybooked")
        self.add_slots(free, [time(14, 0)])
        booked_slots = self.add_slots(fully_booked, [time(14, 0)])
        Booking.objects.create(
            student=self.student_profile,
            tutor=fully_booked,
            availability=booked_slots[0],
            session_date=self.search_date,
            session_mode="Online",
            status="Confirmed",
        )

        # No time range at all -- the old code applied booking exclusions only when a range was
        # given, so a tutor with nothing left still surfaced.
        response = self.recommend(start_time=None, end_time=None)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {free.profile_id})

    def test_mid_payment_booking_blocks_the_slot(self):
        free = self.create_tutor("paymentfree")
        mid_payment = self.create_tutor("paymentpending")
        self.add_slots(free, [time(14, 0)])
        booked_slots = self.add_slots(mid_payment, [time(14, 0)])
        Booking.objects.create(
            student=self.student_profile,
            tutor=mid_payment,
            availability=booked_slots[0],
            session_date=self.search_date,
            session_mode="Online",
            status="Awaiting Payment Verification",
        )

        response = self.recommend(start_time=None, end_time=None)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {free.profile_id})

    def test_booked_slot_excludes_tutor(self):
        available = self.create_tutor("available")
        booked = self.create_tutor("booked")
        self.add_slots(available, [time(14, 0)])
        booked_slots = self.add_slots(booked, [time(14, 0)])
        Booking.objects.create(
            student=self.student_profile,
            tutor=booked,
            availability=booked_slots[0],
            session_date=self.search_date,
            session_mode="Online",
            status="Confirmed",
        )

        response = self.recommend()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {available.profile_id})

    def test_availability_overrides_exclude_tutors(self):
        available = self.create_tutor("plain")
        full_day_blocked = self.create_tutor("fullday")
        slot_blocked = self.create_tutor("slot")
        self.add_slots(available, [time(14, 0), time(14, 30)])
        self.add_slots(full_day_blocked, [time(14, 0), time(14, 30)])
        slot_blocked_slots = self.add_slots(slot_blocked, [time(14, 0), time(14, 30)])
        TutorAvailabilityOverride.objects.create(
            tutor=full_day_blocked,
            override_date=self.search_date,
            is_full_day=True,
        )
        TutorAvailabilityOverride.objects.create(
            tutor=slot_blocked,
            override_date=self.search_date,
            availability=slot_blocked_slots[0],
            is_full_day=False,
        )

        response = self.recommend()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {available.profile_id})

    # A date is required. It used to be optional, and an unreadable one parsed to None and was
    # treated as "no date given" -- silently dropping every date filter and exclusion and returning
    # the widest possible result set with no signal to the caller.

    def test_missing_date_is_rejected(self):
        self.create_tutor("nodate")

        response = self.recommend(date=None, start_time=None, end_time=None)

        self.assertEqual(response.status_code, 400)

    def test_unparseable_date_is_rejected(self):
        self.create_tutor("baddate")

        response = self.recommend(date="tomorrow", start_time=None, end_time=None)

        self.assertEqual(response.status_code, 400)

    def test_past_date_is_rejected(self):
        self.create_tutor("pastdate")
        yesterday = (timezone.localdate() - timedelta(days=1)).isoformat()

        response = self.recommend(date=yesterday, start_time=None, end_time=None)

        self.assertEqual(response.status_code, 400)

    def test_date_beyond_booking_horizon_is_rejected(self):
        self.create_tutor("farfuture")
        too_far = (timezone.localdate() + timedelta(days=30)).isoformat()

        response = self.recommend(date=too_far, start_time=None, end_time=None)

        self.assertEqual(response.status_code, 400)

    def test_response_reports_match_stage_and_card_fields(self):
        tutor = self.create_tutor("carddata")
        self.add_slots(tutor, [time(14, 0)])

        response = self.recommend()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["match_stage"], "exact")
        row = response.data["tutors"][0]
        self.assertEqual(row["available_days"], ["Mon"])
        self.assertEqual(row["total_sessions"], 0)

    def test_subject_only_stage_keeps_tutors_booked_on_that_date(self):
        # The date has been dropped entirely, so filtering by it would be incoherent -- a tutor
        # booked solid on the searched date may be the right choice for another one.
        booked = self.create_tutor("solidbooked")
        booked_slots = self.add_slots(booked, [time(14, 0)])
        Booking.objects.create(
            student=self.student_profile,
            tutor=booked,
            availability=booked_slots[0],
            session_date=self.search_date,
            session_mode="Online",
            status="Confirmed",
        )

        response = self.recommend(start_time=None, end_time=None)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["match_stage"], "subject_only")
        self.assertEqual(self.response_ids(response), {booked.profile_id})
        self.assertEqual(response.data["tutors"][0]["available_days"], ["Mon"])

    # The time range is optional on the booking search UI -- a date with no times means "any time
    # that day" and must land on the date-only candidate stage, not the broad subject-only
    # fallback. The three tests below pin that contract.

    def test_date_without_times_returns_only_tutors_free_that_weekday(self):
        available = self.create_tutor("anytime")
        no_slots_that_day = self.create_tutor("otherday")
        self.add_slots(available, [time(9, 0)])
        TutorAvailability.objects.create(
            tutor=no_slots_that_day,
            day="Tue",
            time_slot=time(9, 0),
            is_active=True,
        )

        response = self.recommend(start_time=None, end_time=None)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {available.profile_id})

    def test_date_without_times_still_excludes_full_day_overrides(self):
        available = self.create_tutor("openday")
        blocked = self.create_tutor("blockedday")
        self.add_slots(available, [time(9, 0)])
        self.add_slots(blocked, [time(9, 0)])
        TutorAvailabilityOverride.objects.create(
            tutor=blocked,
            override_date=self.search_date,
            is_full_day=True,
        )

        response = self.recommend(start_time=None, end_time=None)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.response_ids(response), {available.profile_id})

    def test_date_without_times_keeps_partially_booked_tutors(self):
        partially_booked = self.create_tutor("partial")
        slots = self.add_slots(partially_booked, [time(9, 0), time(14, 0)])
        Booking.objects.create(
            student=self.student_profile,
            tutor=partially_booked,
            availability=slots[0],
            session_date=self.search_date,
            session_mode="Online",
            status="Confirmed",
        )

        response = self.recommend(start_time=None, end_time=None)

        self.assertEqual(response.status_code, 200)
        # Their 14:00 slot is still genuinely free, so a whole-day search must surface them --
        # slot-level booking exclusions need a requested slot range to filter against.
        self.assertEqual(self.response_ids(response), {partially_booked.profile_id})


class ChatFeatureTests(APITestCase):
    def setUp(self):
        self.tutee_user = User.objects.create_user(
            username="chat-tutee",
            email="chat-tutee@example.com",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Chat",
            mname="",
            lname="Tutee",
            role="Tutee",
        )
        self.tutor_user = User.objects.create_user(
            username="chat-tutor",
            email="chat-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Chat",
            mname="",
            lname="Tutor",
            role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=250,
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.subject = Subjects.objects.create(
            subject_code="ETH101",
            subject_name="Ethics",
            department="Humanities",
        )
        TutorSubjects.objects.create(
            tutor=self.tutor,
            subject=self.subject,
            expertise_level=5,
        )
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=time(14, 0),
            is_active=True,
        )
        self.booking_request_id = uuid4()
        self.booking = Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 5, 18),
            session_mode="F2F",
            preferred_location="Library",
            booking_request_id=self.booking_request_id,
            status="Pending",
        )

    def make_booking(self, status, session_mode='F2F', preferred_location='Library',
                     session_date=None, days_ago=0):
        from datetime import date, timedelta
        from uuid import uuid4
        # Cancel the setUp booking so it doesn't interfere with this test's query
        self.booking.status = 'Cancelled'
        self.booking.session_date = date.today() - timedelta(days=8)
        self.booking.save(update_fields=['status', 'session_date'])
        d = session_date or (date.today() - timedelta(days=days_ago))
        return Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=d,
            session_mode=session_mode,
            preferred_location=preferred_location,
            booking_request_id=uuid4(),
            status=status,
        )

    def make_extra_tutor_room(self, username, updated_at=None):
        user = User.objects.create_user(
            username=username,
            email=f"{username}@example.com",
            password="password",
        )
        profile = UserProfile.objects.create(
            user=user,
            fname=username.title(),
            mname="",
            lname="Tutor",
            role="Tutor",
        )
        tutor = Tutor.objects.create(
            profile=profile,
            hourly_rate=250,
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        availability = TutorAvailability.objects.create(
            tutor=tutor,
            day="Mon",
            time_slot=time(14, 0),
            is_active=True,
        )
        room = ChatRoom.objects.create(
            tutee=self.tutee_profile,
            tutor=profile,
            **({'updated_at': updated_at} if updated_at else {}),
        )
        return user, profile, tutor, availability, room

    def test_status_intent_pending_f2f_returns_pending_location(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        booking = self.make_booking('Pending', session_mode='F2F')
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'pending_location')

    def test_status_intent_pending_online_returns_pending(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        booking = self.make_booking('Pending', session_mode='Online')
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'pending')

    def test_status_intent_confirmed_returns_confirmed(self):
        from datetime import date, timedelta
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.make_booking('Confirmed', session_date=date.today() + timedelta(days=1))
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'confirmed')

    def test_confirmed_future_session_returns_upcoming(self):
        from datetime import date, timedelta
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        future_date = date.today() + timedelta(days=2)
        self.make_booking('Confirmed', session_date=future_date)
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'confirmed')
        self.assertEqual(context['status'], 'Upcoming')

    def test_confirmed_active_session_returns_ongoing(self):
        from django.utils import timezone
        from datetime import timedelta
        from uuid import uuid4
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        now_local = timezone.localtime(timezone.now())
        active_availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=(now_local - timedelta(minutes=10)).time(),
            is_active=True,
        )
        self.booking.status = 'Cancelled'
        self.booking.save()

        Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=active_availability,
            session_date=now_local.date(),
            session_mode='Online',
            booking_request_id=uuid4(),
            status='Confirmed',
        )
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'ongoing')
        self.assertEqual(context['status'], 'Ongoing')

    def test_confirmed_past_session_returns_payment_required(self):
        from datetime import date, timedelta
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        past_date = date.today() - timedelta(days=1)
        self.make_booking('Confirmed', session_date=past_date)
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'payment_required')
        self.assertEqual(context['status'], 'Payment Required')

    def test_status_intent_awaiting_payment_returns_awaiting_payment(self):
        from datetime import date, timedelta
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.make_booking('Awaiting Payment Verification', session_date=date.today() + timedelta(days=1))
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'awaiting_payment')

    def test_status_intent_completed_unrated_returns_review_pending(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.make_booking('Completed', days_ago=1)
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'review_pending')

    def test_status_intent_completed_rated_returns_none(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        booking = self.make_booking('Completed', days_ago=1)
        Rating.objects.create(
            booking=booking,
            student=self.tutee_profile,
            tutor=self.tutor,
            rating_score=5,
        )
        context = get_current_booking_context(room)
        self.assertIsNone(context)

    def test_status_intent_rejected_recent_returns_rejected(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.make_booking('Rejected', days_ago=2)
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'rejected')

    def test_status_intent_cancelled_recent_returns_cancelled(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.make_booking('Cancelled', days_ago=3)
        context = get_current_booking_context(room)
        self.assertIsNotNone(context)
        self.assertEqual(context['status_intent'], 'cancelled')

    def test_status_intent_terminal_older_than_7_days_returns_none(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.make_booking('Rejected', days_ago=8)
        context = get_current_booking_context(room)
        self.assertIsNone(context)

    def test_chat_rooms_endpoint_returns_paginated_payload(self):
        from django.utils import timezone
        from datetime import timedelta

        base_time = timezone.now()
        rooms = []
        for index in range(4):
            *_, room = self.make_extra_tutor_room(
                f"page-tutor-{index}",
                updated_at=base_time - timedelta(minutes=index),
            )
            rooms.append(room)

        self.client.force_authenticate(user=self.tutee_user)
        response = self.client.get("/api/chat/rooms/?page_size=2")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["results"]), 2)
        self.assertTrue(response.data["has_more"])
        self.assertIsNotNone(response.data["next_cursor"])
        self.assertEqual(response.data["total_unread"], 0)
        self.assertEqual(
            [room["id"] for room in response.data["results"]],
            [rooms[0].id, rooms[1].id],
        )

    def test_chat_rooms_cursor_returns_next_page_without_duplicates(self):
        from django.utils import timezone
        from datetime import timedelta

        base_time = timezone.now()
        for index in range(4):
            self.make_extra_tutor_room(
                f"cursor-tutor-{index}",
                updated_at=base_time - timedelta(minutes=index),
            )

        self.client.force_authenticate(user=self.tutee_user)
        first_page = self.client.get("/api/chat/rooms/?page_size=2")
        second_page = self.client.get(
            f"/api/chat/rooms/?page_size=2&cursor={first_page.data['next_cursor']}"
        )

        first_ids = {room["id"] for room in first_page.data["results"]}
        second_ids = {room["id"] for room in second_page.data["results"]}
        self.assertEqual(first_page.status_code, 200)
        self.assertEqual(second_page.status_code, 200)
        self.assertEqual(len(second_ids), 2)
        self.assertFalse(first_ids & second_ids)
        self.assertFalse(second_page.data["has_more"])

    def test_chat_rooms_total_unread_includes_rooms_outside_page(self):
        from django.utils import timezone
        from datetime import timedelta

        base_time = timezone.now()
        for index in range(3):
            tutor_user, *_, room = self.make_extra_tutor_room(
                f"unread-tutor-{index}",
                updated_at=base_time - timedelta(minutes=index),
            )
            Message.objects.create(
                room=room,
                sender=tutor_user,
                content=f"Unread {index}",
                is_read=False,
            )

        self.client.force_authenticate(user=self.tutee_user)
        response = self.client.get("/api/chat/rooms/?page_size=1")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["results"]), 1)
        self.assertEqual(response.data["total_unread"], 3)

    def test_batched_current_booking_context_matches_single_room_helper(self):
        from datetime import timedelta
        from uuid import uuid4

        today = date.today()
        room_pending = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)

        _, confirmed_profile, confirmed_tutor, confirmed_availability, room_confirmed = (
            self.make_extra_tutor_room("bulk-confirmed")
        )
        Booking.objects.create(
            student=self.tutee_profile,
            tutor=confirmed_tutor,
            availability=confirmed_availability,
            session_date=today + timedelta(days=2),
            session_mode="Online",
            booking_request_id=uuid4(),
            status="Confirmed",
        )

        _, completed_profile, completed_tutor, completed_availability, room_completed = (
            self.make_extra_tutor_room("bulk-completed")
        )
        Booking.objects.create(
            student=self.tutee_profile,
            tutor=completed_tutor,
            availability=completed_availability,
            session_date=today - timedelta(days=1),
            session_mode="Online",
            booking_request_id=uuid4(),
            status="Completed",
        )

        _, terminal_profile, terminal_tutor, terminal_availability, room_terminal = (
            self.make_extra_tutor_room("bulk-terminal")
        )
        Booking.objects.create(
            student=self.tutee_profile,
            tutor=terminal_tutor,
            availability=terminal_availability,
            session_date=today - timedelta(days=2),
            session_mode="Online",
            booking_request_id=uuid4(),
            status="Cancelled",
        )

        *_, room_none = self.make_extra_tutor_room("bulk-none")
        rooms = [room_pending, room_confirmed, room_completed, room_terminal, room_none]
        batched_contexts = get_current_booking_contexts(rooms)

        for room in rooms:
            self.assertEqual(batched_contexts[room.id], get_current_booking_context(room))

    def test_chat_room_list_query_count_does_not_scale_per_room(self):
        from django.utils import timezone
        from datetime import timedelta
        from uuid import uuid4

        base_time = timezone.now()
        for index in range(6):
            _, _, tutor, availability, room = self.make_extra_tutor_room(
                f"query-tutor-{index}",
                updated_at=base_time - timedelta(minutes=index),
            )
            Booking.objects.create(
                student=self.tutee_profile,
                tutor=tutor,
                availability=availability,
                session_date=date.today() + timedelta(days=index + 1),
                session_mode="Online",
                booking_request_id=uuid4(),
                status="Pending",
            )

        self.client.force_authenticate(user=self.tutee_user)
        with CaptureQueriesContext(connection) as two_room_queries:
            two_room_response = self.client.get("/api/chat/rooms/?page_size=2")
        with CaptureQueriesContext(connection) as six_room_queries:
            six_room_response = self.client.get("/api/chat/rooms/?page_size=6")

        self.assertEqual(two_room_response.status_code, 200)
        self.assertEqual(six_room_response.status_code, 200)
        self.assertLessEqual(len(six_room_queries), len(two_room_queries) + 3)

    def test_partner_context_counts_completed_session_groups_and_hours(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        second_slot = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=time(15, 0),
            is_active=True,
        )
        first_group_id = uuid4()
        self.booking.status = 'Cancelled'
        self.booking.save(update_fields=['status'])
        Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 5, 19),
            session_mode="Online",
            session_group_id=first_group_id,
            status="Completed",
        )
        Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=second_slot,
            session_date=date(2026, 5, 19),
            session_mode="Online",
            session_group_id=first_group_id,
            status="Completed",
        )
        Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 5, 20),
            session_mode="Online",
            status="Completed",
        )

        context = get_partner_context(room, self.tutee_user)

        self.assertEqual(context['sessions_together'], 2)
        self.assertEqual(context['focused_hours'], 3.0)

    def test_partner_context_returns_tutor_topics(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        context = get_partner_context(room, self.tutee_user)

        self.assertEqual(context['topics'], ['Ethics'])

    def test_partner_context_skipped_for_support_rooms(self):
        room = ChatRoom.objects.create(
            tutee=self.tutee_profile,
            tutor=None,
            room_type='support',
        )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get(f"/api/chat/rooms/{room.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.data['partner_context'])

    def test_ticket_context_exposes_subject_category_and_status(self):
        room = ChatRoom.objects.create(
            tutee=self.tutee_profile,
            tutor=None,
            room_type='support',
        )
        SupportTicket.objects.create(
            user=self.tutee_profile,
            chatroom=room,
            category='Technical',
            subject='Cannot join session',
            description='Video call will not load.',
            status='Open',
        )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get(f"/api/chat/rooms/{room.id}/")

        self.assertEqual(response.status_code, 200)
        ticket_context = response.data['ticket_context']
        self.assertEqual(ticket_context['subject'], 'Cannot join session')
        self.assertEqual(ticket_context['category'], 'Technical Problem')
        self.assertEqual(ticket_context['status'], 'Open')
        self.assertIsNotNone(ticket_context['created_at'])
        self.assertIsNone(ticket_context['assigned_agent_name'])

    def test_support_room_read_endpoint_handles_unassigned_ticket_broadcast(self):
        room = ChatRoom.objects.create(
            tutee=self.tutee_profile,
            tutor=None,
            room_type='support',
        )
        SupportTicket.objects.create(
            user=self.tutee_profile,
            chatroom=room,
            category='Dispute',
            subject='Need help with tutor',
            description='The conversation needs moderation.',
            status='Open',
        )
        message = Message.objects.create(
            room=room,
            sender=None,
            content='Ticket created.',
        )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.post(f"/api/chat/rooms/{room.id}/read/")

        message.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(message.is_read)
        self.assertIsNotNone(message.read_at)

    def test_room_read_endpoint_marks_other_user_messages_read(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        message = Message.objects.create(
            room=room,
            sender=self.tutor_user,
            content="Please confirm the location.",
        )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.post(f"/api/chat/rooms/{room.id}/read/")

        message.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(message.is_read)
        self.assertIsNotNone(message.read_at)

    def test_message_history_returns_latest_messages_in_display_order(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        for index in range(55):
            Message.objects.create(
                room=room,
                sender=self.tutee_user if index % 2 == 0 else self.tutor_user,
                content=f"Message {index}",
            )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get(f"/api/chat/rooms/{room.id}/history/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 50)
        self.assertEqual(response.data[0]["content"], "Message 5")
        self.assertEqual(response.data[-1]["content"], "Message 54")

    def test_message_history_after_id_returns_newer_messages(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        first = Message.objects.create(room=room, sender=self.tutee_user, content="First")
        second = Message.objects.create(room=room, sender=self.tutor_user, content="Second")
        third = Message.objects.create(room=room, sender=self.tutee_user, content="Third")
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get(f"/api/chat/rooms/{room.id}/history/?after_id={first.id}")

        self.assertEqual(response.status_code, 200)
        self.assertEqual([message["id"] for message in response.data], [second.id, third.id])

    def test_non_member_cannot_load_message_history(self):
        other_user = User.objects.create_user(
            username="chat-history-other",
            email="chat-history-other@example.com",
            password="password",
        )
        UserProfile.objects.create(
            user=other_user,
            fname="History",
            mname="",
            lname="Other",
            role="Tutee",
        )
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.client.force_authenticate(user=other_user)

        response = self.client.get(f"/api/chat/rooms/{room.id}/history/")

        self.assertEqual(response.status_code, 403)

    def test_message_history_serializes_system_message_without_sender(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        Message.objects.create(
            room=room,
            sender=None,
            content="System note",
            message_type="system",
        )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get(f"/api/chat/rooms/{room.id}/history/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data[0]["sender_name"], "System")
        self.assertIsNone(response.data[0]["sender_profile_id"])

    def make_editable_f2f_booking(self, days_ahead=3, status="Confirmed"):
        """A live face-to-face booking comfortably outside the Grace Cutoff."""
        self.booking.status = "Cancelled"
        self.booking.save(update_fields=["status"])

        return Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=timezone.localdate() + timedelta(days=days_ahead),
            session_mode="F2F",
            preferred_location="Library",
            booking_request_id=uuid4(),
            status=status,
        )

    def test_location_update_posts_booking_event_to_canonical_room(self):
        booking = self.make_editable_f2f_booking()
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.patch(
            f"/api/bookings/{booking.booking_request_id}/location/",
            {"preferred_location": "Study Hall 2"},
            format="json",
        )

        booking.refresh_from_db()
        room = ChatRoom.objects.get(tutee=self.tutee_profile, tutor=self.tutor_profile)
        message = room.messages.latest("created_at")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(booking.preferred_location, "Study Hall 2")
        self.assertEqual(message.message_type, "booking_event")
        self.assertEqual(message.metadata["event_type"], "location_updated")
        self.assertEqual(message.metadata["booking"]["preferred_location"], "Study Hall 2")

    def test_location_update_allowed_while_awaiting_payment_verification(self):
        booking = self.make_editable_f2f_booking(status="Awaiting Payment Verification")
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.patch(
            f"/api/bookings/{booking.booking_request_id}/location/",
            {"preferred_location": "Room 204"},
            format="json",
        )

        booking.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(booking.preferred_location, "Room 204")

    def test_location_update_rejected_inside_grace_cutoff(self):
        # The session is today at 14:00 and the cutoff is 12 hours out, so it has already passed.
        # This is also the is_born_late case: such bookings are frozen from creation.
        booking = self.make_editable_f2f_booking(days_ahead=0)
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.patch(
            f"/api/bookings/{booking.booking_request_id}/location/",
            {"preferred_location": "Too Late Hall"},
            format="json",
        )

        booking.refresh_from_db()
        self.assertEqual(response.status_code, 400)
        self.assertEqual(booking.preferred_location, "Library")

    def test_booking_context_exposes_can_edit_location(self):
        booking = self.make_editable_f2f_booking()

        context = serialize_booking_context(booking)

        self.assertTrue(context["can_edit_location"])

    def test_tutee_can_send_message_via_rest(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.post(
            f"/api/chat/rooms/{room.id}/messages/",
            {"message": "Hello!", "temp_id": "temp-1"},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Message.objects.filter(room=room, message_type="text").count(), 1)
        self.assertEqual(response.data["message"]["content"], "Hello!")
        self.assertEqual(response.data["message"]["temp_id"], "temp-1")
        self.assertTrue(response.data["message"]["is_me"])
        self.assertEqual(response.data["room"]["id"], room.id)
        self.assertEqual(response.data["room"]["last_message"]["content"], "Hello!")

    def test_tutor_can_send_message_via_rest(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            f"/api/chat/rooms/{room.id}/messages/",
            {"message": "Hi!"},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Message.objects.filter(room=room, message_type="text").count(), 1)

    def test_non_member_cannot_send_message_via_rest(self):
        other_user = User.objects.create_user(
            username="chat-other",
            email="chat-other@example.com",
            password="password",
        )
        UserProfile.objects.create(
            user=other_user,
            fname="Chat",
            mname="",
            lname="Other",
            role="Tutee",
        )
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.client.force_authenticate(user=other_user)

        response = self.client.post(
            f"/api/chat/rooms/{room.id}/messages/",
            {"message": "Nope"},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(Message.objects.filter(room=room, message_type="text").count(), 0)

    def test_empty_rest_message_returns_400(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.post(
            f"/api/chat/rooms/{room.id}/messages/",
            {"message": "   "},
            format="json",
        )

        self.assertEqual(response.status_code, 400)

    def test_unauthenticated_rest_message_returns_401(self):
        room = ChatRoom.objects.create(tutee=self.tutee_profile, tutor=self.tutor_profile)

        response = self.client.post(
            f"/api/chat/rooms/{room.id}/messages/",
            {"message": "Hello!"},
            format="json",
        )

        self.assertEqual(response.status_code, 401)


class OnlinePaymentInitiationTests(APITestCase):
    def setUp(self):
        self.tutee_user = User.objects.create_user(
            username="payment-tutee",
            email="payment-tutee@example.com",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Payment",
            mname="",
            lname="Tutee",
            role="Tutee",
        )
        self.tutor_user = User.objects.create_user(
            username="payment-tutor",
            email="payment-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Payment",
            mname="",
            lname="Tutor",
            role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("280.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=time(14, 0),
            is_active=True,
        )
        self.booking = Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 4, 12),
            session_mode="Online",
            booking_request_id=uuid4(),
            status="Confirmed",
        )
        self.client.force_authenticate(user=self.tutee_user)

    def paymongo_response(self, status_code, payload):
        response = Mock()
        response.status_code = status_code
        response.json.return_value = payload
        return response

    def success_payload(self, session_id="cs_test_123"):
        return {
            "data": {
                "id": session_id,
                "attributes": {
                    "checkout_url": "https://checkout.paymongo.com/test-session",
                },
            },
        }

    def retrieve_payload(self, status_value="paid"):
        return {
            "data": {
                "id": "cs_test_123",
                "attributes": {
                    "status": status_value,
                },
            },
        }

    def create_pending_paymongo_payment(self):
        method = PaymentMethod.objects.create(
            code="PAYMONGO",
            method_name="Pay Online (GCash / Card)",
            is_active=True,
        )
        return Payment.objects.create(
            booking=self.booking,
            amount=Decimal("140.00"),
            method=method,
            payment_status="Pending",
            transaction_reference="cs_test_123",
        )

    def create_second_request_slot(self):
        second_availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=time(15, 0),
            is_active=True,
        )
        return Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=second_availability,
            session_date=self.booking.session_date,
            session_mode="Online",
            session_group_id=self.booking.session_group_id,
            booking_request_id=self.booking.booking_request_id,
            status=self.booking.status,
        )

    @patch("studybuddy.views.requests.post")
    def test_initiate_online_payment_accepts_paymongo_200_success(self, mock_post):
        mock_post.return_value = self.paymongo_response(200, self.success_payload())

        response = self.client.post(
            "/api/payments/initiate/",
            {"booking_id": self.booking.id},
            format="json",
        )

        payment = Payment.objects.get(booking=self.booking)
        payload = mock_post.call_args.kwargs["json"]
        attributes = payload["data"]["attributes"]

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.data["payment_url"],
            "https://checkout.paymongo.com/test-session",
        )
        self.assertEqual(payment.amount, Decimal("280.00"))
        self.assertEqual(payment.method.code, "PAYMONGO")
        self.assertEqual(payment.payment_status, "Pending")
        self.assertEqual(payment.transaction_reference, "cs_test_123")
        self.assertIn(f"SB-BK-{self.booking.id}", attributes["description"])
        self.assertEqual(
            attributes["line_items"][0]["name"],
            f"StudyBuddy SB-BK-{self.booking.id}",
        )

    @patch("studybuddy.views.requests.post")
    def test_initiate_online_payment_accepts_paymongo_201_success(self, mock_post):
        mock_post.return_value = self.paymongo_response(201, self.success_payload("cs_test_201"))

        response = self.client.post(
            "/api/payments/initiate/",
            {"booking_id": self.booking.id},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.data["payment_url"],
            "https://checkout.paymongo.com/test-session",
        )
        self.assertEqual(
            Payment.objects.get(booking=self.booking).transaction_reference,
            "cs_test_201",
        )

    @patch("studybuddy.views.requests.post")
    def test_initiate_online_payment_returns_paymongo_validation_error(self, mock_post):
        mock_post.return_value = self.paymongo_response(
            400,
            {"errors": [{"code": "parameter_invalid", "detail": "Invalid payment method type."}]},
        )

        response = self.client.post(
            "/api/payments/initiate/",
            {"booking_id": self.booking.id},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["error"], "Invalid payment method type.")
        self.assertFalse(Payment.objects.filter(booking=self.booking).exists())

    @patch("studybuddy.views.requests.post")
    def test_initiate_online_payment_returns_paymongo_auth_error(self, mock_post):
        mock_post.return_value = self.paymongo_response(
            401,
            {"errors": [{"code": "secret_key_invalid", "detail": "Invalid API key."}]},
        )

        response = self.client.post(
            "/api/payments/initiate/",
            {"booking_id": self.booking.id},
            format="json",
        )

        self.assertEqual(response.status_code, 502)
        self.assertEqual(
            response.data["error"],
            "Payment provider authentication failed. Check the PayMongo secret key.",
        )
        self.assertFalse(Payment.objects.filter(booking=self.booking).exists())

    @patch("studybuddy.views.requests.post")
    def test_initiate_online_payment_requires_checkout_url_on_success(self, mock_post):
        mock_post.return_value = self.paymongo_response(
            200,
            {"data": {"id": "cs_missing_url", "attributes": {}}},
        )

        response = self.client.post(
            "/api/payments/initiate/",
            {"booking_id": self.booking.id},
            format="json",
        )

        self.assertEqual(response.status_code, 502)
        self.assertEqual(response.data["error"], "Payment provider did not return a checkout URL.")
        self.assertFalse(Payment.objects.filter(booking=self.booking).exists())

    @patch("studybuddy.views.requests.post")
    def test_initiate_online_payment_returns_502_on_network_error(self, mock_post):
        import requests

        mock_post.side_effect = requests.RequestException("connection reset")

        response = self.client.post(
            "/api/payments/initiate/",
            {"booking_id": self.booking.id},
            format="json",
        )

        self.assertEqual(response.status_code, 502)
        self.assertEqual(
            response.data["error"],
            "Could not reach the payment provider. Please try again.",
        )
        self.assertFalse(Payment.objects.filter(booking=self.booking).exists())

    @patch("studybuddy.views.requests.get")
    def test_verify_online_payment_marks_paymongo_payment_paid(self, mock_get):
        payment = self.create_pending_paymongo_payment()
        mock_get.return_value = self.paymongo_response(200, self.retrieve_payload("paid"))

        response = self.client.post(f"/api/bookings/{self.booking.id}/verify-online-payment/")

        payment.refresh_from_db()
        self.booking.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["session"]["status"], "Awaiting Verification")
        self.assertEqual(payment.payment_status, "Paid")
        self.assertIsNotNone(payment.paid_at)
        self.assertEqual(self.booking.status, "Awaiting Payment Verification")
        self.assertTrue(self.booking.tutee_confirmed)

    @patch("studybuddy.views.requests.get")
    def test_verify_online_payment_returns_502_on_network_error(self, mock_get):
        import requests

        payment = self.create_pending_paymongo_payment()
        mock_get.side_effect = requests.RequestException("gateway timeout")

        response = self.client.post(f"/api/bookings/{self.booking.id}/verify-online-payment/")

        payment.refresh_from_db()
        self.assertEqual(response.status_code, 502)
        self.assertEqual(
            response.data["error"],
            "Could not reach the payment provider. Please try again.",
        )
        self.assertEqual(payment.payment_status, "Pending")

    @patch("studybuddy.views.requests.get")
    def test_verify_online_payment_updates_all_booking_request_slots(self, mock_get):
        second_booking = self.create_second_request_slot()
        self.create_pending_paymongo_payment()
        mock_get.return_value = self.paymongo_response(200, self.retrieve_payload("paid"))

        response = self.client.post(f"/api/bookings/{self.booking.id}/verify-online-payment/")

        self.booking.refresh_from_db()
        second_booking.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["session"]["status"], "Awaiting Verification")
        self.assertEqual(self.booking.status, "Awaiting Payment Verification")
        self.assertEqual(second_booking.status, "Awaiting Payment Verification")
        self.assertTrue(self.booking.tutee_confirmed)
        self.assertTrue(second_booking.tutee_confirmed)

    @patch("studybuddy.views.requests.get")
    @patch("studybuddy.views.requests.post")
    def test_verify_payment_returns_all_booking_request_slots(self, mock_post, mock_get):
        """Response must include all booking_request_id siblings, not just session_group siblings."""
        second_booking = self.create_second_request_slot()
        self.create_pending_paymongo_payment()

        mock_get.return_value = self.paymongo_response(200, self.retrieve_payload("paid"))

        response = self.client.post(
            f"/api/bookings/{self.booking.id}/verify-online-payment/",
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        # Both slots should be reflected in the response. With 2 x 60-min slots the
        # duration_hours must be 2.0; if the bug is present (get_session_group_bookings
        # is used and session_group_id is None) only one slot is included → 1.0.
        self.assertEqual(response.data["session"]["duration_hours"], 2.0)
        _ = second_booking  # referenced to avoid unused-variable warning

    def test_manual_payment_submission_updates_all_session_group_slots(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        # CASH is only valid for F2F sessions now that payment method is derived from
        # session_mode; this test's booking defaults to Online, so both slots are flipped
        # to F2F to keep exercising the CASH path.
        self.booking.session_mode = "F2F"
        self.booking.save(update_fields=["session_mode"])
        second_booking = self.create_second_request_slot()
        second_booking.session_mode = "F2F"
        second_booking.save(update_fields=["session_mode"])
        method = PaymentMethod.objects.create(
            code="CASH",
            method_name="Cash",
            is_active=True,
        )
        receipt = SimpleUploadedFile("receipt.jpg", b"receipt-bytes", content_type="image/jpeg")

        response = self.client.post(
            f"/api/bookings/{self.booking.id}/submit-payment/",
            {"payment_method": method.method_id, "receipt_image": receipt},
            format="multipart",
        )

        self.booking.refresh_from_db()
        second_booking.refresh_from_db()

        self.assertEqual(response.status_code, 201)
        self.assertEqual(self.booking.status, "Awaiting Payment Verification")
        self.assertEqual(second_booking.status, "Awaiting Payment Verification")
        self.assertTrue(self.booking.tutee_confirmed)
        self.assertTrue(second_booking.tutee_confirmed)

    def test_tutor_completion_updates_all_session_group_slots(self):
        second_booking = self.create_second_request_slot()
        self.booking.status = "Awaiting Payment Verification"
        self.booking.tutee_confirmed = True
        self.booking.save(update_fields=["status", "tutee_confirmed"])
        second_booking.status = "Awaiting Payment Verification"
        second_booking.tutee_confirmed = True
        second_booking.save(update_fields=["status", "tutee_confirmed"])
        method = PaymentMethod.objects.create(
            code="CASH",
            method_name="Cash",
            is_active=True,
        )
        Payment.objects.create(
            booking=self.booking,
            amount=Decimal("280.00"),
            method=method,
            payment_status="Pending",
        )
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(f"/api/bookings/{self.booking.id}/tutor-confirm/")

        self.booking.refresh_from_db()
        second_booking.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.booking.status, "Completed")
        self.assertEqual(second_booking.status, "Completed")
        self.assertTrue(self.booking.tutor_confirmed)
        self.assertTrue(second_booking.tutor_confirmed)

    def test_tutor_completion_credits_wallet_for_paymongo_payment(self):
        self.booking.status = "Awaiting Payment Verification"
        self.booking.tutee_confirmed = True
        self.booking.save(update_fields=["status", "tutee_confirmed"])
        method = PaymentMethod.objects.create(
            code="PAYMONGO",
            method_name="Pay Online (GCash / Card)",
            is_active=True,
        )
        Payment.objects.create(
            booking=self.booking,
            amount=Decimal("280.00"),
            method=method,
            payment_status="Paid",
            transaction_reference="cs_test_wallet",
        )
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(f"/api/bookings/{self.booking.id}/tutor-confirm/")

        wallet = Wallet.objects.get(tutor=self.tutor)
        transaction = Transaction.objects.get(reference_id=f"BK-{self.booking.id}")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(wallet.balance, Decimal("252.00"))
        self.assertEqual(transaction.transaction_type, "session_credit")
        self.assertEqual(transaction.amount, Decimal("252.00"))
        self.assertIn("Student: Payment Tutee", transaction.description)
        self.assertNotIn("Transaction ID", transaction.description)

    def test_wallet_transactions_include_paymongo_transaction_id(self):
        self.booking.status = "Awaiting Payment Verification"
        self.booking.tutee_confirmed = True
        self.booking.save(update_fields=["status", "tutee_confirmed"])
        method = PaymentMethod.objects.create(
            code="PAYMONGO",
            method_name="Pay Online (GCash / Card)",
            is_active=True,
        )
        Payment.objects.create(
            booking=self.booking,
            amount=Decimal("280.00"),
            method=method,
            payment_status="Paid",
            transaction_reference="cs_test_wallet",
        )
        self.client.force_authenticate(user=self.tutor_user)
        self.client.post(f"/api/bookings/{self.booking.id}/tutor-confirm/")
        Transaction.objects.filter(reference_id=f"BK-{self.booking.id}").update(
            description=(
                "Session Credit for 2026-04-12 (Less 10% Platform Fee) "
                "- Transaction ID: cs_test_wallet"
            )
        )

        response = self.client.get("/api/wallet/transactions/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data[0]["reference_id"], f"BK-{self.booking.id}")
        self.assertEqual(response.data[0]["payment_transaction_id"], "cs_test_wallet")
        self.assertEqual(response.data[0]["student_name"], "Payment Tutee")
        self.assertIn("Student: Payment Tutee", response.data[0]["description"])
        self.assertNotIn("Transaction ID", response.data[0]["description"])

    def test_tutor_completion_does_not_duplicate_wallet_credit(self):
        self.booking.status = "Awaiting Payment Verification"
        self.booking.tutee_confirmed = True
        self.booking.save(update_fields=["status", "tutee_confirmed"])
        method = PaymentMethod.objects.create(
            code="PAYMONGO",
            method_name="Pay Online (GCash / Card)",
            is_active=True,
        )
        Payment.objects.create(
            booking=self.booking,
            amount=Decimal("280.00"),
            method=method,
            payment_status="Paid",
            transaction_reference="cs_test_wallet",
        )
        self.client.force_authenticate(user=self.tutor_user)

        first_response = self.client.post(f"/api/bookings/{self.booking.id}/tutor-confirm/")
        second_response = self.client.post(f"/api/bookings/{self.booking.id}/tutor-confirm/")

        wallet = Wallet.objects.get(tutor=self.tutor)

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 400)
        self.assertEqual(wallet.balance, Decimal("252.00"))
        self.assertEqual(Transaction.objects.filter(reference_id=f"BK-{self.booking.id}").count(), 1)

    @patch("studybuddy.views.requests.get")
    def test_verify_online_payment_rejects_incomplete_checkout(self, mock_get):
        payment = self.create_pending_paymongo_payment()
        mock_get.return_value = self.paymongo_response(200, self.retrieve_payload("active"))

        response = self.client.post(f"/api/bookings/{self.booking.id}/verify-online-payment/")

        payment.refresh_from_db()
        self.booking.refresh_from_db()

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["error"], "Online payment has not been completed yet.")
        self.assertEqual(payment.payment_status, "Pending")
        self.assertEqual(self.booking.status, "Confirmed")
        self.assertFalse(self.booking.tutee_confirmed)

    def test_verify_online_payment_requires_paymongo_payment(self):
        response = self.client.post(f"/api/bookings/{self.booking.id}/verify-online-payment/")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["error"], "No PayMongo checkout is pending for this session.")


class PaymentMethodTests(APITestCase):
    def test_payment_methods_hides_legacy_online_methods_when_paymongo_exists(self):
        PaymentMethod.objects.all().delete()

        PaymentMethod.objects.create(
            code="CASH",
            method_name="Cash",
            is_active=True,
        )
        PaymentMethod.objects.create(
            code="ONLINE",
            method_name="Online Payment",
            is_active=True,
        )
        PaymentMethod.objects.create(
            code="online",
            method_name="Online Payment",
            is_active=True,
        )
        PaymentMethod.objects.create(
            code="PAYMONGO",
            method_name="Pay Online (GCash / Card)",
            is_active=True,
        )

        response = self.client.get("/api/payment-methods/")
        codes = [method["code"] for method in response.data]

        self.assertEqual(response.status_code, 200)
        self.assertEqual(codes, ["CASH", "PAYMONGO"])


class SessionModePaymentMethodEnforcementTests(APITestCase):
    """Payment method must match the booking's session_mode (F2F->CASH, Online->PAYMONGO),
    enforced server-side in both submit_session_payment and confirm_payment_and_book,
    independent of whatever the client claims."""

    def setUp(self):
        self.tutee_user = User.objects.create_user(
            username="mode-enforce-tutee",
            email="mode-enforce-tutee@example.com",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user, fname="Mode", mname="", lname="Tutee", role="Tutee",
        )
        self.tutor_user = User.objects.create_user(
            username="mode-enforce-tutor",
            email="mode-enforce-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Mode", mname="", lname="Tutor", role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("280.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.cash_method = PaymentMethod.objects.create(
            code="CASH", method_name="Cash", is_active=True,
        )
        self.paymongo_method = PaymentMethod.objects.create(
            code="PAYMONGO", method_name="Pay Online (GCash / Card)", is_active=True,
        )
        self.client.force_authenticate(user=self.tutee_user)

    def make_confirmed_booking(self, session_mode):
        availability = TutorAvailability.objects.create(
            tutor=self.tutor, day="Mon", time_slot=time(14, 0), is_active=True,
        )
        return Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=availability,
            session_date=date(2026, 4, 12),
            session_mode=session_mode,
            booking_request_id=uuid4(),
            status="Confirmed",
        )

    def test_submit_session_payment_rejects_cash_for_online_session(self):
        booking = self.make_confirmed_booking("Online")

        response = self.client.post(
            f"/api/bookings/{booking.id}/submit-payment/",
            {"payment_method": self.cash_method.method_id},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.data["error"], "Payment method does not match this session's mode."
        )
        self.assertFalse(hasattr(booking, "payment"))

    def test_submit_session_payment_rejects_paymongo_for_f2f_session(self):
        booking = self.make_confirmed_booking("F2F")

        response = self.client.post(
            f"/api/bookings/{booking.id}/submit-payment/",
            {"payment_method": self.paymongo_method.method_id},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.data["error"], "Payment method does not match this session's mode."
        )

    def test_submit_session_payment_cash_requires_receipt_but_not_reference(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        booking = self.make_confirmed_booking("F2F")

        no_receipt_response = self.client.post(
            f"/api/bookings/{booking.id}/submit-payment/",
            {"payment_method": self.cash_method.method_id},
        )
        self.assertEqual(no_receipt_response.status_code, 400)
        self.assertEqual(
            no_receipt_response.data["error"],
            "Receipt image is required for cash payments.",
        )

        receipt = SimpleUploadedFile("receipt.jpg", b"receipt-bytes", content_type="image/jpeg")
        success_response = self.client.post(
            f"/api/bookings/{booking.id}/submit-payment/",
            {"payment_method": self.cash_method.method_id, "receipt_image": receipt},
            format="multipart",
        )

        self.assertEqual(success_response.status_code, 201)
        booking.refresh_from_db()
        self.assertEqual(booking.payment.method.code, "CASH")
        self.assertTrue(bool(booking.payment.receipt_image))
        self.assertIsNone(booking.payment.transaction_reference)

    def test_submit_session_payment_paymongo_still_requires_receipt_and_reference(self):
        booking = self.make_confirmed_booking("Online")

        response = self.client.post(
            f"/api/bookings/{booking.id}/submit-payment/",
            {"payment_method": self.paymongo_method.method_id},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.data["error"], "Receipt image is required for online payments."
        )

    @override_settings(TUTEE_VERIFICATION_ENFORCEMENT_START_DATE=None)
    def test_confirm_payment_and_book_rejects_mismatched_method(self):
        # Not-yet-enforced verification, so the unrelated verification gate in
        # confirm_payment_and_book doesn't 403 before this test's own check is exercised.
        availability = TutorAvailability.objects.create(
            tutor=self.tutor, day="Mon", time_slot=time(15, 0), is_active=True,
        )
        future_date = date.today() + timedelta(days=7)
        while future_date.strftime("%a")[:3] != "Mon":
            future_date += timedelta(days=1)

        response = self.client.post(
            "/api/bookings/confirm/",
            {
                "tutor_id": self.tutor_profile.id,
                "slots": [{
                    "availability_id": availability.id,
                    "session_date": future_date.isoformat(),
                    "session_mode": "Online",
                }],
                "payment_method": self.cash_method.method_id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.data["error"], "Payment method does not match this session's mode."
        )
        self.assertFalse(Booking.objects.filter(tutor=self.tutor).exists())


class ProfileStatusWalletNegativeTests(APITestCase):
    """profile_status exposes wallet_negative (boolean only, never the balance) for tutors
    whose Wallet.balance has gone negative -- the signal the tutor debt banner reads."""

    def setUp(self):
        self.tutor_user = User.objects.create_user(
            username="wallet-negative-tutor",
            email="wallet-negative-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Debt", mname="", lname="Tutor", role="Tutor",
            profile_completed=True,
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("280.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.tutee_user = User.objects.create_user(
            username="wallet-negative-tutee",
            email="wallet-negative-tutee@example.com",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user, fname="Debt", mname="", lname="Tutee", role="Tutee",
            profile_completed=True,
        )

    def test_wallet_negative_false_when_balance_zero_or_positive(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.get("/api/profile/status/")
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["wallet_negative"])

        wallet = Wallet.objects.get(tutor=self.tutor)
        wallet.balance = Decimal("50.00")
        wallet.save(update_fields=["balance"])

        response = self.client.get("/api/profile/status/")
        self.assertFalse(response.data["wallet_negative"])

    def test_wallet_negative_true_when_balance_below_zero(self):
        wallet = Wallet.objects.get(tutor=self.tutor)
        wallet.balance = Decimal("-25.00")
        wallet.save(update_fields=["balance"])
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.get("/api/profile/status/")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["wallet_negative"])
        self.assertNotIn("wallet_balance", response.data)

    def test_wallet_negative_false_for_tutee(self):
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get("/api/profile/status/")

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["wallet_negative"])


class CommissionTermsDisclosureTests(APITestCase):
    """Commission-disclosure acceptance (ADR-0010): tutor_setup blocks without acknowledgement,
    accept_commission_terms covers already-onboarded tutors retroactively, and profile_status
    exposes commission_terms_accepted for the router guard."""

    def setUp(self):
        self.tutor_user = User.objects.create_user(
            username="commission-tutor",
            email="commission-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Comm", mname="", lname="Tutor", role="Tutor",
        )
        self.tutor = Tutor.objects.create(profile=self.tutor_profile)

        self.tutee_user = User.objects.create_user(
            username="commission-tutee",
            email="commission-tutee@example.com",
            password="password",
        )
        UserProfile.objects.create(
            user=self.tutee_user, fname="Comm", mname="", lname="Tutee", role="Tutee",
            profile_completed=True,
        )

    def test_tutor_setup_rejects_without_acknowledgement(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            "/api/tutor/setup/",
            {"teaching_level": "College", "hourly_rate": "300.00"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.tutor.refresh_from_db()
        self.assertIsNone(self.tutor.commission_terms_accepted_at)

    def test_tutor_setup_accepts_and_stamps_timestamp(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            "/api/tutor/setup/",
            {
                "teaching_level": "College",
                "hourly_rate": "300.00",
                "commission_terms_accepted": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.tutor.refresh_from_db()
        self.assertIsNotNone(self.tutor.commission_terms_accepted_at)

    def test_tutor_setup_does_not_require_reacknowledgement_once_accepted(self):
        self.tutor.commission_terms_accepted_at = timezone.now()
        self.tutor.save(update_fields=["commission_terms_accepted_at"])
        original_timestamp = self.tutor.commission_terms_accepted_at
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            "/api/tutor/setup/",
            {"teaching_level": "College", "hourly_rate": "350.00"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.tutor.refresh_from_db()
        self.assertEqual(self.tutor.commission_terms_accepted_at, original_timestamp)

    def test_accept_commission_terms_endpoint_is_idempotent(self):
        self.client.force_authenticate(user=self.tutor_user)

        first_response = self.client.post("/api/tutor/accept-commission-terms/")
        self.assertEqual(first_response.status_code, 200)
        self.tutor.refresh_from_db()
        first_timestamp = self.tutor.commission_terms_accepted_at
        self.assertIsNotNone(first_timestamp)

        second_response = self.client.post("/api/tutor/accept-commission-terms/")
        self.assertEqual(second_response.status_code, 200)
        self.tutor.refresh_from_db()
        self.assertEqual(self.tutor.commission_terms_accepted_at, first_timestamp)

    def test_profile_status_reflects_commission_terms_accepted_for_tutor(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.get("/api/profile/status/")
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["commission_terms_accepted"])

        self.tutor.commission_terms_accepted_at = timezone.now()
        self.tutor.save(update_fields=["commission_terms_accepted_at"])

        response = self.client.get("/api/profile/status/")
        self.assertTrue(response.data["commission_terms_accepted"])

    def test_profile_status_commission_terms_accepted_true_for_tutee(self):
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get("/api/profile/status/")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["commission_terms_accepted"])


@override_settings(
    PAYMONGO_WALLET_ID="wallet_test",
    PAYMONGO_CASHOUT_CALLBACK_URL="https://api.test/api/wallet/paymongo/callback/",
    CASHOUT_PROVIDER_FEE_PHP="10",
    CASHOUT_MIN_PHP="50",
)
class TutorCashOutTests(APITestCase):
    def setUp(self):
        self.tutor_user = User.objects.create_user(
            username="cashout-tutor",
            email="cashout-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Cash",
            mname="",
            lname="Tutor",
            role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("300.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.wallet = Wallet.objects.get(tutor=self.tutor)
        self.wallet.balance = Decimal("1000.00")
        self.wallet.save(update_fields=["balance"])
        self.client.force_authenticate(user=self.tutor_user)

    def destination_fields(self, **overrides):
        fields = {
            "destination_type": "gcash",
            "receiving_institution_id": "inst_gcash",
            "receiving_institution_name": "GCash",
            "receiving_institution_code": "GCASH",
            "account_number": "09171234567",
            "account_name": "Cash Tutor",
        }
        fields.update(overrides)
        return fields

    def paymongo_response(self, status_code, payload):
        response = Mock()
        response.status_code = status_code
        response.json.return_value = payload
        return response

    def provider_result(self, status_value="pending", transaction_id="wt_test_123"):
        return {
            "id": transaction_id,
            "status": status_value,
            "provider": "paymongo",
            "reference_number": "PMO-123",
            "fee": Decimal("0.00"),
            "net_amount": Decimal("500.00"),
        }

    def callback_payload(self, status_value="failed", transaction_id="wt_test_123"):
        return {
            "data": {
                "id": transaction_id,
                "attributes": {
                    "status": status_value,
                    "provider": "paymongo",
                    "reference_number": "PMO-123",
                    "provider_error": {"detail": "Receiving account rejected the transfer."},
                },
            }
        }

    def test_cashout_rejects_invalid_amount_and_insufficient_balance(self):
        destination = self.destination_fields()

        small_response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "49.99", **destination},
            format="json",
        )
        insufficient_response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "995.00", **destination},
            format="json",
        )

        self.assertEqual(small_response.status_code, 400)
        self.assertEqual(insufficient_response.status_code, 400)
        self.wallet.refresh_from_db()
        self.assertEqual(self.wallet.balance, Decimal("1000.00"))

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_deducts_amount_and_fee_and_stores_provider_data(self, mock_create):
        destination = self.destination_fields()
        mock_create.return_value = self.provider_result()

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.wallet.refresh_from_db()
        withdrawal = WithdrawalRequest.objects.get(id=response.data["id"])

        self.assertEqual(response.status_code, 201)
        self.assertEqual(self.wallet.balance, Decimal("490.00"))
        self.assertEqual(withdrawal.status, "pending")
        self.assertEqual(withdrawal.provider_wallet_transaction_id, "wt_test_123")
        self.assertEqual(withdrawal.provider_reference_number, "PMO-123")
        self.assertEqual(withdrawal.provider_fee, Decimal("10.00"))
        self.assertTrue(
            Transaction.objects.filter(
                wallet=self.wallet,
                transaction_type="withdrawal",
                amount=Decimal("-500.00"),
            ).exists()
        )
        self.assertTrue(
            Transaction.objects.filter(
                wallet=self.wallet,
                transaction_type="cashout_fee",
                amount=Decimal("-10.00"),
            ).exists()
        )

    @override_settings(PAYMONGO_CASHOUT_MOCK=False)
    @patch("studybuddy.paymongo_money_movement.requests.post")
    def test_cashout_sends_centavos_and_normalizes_provider_amounts(self, mock_post):
        destination = self.destination_fields()
        mock_post.return_value = self.paymongo_response(
            201,
            {
                "data": {
                    "id": "wt_test_live_123",
                    "attributes": {
                        "status": "pending",
                        "provider": "paymongo",
                        "reference_number": "PMO-LIVE-123",
                        "fee": 1250,
                        "net_amount": 48750,
                    },
                }
            },
        )

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["provider_fee"], 12.5)
        self.assertEqual(response.data["net_amount"], 487.5)
        self.assertEqual(
            mock_post.call_args.kwargs["json"]["data"]["attributes"]["amount"],
            50000,
        )
        self.assertIn(
            "/api/wallet/paymongo/callback/",
            mock_post.call_args.kwargs["json"]["data"]["attributes"]["callback_url"],
        )

    @override_settings(PAYMONGO_CASHOUT_CALLBACK_SECRET="")
    @patch("studybuddy.views.create_wallet_transaction")
    def test_failed_callback_refunds_amount_and_fee_once(self, mock_create):
        destination = self.destination_fields()
        mock_create.return_value = self.provider_result()

        create_response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )
        self.client.force_authenticate(user=None)

        first_callback = self.client.post(
            "/api/wallet/paymongo/callback/",
            self.callback_payload(),
            format="json",
        )
        second_callback = self.client.post(
            "/api/wallet/paymongo/callback/",
            self.callback_payload(),
            format="json",
        )

        self.wallet.refresh_from_db()
        withdrawal = WithdrawalRequest.objects.get(id=create_response.data["id"])

        self.assertEqual(first_callback.status_code, 200)
        self.assertEqual(second_callback.status_code, 200)
        self.assertEqual(withdrawal.status, "failed")
        self.assertEqual(self.wallet.balance, Decimal("1000.00"))
        self.assertEqual(
            Transaction.objects.filter(
                wallet=self.wallet,
                transaction_type__in=["withdrawal_reversal", "cashout_fee_reversal"],
            ).count(),
            2,
        )

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_with_inline_destination_fields_succeeds(self, mock_create):
        destination = self.destination_fields()
        mock_create.return_value = self.provider_result()

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", "note": "Rent money", **destination},
            format="json",
        )

        self.wallet.refresh_from_db()
        withdrawal = WithdrawalRequest.objects.get(id=response.data["id"])

        self.assertIn(response.status_code, (200, 201))
        self.assertEqual(self.wallet.balance, Decimal("490.00"))
        self.assertEqual(withdrawal.method, destination["destination_type"])
        self.assertEqual(withdrawal.account_number, destination["account_number"])
        self.assertEqual(withdrawal.account_name, destination["account_name"])
        self.assertEqual(withdrawal.note, "Rent money")

    @patch("studybuddy.views.create_wallet_transaction")
    def test_recent_cash_outs_returns_last_four_most_recent_first(self, mock_create):
        mock_create.return_value = self.provider_result()
        destination = self.destination_fields()
        self.wallet.balance = Decimal("10000.00")
        self.wallet.save(update_fields=["balance"])
        statuses = ["pending", "processed", "rejected", "failed", "flagged", "processed"]

        created_ids = []
        for status_value in statuses:
            response = self.client.post(
                "/api/wallet/cash-outs/",
                {"amount": "500.00", **destination},
                format="json",
            )
            withdrawal = WithdrawalRequest.objects.get(id=response.data["id"])
            withdrawal.status = status_value
            withdrawal.save(update_fields=["status"])
            created_ids.append(withdrawal.id)

        response = self.client.get("/api/wallet/cash-outs/recent/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 4)
        self.assertEqual(
            [item["id"] for item in response.data],
            list(reversed(created_ids[-4:])),
        )

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_new_destination_requires_confirmation(self, mock_create):
        mock_create.return_value = self.provider_result()
        self.wallet.balance = Decimal("2000.00")
        self.wallet.save(update_fields=["balance"])
        known_destination = self.destination_fields()

        self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **known_destination},
            format="json",
        )

        different_destination = self.destination_fields(
            destination_type="bank",
            receiving_institution_id="inst_bdo",
            receiving_institution_name="BDO",
            receiving_institution_code="BDO",
            account_number="0011223344",
            account_name="Cash Tutor",
            bank_name="BDO",
        )

        rejected_response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **different_destination},
            format="json",
        )

        self.assertEqual(rejected_response.status_code, 409)
        self.assertEqual(
            rejected_response.data["error"], "new_destination_confirmation_required"
        )
        self.assertEqual(WithdrawalRequest.objects.filter(tutor=self.tutor).count(), 1)

        confirmed_response = self.client.post(
            "/api/wallet/cash-outs/",
            {
                "amount": "500.00",
                "confirm_new_destination": True,
                **different_destination,
            },
            format="json",
        )

        self.assertIn(confirmed_response.status_code, (200, 201))
        self.assertEqual(WithdrawalRequest.objects.filter(tutor=self.tutor).count(), 2)

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_matching_recent_destination_skips_confirmation(self, mock_create):
        mock_create.return_value = self.provider_result()
        destination = self.destination_fields()

        self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )
        self.wallet.balance = Decimal("1000.00")
        self.wallet.save(update_fields=["balance"])

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertIn(response.status_code, (200, 201))

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_first_time_destination_skips_confirmation_when_no_history(self, mock_create):
        mock_create.return_value = self.provider_result()
        destination = self.destination_fields()
        self.assertFalse(WithdrawalRequest.objects.filter(tutor=self.tutor).exists())

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertIn(response.status_code, (200, 201))

    @override_settings(PAYMONGO_WALLET_ID="", PAYMONGO_CASHOUT_MOCK=True)
    @patch("studybuddy.paymongo_money_movement.requests.post")
    def test_cashout_mock_mode_succeeds_without_wallet_id_or_http_call(self, mock_post):
        destination = self.destination_fields()

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        withdrawal = WithdrawalRequest.objects.get(id=response.data["id"])

        self.assertEqual(response.status_code, 201)
        self.assertEqual(withdrawal.status, "processed")
        self.assertEqual(withdrawal.provider_wallet_transaction_id, f"mock_wtx_{withdrawal.id}")
        mock_post.assert_not_called()

    @override_settings(PAYMONGO_CASHOUT_MOCK=True)
    @patch("studybuddy.paymongo_money_movement.requests.get")
    def test_receiving_institutions_mock_mode_returns_list_without_http_call(self, mock_get):
        response = self.client.get("/api/wallet/receiving-institutions/")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(len(response.data) >= 1)
        names = {item["attributes"]["name"] for item in response.data}
        self.assertIn("GCash", names)
        mock_get.assert_not_called()

    @override_settings(PAYMONGO_WALLET_ID="", PAYMONGO_CASHOUT_MOCK=True)
    @patch("studybuddy.paymongo_money_movement.requests.post")
    def test_auto_processed_cashout_logs_platform_activity(self, mock_post):
        destination = self.destination_fields()

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        withdrawal = WithdrawalRequest.objects.get(id=response.data["id"])
        activities = PlatformActivity.objects.filter(activity_type="withdrawal_processed")

        self.assertEqual(withdrawal.status, "processed")
        self.assertEqual(activities.count(), 1)
        self.assertIn(str(withdrawal.id), activities.first().message)


@override_settings(
    PAYMONGO_WALLET_ID="wallet_test",
    PAYMONGO_CASHOUT_CALLBACK_URL="https://api.test/api/wallet/paymongo/callback/",
    CASHOUT_PROVIDER_FEE_PHP="10",
    CASHOUT_MIN_PHP="50",
)
class WalletCashOutEdgeCaseTests(APITestCase):
    def setUp(self):
        self.tutor_user = User.objects.create_user(
            username="cashout-edge-tutor",
            email="cashout-edge-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Edge",
            mname="",
            lname="Tutor",
            role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("300.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.wallet = Wallet.objects.get(tutor=self.tutor)
        self.wallet.balance = Decimal("1000.00")
        self.wallet.save(update_fields=["balance"])
        self.client.force_authenticate(user=self.tutor_user)

    def destination_fields(self, **overrides):
        fields = {
            "destination_type": "gcash",
            "receiving_institution_id": "inst_gcash",
            "receiving_institution_name": "GCash",
            "receiving_institution_code": "GCASH",
            "account_number": "09171234567",
            "account_name": "Edge Tutor",
        }
        fields.update(overrides)
        return fields

    def provider_result(self, transaction_id="wt_test_edge"):
        return {
            "id": transaction_id,
            "status": "pending",
            "provider": "paymongo",
            "reference_number": "PMO-EDGE",
            "fee": Decimal("0.00"),
            "net_amount": Decimal("0.00"),
        }

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_same_destination_works_across_different_amounts(self, mock_create):
        self.wallet.balance = Decimal("100000.00")
        self.wallet.save(update_fields=["balance"])
        destination = self.destination_fields()
        mock_create.return_value = self.provider_result()

        first_response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "600.00", **destination},
            format="json",
        )
        second_response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "50000.00", **destination},
            format="json",
        )

        self.assertEqual(first_response.status_code, 201)
        self.assertEqual(second_response.status_code, 201)
        self.assertEqual(
            WithdrawalRequest.objects.filter(
                tutor=self.tutor, account_number=destination["account_number"]
            ).count(),
            2,
        )

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_allows_exact_cap_amount(self, mock_create):
        self.wallet.balance = Decimal("100000.00")
        self.wallet.save(update_fields=["balance"])
        destination = self.destination_fields()
        mock_create.return_value = self.provider_result()

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "50000.00", **destination},
            format="json",
        )

        self.assertEqual(response.status_code, 201)

    def test_cashout_rejects_amount_above_cap(self):
        self.wallet.balance = Decimal("100000.00")
        self.wallet.save(update_fields=["balance"])
        destination = self.destination_fields()

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "50000.01", **destination},
            format="json",
        )

        self.wallet.refresh_from_db()
        self.assertEqual(response.status_code, 400)
        self.assertIn("Maximum cash-out", response.data["error"])
        self.assertEqual(self.wallet.balance, Decimal("100000.00"))
        self.assertFalse(WithdrawalRequest.objects.filter(tutor=self.tutor).exists())

    def test_cashout_rejects_missing_destination_type(self):
        destination = self.destination_fields()
        destination.pop("destination_type")

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(WithdrawalRequest.objects.filter(tutor=self.tutor).exists())

    def test_cashout_rejects_invalid_destination_type(self):
        destination = self.destination_fields(destination_type="paypal")

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(WithdrawalRequest.objects.filter(tutor=self.tutor).exists())

    def test_cashout_rejects_missing_account_number(self):
        destination = self.destination_fields()
        destination.pop("account_number")

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(WithdrawalRequest.objects.filter(tutor=self.tutor).exists())

    def test_cashout_rejects_missing_account_name(self):
        destination = self.destination_fields()
        destination.pop("account_name")

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(WithdrawalRequest.objects.filter(tutor=self.tutor).exists())

    def test_cashout_rejects_missing_bank_name_for_bank_destination(self):
        destination = self.destination_fields(
            destination_type="bank",
            receiving_institution_id="bdo",
            receiving_institution_name="BDO",
            receiving_institution_code="BDO",
        )

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(WithdrawalRequest.objects.filter(tutor=self.tutor).exists())

    def test_cashout_rejects_missing_receiving_institution(self):
        destination = self.destination_fields()
        destination.pop("receiving_institution_id")
        destination.pop("receiving_institution_name")

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(WithdrawalRequest.objects.filter(tutor=self.tutor).exists())

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_synchronous_paymongo_failure_reverses_immediately(self, mock_create):
        destination = self.destination_fields()
        mock_create.side_effect = PayMongoCashOutError("Receiving account rejected the transfer.")

        response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.wallet.refresh_from_db()
        withdrawal = WithdrawalRequest.objects.get(id=response.data["id"])

        self.assertEqual(response.status_code, 502)
        self.assertEqual(response.data["provider_error_message"], "Receiving account rejected the transfer.")
        self.assertEqual(withdrawal.status, "failed")
        self.assertEqual(self.wallet.balance, Decimal("1000.00"))
        self.assertEqual(
            Transaction.objects.filter(
                wallet=self.wallet,
                transaction_type__in=["withdrawal_reversal", "cashout_fee_reversal"],
            ).count(),
            2,
        )

    @patch("studybuddy.views.create_wallet_transaction")
    def test_cashout_new_destination_requires_confirmation(self, mock_create):
        WithdrawalRequest.objects.create(
            tutor=self.tutor,
            amount=Decimal("500.00"),
            method="bank",
            receiving_institution_id="bdo",
            receiving_institution_name="BDO",
            account_number="123456",
            account_name="Cash Tutor",
            status="pending",
        )
        mock_create.return_value = self.provider_result()

        destination = self.destination_fields(
            destination_type="bank",
            receiving_institution_id="bpi",
            receiving_institution_name="BPI",
            receiving_institution_code="BPI",
            account_number="123456",
            account_name="Cash Tutor",
            bank_name="BPI",
        )

        unconfirmed_response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", **destination},
            format="json",
        )

        self.assertEqual(unconfirmed_response.status_code, 409)
        self.assertEqual(
            WithdrawalRequest.objects.filter(tutor=self.tutor, receiving_institution_id="bpi").count(),
            0,
        )

        confirmed_response = self.client.post(
            "/api/wallet/cash-outs/",
            {"amount": "500.00", "confirm_new_destination": True, **destination},
            format="json",
        )

        self.assertEqual(confirmed_response.status_code, 201)
        self.assertEqual(
            WithdrawalRequest.objects.filter(tutor=self.tutor, receiving_institution_id="bpi").count(),
            1,
        )

    def test_payout_destinations_endpoints_removed(self):
        get_response = self.client.get("/api/wallet/payout-destinations/")
        post_response = self.client.post(
            "/api/wallet/payout-destinations/",
            self.destination_fields(),
            format="json",
        )

        self.assertEqual(get_response.status_code, 404)
        self.assertEqual(post_response.status_code, 404)


class WalletCashInTests(APITestCase):
    def setUp(self):
        self.tutor_user = User.objects.create_user(
            username="cashin-tutor",
            email="cashin-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="CashIn",
            mname="",
            lname="Tutor",
            role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("300.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.wallet = Wallet.objects.get(tutor=self.tutor)
        self.client.force_authenticate(user=self.tutor_user)

    def paymongo_response(self, status_code, payload):
        response = Mock()
        response.status_code = status_code
        response.json.return_value = payload
        return response

    def checkout_session_payload(self, session_id="cs_topup_123"):
        return {
            "data": {
                "id": session_id,
                "attributes": {"checkout_url": "https://checkout.paymongo.com/topup-session"},
            },
        }

    def retrieve_payload(self, status_value="paid"):
        return {
            "data": {
                "id": "cs_topup_123",
                "attributes": {"status": status_value},
            },
        }

    @patch("studybuddy.views.requests.post")
    def test_initiate_cash_in_creates_pending_topup_and_checkout(self, mock_post):
        mock_post.return_value = self.paymongo_response(200, self.checkout_session_payload())

        response = self.client.post("/api/wallet/cash-in/", {"amount": "500.00"}, format="json")

        topup = WalletTopUp.objects.get(id=response.data["id"])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["checkout_url"], "https://checkout.paymongo.com/topup-session")
        self.assertEqual(topup.status, "pending")
        self.assertEqual(topup.amount, Decimal("500.00"))
        self.assertEqual(topup.provider_reference, "cs_topup_123")

    def test_initiate_cash_in_rejects_invalid_amount(self):
        response = self.client.post("/api/wallet/cash-in/", {"amount": "0"}, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertFalse(WalletTopUp.objects.filter(tutor=self.tutor).exists())

    def test_initiate_cash_in_rejects_amount_below_minimum(self):
        response = self.client.post("/api/wallet/cash-in/", {"amount": "49.99"}, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertIn("Minimum cash-in", response.data["error"])
        self.assertFalse(WalletTopUp.objects.filter(tutor=self.tutor).exists())

    @patch("studybuddy.views.requests.post")
    def test_initiate_cash_in_marks_failed_on_missing_checkout_url(self, mock_post):
        mock_post.return_value = self.paymongo_response(
            200, {"data": {"id": "cs_missing_url", "attributes": {}}}
        )

        response = self.client.post("/api/wallet/cash-in/", {"amount": "500.00"}, format="json")

        topup = WalletTopUp.objects.get(tutor=self.tutor)
        self.assertEqual(response.status_code, 502)
        self.assertEqual(topup.status, "failed")

    @patch("studybuddy.views.requests.post")
    def test_initiate_cash_in_marks_failed_on_provider_error(self, mock_post):
        mock_post.return_value = self.paymongo_response(
            400, {"errors": [{"detail": "Invalid line item."}]}
        )

        response = self.client.post("/api/wallet/cash-in/", {"amount": "500.00"}, format="json")

        topup = WalletTopUp.objects.get(tutor=self.tutor)
        self.assertEqual(response.status_code, 502)
        self.assertEqual(response.data["error"], "Invalid line item.")
        self.assertEqual(topup.status, "failed")

    @patch("studybuddy.views.requests.post")
    def test_initiate_cash_in_marks_failed_on_network_error(self, mock_post):
        import requests

        mock_post.side_effect = requests.RequestException("connection reset")

        response = self.client.post("/api/wallet/cash-in/", {"amount": "500.00"}, format="json")

        topup = WalletTopUp.objects.get(tutor=self.tutor)
        self.assertEqual(response.status_code, 502)
        self.assertEqual(topup.status, "failed")

    @patch("studybuddy.views.requests.get")
    @patch("studybuddy.views.requests.post")
    def test_verify_cash_in_credits_wallet_exactly_once(self, mock_post, mock_get):
        mock_post.return_value = self.paymongo_response(200, self.checkout_session_payload())
        mock_get.return_value = self.paymongo_response(200, self.retrieve_payload("paid"))

        create_response = self.client.post("/api/wallet/cash-in/", {"amount": "500.00"}, format="json")
        topup_id = create_response.data["id"]

        first_verify = self.client.post(f"/api/wallet/cash-in/{topup_id}/verify/")
        second_verify = self.client.post(f"/api/wallet/cash-in/{topup_id}/verify/")

        self.wallet.refresh_from_db()
        self.assertEqual(first_verify.status_code, 200)
        self.assertEqual(second_verify.status_code, 200)
        self.assertEqual(self.wallet.balance, Decimal("500.00"))
        self.assertEqual(mock_get.call_count, 1)
        self.assertEqual(
            Transaction.objects.filter(wallet=self.wallet, transaction_type="cash_in").count(),
            1,
        )

    @patch("studybuddy.views.requests.get")
    @patch("studybuddy.views.requests.post")
    def test_verify_cash_in_rejects_unpaid_checkout(self, mock_post, mock_get):
        mock_post.return_value = self.paymongo_response(200, self.checkout_session_payload())
        mock_get.return_value = self.paymongo_response(200, self.retrieve_payload("pending"))

        create_response = self.client.post("/api/wallet/cash-in/", {"amount": "500.00"}, format="json")
        response = self.client.post(f"/api/wallet/cash-in/{create_response.data['id']}/verify/")

        self.wallet.refresh_from_db()
        self.assertEqual(response.status_code, 400)
        self.assertEqual(self.wallet.balance, Decimal("0.00"))

    def test_verify_cash_in_rejects_missing_provider_reference(self):
        topup = WalletTopUp.objects.create(tutor=self.tutor, amount=Decimal("500.00"), status="pending")

        response = self.client.post(f"/api/wallet/cash-in/{topup.id}/verify/")

        self.assertEqual(response.status_code, 400)

    @patch("studybuddy.views.requests.get")
    @patch("studybuddy.views.requests.post")
    def test_verify_cash_in_returns_502_on_provider_error(self, mock_post, mock_get):
        mock_post.return_value = self.paymongo_response(200, self.checkout_session_payload())
        mock_get.return_value = self.paymongo_response(500, {"errors": []})

        create_response = self.client.post("/api/wallet/cash-in/", {"amount": "500.00"}, format="json")
        response = self.client.post(f"/api/wallet/cash-in/{create_response.data['id']}/verify/")

        self.wallet.refresh_from_db()
        self.assertEqual(response.status_code, 502)
        self.assertEqual(self.wallet.balance, Decimal("0.00"))

    @patch("studybuddy.views.requests.post")
    def test_verify_cash_in_404s_for_another_tutors_topup(self, mock_post):
        mock_post.return_value = self.paymongo_response(200, self.checkout_session_payload())
        create_response = self.client.post("/api/wallet/cash-in/", {"amount": "500.00"}, format="json")
        topup_id = create_response.data["id"]

        other_user = User.objects.create_user(
            username="other-cashin-tutor",
            email="other-cashin-tutor@example.com",
            password="password",
        )
        other_profile = UserProfile.objects.create(
            user=other_user, fname="Other", mname="", lname="Tutor", role="Tutor",
        )
        Tutor.objects.create(
            profile=other_profile,
            hourly_rate=Decimal("300.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.client.force_authenticate(user=other_user)

        response = self.client.post(f"/api/wallet/cash-in/{topup_id}/verify/")

        self.assertEqual(response.status_code, 404)


class SessionCreditWalletTests(APITestCase):
    def setUp(self):
        self.tutee_profile = UserProfile.objects.create(
            user=User.objects.create_user(
                username="credit-tutee", email="credit-tutee@example.com", password="password",
            ),
            fname="Credit",
            mname="",
            lname="Tutee",
            role="Tutee",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=User.objects.create_user(
                username="credit-tutor", email="credit-tutor@example.com", password="password",
            ),
            fname="Credit",
            mname="",
            lname="Tutor",
            role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("300.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.wallet = Wallet.objects.get(tutor=self.tutor)
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor, day="Mon", time_slot=time(14, 0), is_active=True,
        )
        self.booking = Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 4, 12),
            session_mode="Online",
            booking_request_id=uuid4(),
            status="Confirmed",
        )

    def create_payment(self, code, amount=Decimal("1000.00")):
        method, _ = PaymentMethod.objects.get_or_create(
            code=code, defaults={"method_name": code, "is_active": True}
        )
        return Payment.objects.create(
            booking=self.booking, amount=amount, method=method, payment_status="Paid",
        )

    def test_paymongo_settled_booking_credits_90_percent(self):
        self.create_payment("PAYMONGO")

        credit_tutor_wallet(self.booking)

        self.wallet.refresh_from_db()
        self.assertEqual(self.wallet.balance, Decimal("900.00"))
        self.assertTrue(
            Transaction.objects.filter(
                wallet=self.wallet, transaction_type="session_credit", amount=Decimal("900.00"),
            ).exists()
        )

    def test_cash_booking_deducts_10_percent_commission(self):
        self.create_payment("CASH")

        credit_tutor_wallet(self.booking)

        self.wallet.refresh_from_db()
        self.assertEqual(self.wallet.balance, Decimal("-100.00"))
        self.assertTrue(
            Transaction.objects.filter(
                wallet=self.wallet, transaction_type="commission_deduction", amount=Decimal("-100.00"),
            ).exists()
        )

    def test_credit_tutor_wallet_is_idempotent(self):
        self.create_payment("PAYMONGO")

        credit_tutor_wallet(self.booking)
        credit_tutor_wallet(self.booking)

        self.wallet.refresh_from_db()
        self.assertEqual(self.wallet.balance, Decimal("900.00"))
        self.assertEqual(
            Transaction.objects.filter(wallet=self.wallet, transaction_type="session_credit").count(),
            1,
        )


class WalletStatusAndTransactionsTests(APITestCase):
    def setUp(self):
        self.tutor_user = User.objects.create_user(
            username="status-tutor", email="status-tutor@example.com", password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Status", mname="", lname="Tutor", role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("300.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.wallet = Wallet.objects.get(tutor=self.tutor)
        self.wallet.balance = Decimal("750.00")
        self.wallet.pending_amount = Decimal("50.00")
        self.wallet.save(update_fields=["balance", "pending_amount"])
        self.client.force_authenticate(user=self.tutor_user)

    def test_wallet_status_returns_balance_and_cashout_settings(self):
        response = self.client.get("/api/wallet/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["balance"], 750.00)
        self.assertEqual(response.data["pending_amount"], 50.00)
        self.assertIn("cashin_minimum", response.data)
        self.assertIn("cashout_minimum", response.data)
        self.assertIn("cashout_provider_fee", response.data)

    def test_wallet_transactions_returns_history_ordered_newest_first(self):
        older = Transaction.objects.create(
            wallet=self.wallet,
            transaction_type="session_credit",
            amount=Decimal("100.00"),
            description="Older credit",
        )
        newer = Transaction.objects.create(
            wallet=self.wallet,
            transaction_type="cash_in",
            amount=Decimal("200.00"),
            description="Newer top-up",
        )

        response = self.client.get("/api/wallet/transactions/")

        self.assertEqual(response.status_code, 200)
        ids = [tx["id"] for tx in response.data]
        self.assertEqual(ids[0], newer.id)
        self.assertIn(older.id, ids)


class DevWalletFundsTests(APITestCase):
    """Calls the dev_add/remove_wallet_funds views directly via APIRequestFactory.

    /api/dev/wallet/add/ and /remove/ are only added to urlpatterns when
    settings.DEBUG is True at process startup (see urls.py); override_settings
    in a test can't retroactively register them. Calling the view functions
    directly exercises their real DEBUG check without depending on urlconf.
    """

    def setUp(self):
        self.factory = APIRequestFactory()
        self.tutor_user = User.objects.create_user(
            username="dev-funds-tutor", email="dev-funds-tutor@example.com", password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Dev", mname="", lname="Tutor", role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("300.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.wallet = Wallet.objects.get(tutor=self.tutor)

    def call_view(self, view, amount):
        request = self.factory.post(f"/api/dev/wallet/_/", {"amount": amount}, format="json")
        force_authenticate(request, user=self.tutor_user)
        return view(request)

    @override_settings(DEBUG=True)
    def test_dev_add_wallet_funds_increases_balance(self):
        response = self.call_view(dev_add_wallet_funds, "300")

        self.wallet.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.wallet.balance, Decimal("300.00"))

    @override_settings(DEBUG=True)
    def test_dev_remove_wallet_funds_decreases_balance(self):
        self.wallet.balance = Decimal("300.00")
        self.wallet.save(update_fields=["balance"])

        response = self.call_view(dev_remove_wallet_funds, "100")

        self.wallet.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.wallet.balance, Decimal("200.00"))

    @override_settings(BOOKING_DEV_TOOLS_ENABLED=False)
    def test_dev_wallet_funds_404s_when_dev_tools_disabled(self):
        """Gated by BOOKING_DEV_TOOLS_ENABLED, not DEBUG -- see
        test_force_live_returns_404_when_booking_dev_tools_disabled for the same pattern."""
        add_response = self.call_view(dev_add_wallet_funds, "300")
        remove_response = self.call_view(dev_remove_wallet_funds, "100")

        self.assertEqual(add_response.status_code, 404)
        self.assertEqual(remove_response.status_code, 404)


class TuteeProfileTests(APITestCase):
    def setUp(self):
        self.tutee_user = User.objects.create_user(
            username="tutee-test",
            email="tutee@example.com",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Tutee",
            mname="",
            lname="Test",
            role="Tutee",
            year_level=11,
        )

    def test_get_profile_includes_avatar_url(self):
        self.client.force_authenticate(user=self.tutee_user)
        response = self.client.get('/api/tutee/profile/')
        self.assertEqual(response.status_code, 200)
        self.assertIn('profile_picture_url', response.data)

    def test_upload_avatar_success(self):
        from io import BytesIO
        from PIL import Image
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.client.force_authenticate(user=self.tutee_user)
        buffer = BytesIO()
        Image.new("RGB", (64, 64), (120, 80, 200)).save(buffer, format="JPEG")
        buffer.seek(0)
        image = SimpleUploadedFile("avatar.jpg", buffer.read(), content_type="image/jpeg")
        response = self.client.post('/api/tutee/profile/avatar/', {'avatar': image}, format='multipart')
        self.assertEqual(response.status_code, 200)
        self.assertIn('profile_picture_url', response.data)

    def test_update_profile_can_clear_course_and_subjects(self):
        course = Course.objects.create(
            course_code="BSCS",
            course_name="BS Computer Science",
        )
        subject = Subjects.objects.create(
            subject_code="MATH101",
            subject_name="College Algebra",
            department="Math",
            category="College",
        )
        self.tutee_profile.course = course
        self.tutee_profile.save()
        preference = Preference.objects.create(user=self.tutee_profile)
        preference.subjects.add(subject)

        self.client.force_authenticate(user=self.tutee_user)
        response = self.client.put('/api/tutee/profile/update/', {
            "course": "",
            "year_level": 7,
            "subjects": [],
        }, format='json')

        self.assertEqual(response.status_code, 200)
        self.tutee_profile.refresh_from_db()
        preference.refresh_from_db()
        self.assertIsNone(self.tutee_profile.course)
        self.assertEqual(preference.subjects.count(), 0)


class TutorProfileTests(APITestCase):
    def setUp(self):
        self.tutor_user = User.objects.create_user(
            username="tutor-test",
            email="tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Tutor",
            mname="",
            lname="Test",
            role="Tutor",
            year_level=12,
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("250.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )

    def test_get_profile_includes_profile_picture_url(self):
        self.client.force_authenticate(user=self.tutor_user)
        response = self.client.get('/api/tutor/profile/')
        self.assertEqual(response.status_code, 200)
        self.assertIn('profile_picture_url', response.data)

    def test_upload_avatar_success(self):
        from io import BytesIO
        from PIL import Image
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.client.force_authenticate(user=self.tutor_user)
        buffer = BytesIO()
        Image.new("RGB", (64, 64), (120, 80, 200)).save(buffer, format="JPEG")
        buffer.seek(0)
        image = SimpleUploadedFile("avatar.jpg", buffer.read(), content_type="image/jpeg")
        response = self.client.post('/api/tutor/profile/avatar/', {'avatar': image}, format='multipart')
        self.assertEqual(response.status_code, 200)
        self.assertIn('profile_picture_url', response.data)

    def test_upload_avatar_rejects_non_image(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.client.force_authenticate(user=self.tutor_user)
        document = SimpleUploadedFile("avatar.pdf", b"fake_pdf_content", content_type="application/pdf")
        response = self.client.post('/api/tutor/profile/avatar/', {'avatar': document}, format='multipart')
        self.assertEqual(response.status_code, 400)

    def test_upload_avatar_rejects_missing_file(self):
        self.client.force_authenticate(user=self.tutor_user)
        response = self.client.post('/api/tutor/profile/avatar/', {}, format='multipart')
        self.assertEqual(response.status_code, 400)


class TutorDocumentRenewalTests(APITestCase):
    def setUp(self):
        from datetime import timedelta
        from django.utils import timezone

        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU",
            school_email_domain="cpu.edu",
            is_active=True,
        )
        self.other_institution = PartnerInstitution.objects.create(
            institution_name="Other University",
            school_email_domain="other.edu",
            is_active=True,
        )
        self.tutor_user = User.objects.create_user(
            username="renewal-tutor@cpu.edu",
            email="renewal-tutor@cpu.edu",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Renewal",
            mname="",
            lname="Tutor",
            role="Tutor",
            institution=self.institution,
        )
        Tutor.objects.create(profile=self.tutor_profile)
        self.reviewed_at = timezone.now() - timedelta(days=91)
        self.application = TutorApplication.objects.create(
            profile=self.tutor_profile,
            school_id=self.upload("initial-id.jpg", b"initial-id", "image/jpeg"),
            enrollment_proof=self.upload("initial-rf.pdf", b"initial-rf", "application/pdf"),
            application_status="approved",
            reviewed_at=self.reviewed_at,
        )
        self.admin_user = User.objects.create_user(
            username="admin@cpu.edu",
            email="admin@cpu.edu",
            password="password",
        )
        self.admin_profile = UserProfile.objects.create(
            user=self.admin_user,
            fname="Admin",
            mname="",
            lname="User",
            role="SuperAdmin",
            institution=self.institution,
        )

    def upload(self, name, content, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, content, content_type=content_type)

    def renewal_payload(self):
        return {
            "school_id": self.upload("renewal-id.jpg", b"renewal-id", "image/jpeg"),
            "enrollment_proof": self.upload("renewal-rf.pdf", b"renewal-rf", "application/pdf"),
            "reason_to_tutor": "Still enrolled and available to tutor.",
        }

    def test_profile_status_exposes_due_document_renewal_state(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.get("/api/profile/status/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["application_status"], "approved")
        self.assertEqual(response.data["document_renewal_status"], "due")
        self.assertTrue(response.data["can_submit_document_renewal"])
        self.assertIsNotNone(response.data["document_renewal_due_at"])

    def test_tutor_can_submit_due_document_renewal(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            "/api/tutor-application/renewal/",
            self.renewal_payload(),
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        renewal = TutorDocumentRenewalReview.objects.get(application=self.application)
        self.assertEqual(renewal.status, "pending")
        self.assertEqual(response.data["document_renewal_status"], "pending")

    def test_legacy_resubmit_endpoint_creates_document_renewal_for_approved_tutor(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            "/api/tutor-application/resubmit/",
            self.renewal_payload(),
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        self.application.refresh_from_db()
        self.assertEqual(self.application.application_status, "approved")
        self.assertEqual(TutorDocumentRenewalReview.objects.filter(status="pending").count(), 1)

    def test_tutor_cannot_submit_verified_document_renewal_before_due(self):
        from django.utils import timezone

        self.application.reviewed_at = timezone.now()
        self.application.save(update_fields=["reviewed_at"])
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            "/api/tutor-application/renewal/",
            self.renewal_payload(),
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(TutorDocumentRenewalReview.objects.count(), 0)

    def test_admin_can_approve_document_renewal_and_due_date_moves_forward(self):
        renewal = TutorDocumentRenewalReview.objects.create(
            application=self.application,
            profile=self.tutor_profile,
            school_id=self.upload("pending-id.jpg", b"pending-id", "image/jpeg"),
            enrollment_proof=self.upload("pending-rf.pdf", b"pending-rf", "application/pdf"),
            status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.patch(
            f"/api/admin/tutor-document-renewals/{renewal.id}/",
            {"application_status": "approved"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        renewal.refresh_from_db()
        self.assertEqual(renewal.status, "approved")
        self.assertEqual(self.application.document_renewal_status(), "verified")
        self.assertGreater(self.application.document_renewal_due_at(), renewal.reviewed_at)

    def test_admin_can_reject_document_renewal_and_tutor_can_resubmit(self):
        renewal = TutorDocumentRenewalReview.objects.create(
            application=self.application,
            profile=self.tutor_profile,
            school_id=self.upload("pending-id.jpg", b"pending-id", "image/jpeg"),
            enrollment_proof=self.upload("pending-rf.pdf", b"pending-rf", "application/pdf"),
            status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)

        reject_response = self.client.patch(
            f"/api/admin/tutor-document-renewals/{renewal.id}/",
            {
                "application_status": "rejected",
                "rejection_reason": "RF is not for the current term.",
            },
            format="json",
        )

        self.assertEqual(reject_response.status_code, 200)
        renewal.refresh_from_db()
        self.assertEqual(renewal.status, "rejected")
        self.client.force_authenticate(user=self.tutor_user)
        resubmit_response = self.client.post(
            "/api/tutor-application/renewal/",
            self.renewal_payload(),
            format="multipart",
        )
        self.assertEqual(resubmit_response.status_code, 201)
        self.assertEqual(TutorDocumentRenewalReview.objects.filter(status="pending").count(), 1)



class ApplicationVerificationSharedBaseTests(APITestCase):
    """Proves TutorApplication and TuteeApplication share identical renewal-cadence
    behavior via their common abstract base (Phase 1 of tutee enrollment verification,
    see docs/plans/2026-07-01-tutee-verification-phase1-model.md)."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU",
            school_email_domain="cpu.edu",
            is_active=True,
        )
        self.tutor_user = User.objects.create_user(
            username="shared-base-tutor@cpu.edu",
            email="shared-base-tutor@cpu.edu",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Shared",
            mname="",
            lname="Tutor",
            role="Tutor",
            institution=self.institution,
        )
        self.tutee_user = User.objects.create_user(
            username="shared-base-tutee@cpu.edu",
            email="shared-base-tutee@cpu.edu",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Shared",
            mname="",
            lname="Tutee",
            role="Tutee",
            institution=self.institution,
        )

    def upload(self, name, content, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, content, content_type=content_type)

    def make_application(self, model, profile, reviewed_at):
        return model.objects.create(
            profile=profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
            reviewed_at=reviewed_at,
        )

    def assert_shared_renewal_behavior(self, model, review_model, profile):
        from datetime import timedelta
        from django.utils import timezone

        due_reviewed_at = timezone.now() - timedelta(days=91)
        application = self.make_application(model, profile, due_reviewed_at)

        # Verified just after approval, due once DOCUMENT_RENEWAL_INTERVAL_DAYS has elapsed.
        self.assertEqual(
            application.document_renewal_due_at(),
            due_reviewed_at + timedelta(days=model.DOCUMENT_RENEWAL_INTERVAL_DAYS),
        )
        self.assertEqual(application.document_renewal_status(), "due")
        self.assertTrue(application.can_submit_document_renewal())

        # A pending renewal review takes precedence over "due".
        pending_renewal = review_model.objects.create(
            application=application,
            profile=profile,
            school_id=self.upload("renewal-id.jpg", b"renewal-id", "image/jpeg"),
            enrollment_proof=self.upload("renewal-rf.pdf", b"renewal-rf", "application/pdf"),
            status="pending",
        )
        self.assertEqual(application.latest_document_renewal_review(), pending_renewal)
        self.assertEqual(application.document_renewal_status(), "pending")
        self.assertFalse(application.can_submit_document_renewal())

        # Approving the renewal moves the renewal clock forward and re-verifies.
        pending_renewal.status = "approved"
        pending_renewal.reviewed_at = timezone.now()
        pending_renewal.save(update_fields=["status", "reviewed_at"])
        self.assertEqual(application.latest_approved_document_review_at(), pending_renewal.reviewed_at)
        self.assertEqual(application.document_renewal_status(), "verified")
        self.assertGreater(application.document_renewal_due_at(), due_reviewed_at)

        # A rejected renewal makes the application resubmittable, not "verified".
        pending_renewal.status = "rejected"
        pending_renewal.reviewed_at = timezone.now()
        pending_renewal.save(update_fields=["status", "reviewed_at"])
        self.assertEqual(application.document_renewal_status(), "rejected")
        self.assertTrue(application.can_submit_document_renewal())

    def test_tutor_application_shared_renewal_behavior(self):
        self.assert_shared_renewal_behavior(
            TutorApplication, TutorDocumentRenewalReview, self.tutor_profile
        )

    def test_tutee_application_shared_renewal_behavior(self):
        self.assert_shared_renewal_behavior(
            TuteeApplication, TuteeDocumentRenewalReview, self.tutee_profile
        )

    def test_reminder_dedup_fields_default_null_for_both_roles(self):
        tutor_application = self.make_application(
            TutorApplication, self.tutor_profile, timezone_now_minus_days(1)
        )
        tutee_application = self.make_application(
            TuteeApplication, self.tutee_profile, timezone_now_minus_days(1)
        )

        for application in (tutor_application, tutee_application):
            self.assertIsNone(application.reminder_7day_sent_at)
            self.assertIsNone(application.reminder_1day_sent_at)

    def test_generalized_document_review_context_matches_tutor_shape(self):
        from .views import get_document_review_context, get_tutor_document_review_context

        due_reviewed_at = timezone_now_minus_days(91)
        tutor_application = self.make_application(TutorApplication, self.tutor_profile, due_reviewed_at)
        tutee_application = self.make_application(TuteeApplication, self.tutee_profile, due_reviewed_at)

        tutor_context = get_tutor_document_review_context(self.tutor_profile)
        generalized_tutor_context = get_document_review_context(tutor_application, 'tutor')
        generalized_tutee_context = get_document_review_context(tutee_application, 'tutee')

        self.assertEqual(tutor_context, generalized_tutor_context)
        self.assertEqual(set(generalized_tutor_context.keys()), set(generalized_tutee_context.keys()))
        self.assertEqual(generalized_tutee_context["document_renewal_status"], "due")
        self.assertTrue(generalized_tutee_context["can_submit_document_renewal"])


def timezone_now_minus_days(days):
    from datetime import timedelta
    from django.utils import timezone

    return timezone.now() - timedelta(days=days)


class BookingVerificationGateTests(APITestCase):
    """Phase 2 of tutee enrollment verification: can_create_new_booking gates new booking creation
    (tutee) and accepting a pending booking request (tutor). See
    docs/plans/2026-07-01-tutee-verification-phase2-gate.md."""

    ENFORCED = {'TUTEE_VERIFICATION_ENFORCEMENT_START_DATE': '2020-01-01'}
    NOT_YET_ENFORCED = {'TUTEE_VERIFICATION_ENFORCEMENT_START_DATE': None}
    MALFORMED = {'TUTEE_VERIFICATION_ENFORCEMENT_START_DATE': 'not-a-date'}

    def setUp(self):
        from datetime import timedelta
        from django.utils import timezone
        from .views import WEEKDAY_MAP

        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU",
            school_email_domain="cpu.edu",
            is_active=True,
        )
        self.course = Course.objects.create(
            course_code="GATE101",
            course_name="Gatekeeping Studies",
        )
        self.subject = Subjects.objects.create(
            subject_code="GATE-SUBJ",
            subject_name="Catalog Approved Subject",
            department="Gate",
            category=self.course.course_code,
        )
        self.tutee_user = User.objects.create_user(
            username="gate-tutee@cpu.edu",
            email="gate-tutee@cpu.edu",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Gate",
            mname="",
            lname="Tutee",
            role="Tutee",
            institution=self.institution,
            course=self.course,
        )
        self.tutor_user = User.objects.create_user(
            username="gate-tutor@cpu.edu",
            email="gate-tutor@cpu.edu",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Gate",
            mname="",
            lname="Tutor",
            role="Tutor",
            institution=self.institution,
            course=self.course,
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("280.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        TutorSubjects.objects.create(
            tutor=self.tutor,
            subject=self.subject,
            expertise_level=5,
        )
        self.future_date = timezone.now().date() + timedelta(days=5)
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day=WEEKDAY_MAP[self.future_date.weekday()],
            time_slot=time(14, 0),
            is_active=True,
        )
        self.make_verified_application(TutorApplication, self.tutor_profile)

    def upload(self, name, content, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, content, content_type=content_type)

    def make_verified_application(self, model, profile):
        return model.objects.create(
            profile=profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
            reviewed_at=timezone_now_minus_days(1),
        )

    def make_due_application(self, model, profile):
        return model.objects.create(
            profile=profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
            reviewed_at=timezone_now_minus_days(91),
        )

    def post_confirm_booking(self, *, subject=None):
        self.client.force_authenticate(user=self.tutee_user)
        payload = {
            "tutor_id": self.tutor_profile.id,
            "slots": [{
                "availability_id": self.availability.id,
                "session_date": self.future_date.isoformat(),
                "session_mode": "Online",
            }],
        }
        if subject is not None:
            payload["subject"] = subject
        return self.client.post(
            "/api/bookings/confirm/",
            payload,
            format="json",
        )

    def create_accepted_booking_group(self, *, start_hour, group_id=None, request_id=None):
        from .views import WEEKDAY_MAP

        availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day=WEEKDAY_MAP[self.future_date.weekday()],
            time_slot=time(start_hour, 0),
            is_active=True,
        )

        return Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=availability,
            session_date=self.future_date,
            session_mode="Online",
            booking_request_id=request_id or uuid4(),
            session_group_id=group_id or uuid4(),
            status="Confirmed",
        )

    @override_settings(**ENFORCED)
    def test_tutee_without_application_blocked_once_enforced(self):
        response = self.post_confirm_booking()

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.data.get("code"), "verification_required")

    @override_settings(**NOT_YET_ENFORCED)
    def test_tutee_without_application_allowed_during_grace_period(self):
        response = self.post_confirm_booking()

        self.assertEqual(response.status_code, 200)

    @override_settings(**MALFORMED)
    def test_malformed_enforcement_date_fails_safe_as_not_enforced(self):
        response = self.post_confirm_booking()

        self.assertEqual(response.status_code, 200)

    @override_settings(**ENFORCED)
    def test_tutee_with_verified_application_can_book(self):
        self.make_verified_application(TuteeApplication, self.tutee_profile)

        response = self.post_confirm_booking()

        self.assertEqual(response.status_code, 200)
        booking = Booking.objects.get(id=response.data['booking_ids'][0])
        self.assertEqual(booking.status, 'Confirmed')
        self.assertTrue(booking.meeting_link)

    @override_settings(**NOT_YET_ENFORCED)
    def test_confirm_booking_persists_subject(self):
        response = self.post_confirm_booking(subject=self.subject.subject_code)

        self.assertEqual(response.status_code, 200)
        booking = Booking.objects.get(id=response.data["booking_ids"][0])
        self.assertEqual(booking.subject.subject_code, self.subject.subject_code)

    @override_settings(**NOT_YET_ENFORCED)
    def test_confirm_booking_without_subject_still_succeeds(self):
        response = self.post_confirm_booking()

        self.assertEqual(response.status_code, 200)
        booking = Booking.objects.get(id=response.data["booking_ids"][0])
        self.assertIsNone(booking.subject)

    @override_settings(**NOT_YET_ENFORCED)
    def test_confirm_booking_rejects_unrecognized_subject(self):
        response = self.post_confirm_booking(subject="NOT-RECOGNIZED")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.data["error"],
            "This subject is not recognized for your course catalog.",
        )

    @override_settings(**ENFORCED)
    def test_tutee_with_due_renewal_blocked(self):
        self.make_due_application(TuteeApplication, self.tutee_profile)

        response = self.post_confirm_booking()

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.data.get("code"), "verification_required")

    def create_pending_booking(self):
        return Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=self.future_date,
            session_mode="Online",
            booking_request_id=uuid4(),
            status="Pending",
        )

    def test_removed_approval_route_returns_404(self):
        booking = self.create_pending_booking()
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(f"/api/bookings/{booking.id}/approve/")

        self.assertEqual(response.status_code, 404)

    def test_removed_rejection_route_returns_404(self):
        booking = self.create_pending_booking()
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(f"/api/bookings/{booking.id}/reject/")

        self.assertEqual(response.status_code, 404)

    @override_settings(**ENFORCED)
    def test_tutor_acceptance_load_counts_session_groups_not_rows(self):
        shared_group_id = uuid4()
        self.create_accepted_booking_group(start_hour=15, group_id=shared_group_id)
        self.create_accepted_booking_group(start_hour=16, group_id=shared_group_id)

        self.assertEqual(self.tutor.accepted_session_load(), 1)

    @override_settings(**ENFORCED)
    def test_tutor_at_acceptance_limit_cannot_receive_new_booking(self):
        self.make_verified_application(TuteeApplication, self.tutee_profile)
        accepted_hours = [8, 9, 10, 11, 12, 13, 15, 16, 17, 18]
        for start_hour in accepted_hours:
            self.create_accepted_booking_group(start_hour=start_hour)

        response = self.post_confirm_booking()

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.data.get("code"), "session_load_limit_reached")
        self.assertEqual(response.data.get("accepted_session_load"), 10)
        self.assertEqual(response.data.get("session_load_limit"), 10)


class TuteeVerificationPhase3Tests(APITestCase):
    """Phase 3 of tutee enrollment verification: the mirrored tutee-facing endpoints, admin views, and
    the profile_status role dispatcher. See docs/plans/2026-07-01-tutee-verification-phase3-ui.md."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU",
            school_email_domain="cpu.edu",
            is_active=True,
        )
        self.other_institution = PartnerInstitution.objects.create(
            institution_name="Other University",
            school_email_domain="other.edu",
            is_active=True,
        )
        self.tutee_user = User.objects.create_user(
            username="phase3-tutee@cpu.edu",
            email="phase3-tutee@cpu.edu",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Phase3",
            mname="",
            lname="Tutee",
            role="Tutee",
            institution=self.institution,
        )
        self.admin_user = User.objects.create_user(
            username="phase3-admin@cpu.edu",
            email="phase3-admin@cpu.edu",
            password="password",
        )
        self.admin_profile = UserProfile.objects.create(
            user=self.admin_user,
            fname="Phase3",
            mname="",
            lname="Admin",
            role="SuperAdmin",
            institution=self.institution,
        )

    def upload(self, name, content, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, content, content_type=content_type)

    def test_tutee_application_status_returns_404_when_none_exists(self):
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get("/api/tutee-application/status/")

        self.assertEqual(response.status_code, 404)

    def test_tutee_application_resubmit_creates_initial_application_when_none_exists(self):
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.post(
            "/api/tutee-application/resubmit/",
            {
                "school_id": self.upload("id.jpg", b"id", "image/jpeg"),
                "enrollment_proof": self.upload("rf.pdf", b"rf", "application/pdf"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        application = TuteeApplication.objects.get(profile=self.tutee_profile)
        self.assertEqual(application.application_status, "pending")
        self.assertEqual(
            PlatformActivity.objects.filter(activity_type="tutee_application").count(), 1
        )

    def test_tutee_application_resubmit_rejected_goes_back_to_pending(self):
        application = TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="rejected",
            rejection_reason="Blurry photo",
        )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.post(
            "/api/tutee-application/resubmit/",
            {
                "school_id": self.upload("new-id.jpg", b"new-id", "image/jpeg"),
                "enrollment_proof": self.upload("new-rf.pdf", b"new-rf", "application/pdf"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        application.refresh_from_db()
        self.assertEqual(application.application_status, "pending")
        self.assertEqual(
            PlatformActivity.objects.filter(activity_type="tutee_application").count(), 1
        )

    def test_tutee_document_renewal_submit_when_due(self):
        application = TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
            reviewed_at=timezone_now_minus_days(91),
        )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.post(
            "/api/tutee-application/renewal/",
            {
                "school_id": self.upload("new-id.jpg", b"new-id", "image/jpeg"),
                "enrollment_proof": self.upload("new-rf.pdf", b"new-rf", "application/pdf"),
                "reason_to_tutor": "Still enrolled",
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["document_renewal_status"], "pending")
        renewal = TuteeDocumentRenewalReview.objects.get(application=application)
        self.assertEqual(renewal.status, "pending")

    def test_admin_tutee_application_list_combines_applications_and_renewals(self):
        TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/admin/tutee-applications/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["applicant_name"], "Phase3 Tutee")

    def test_admin_tutee_application_list_query_count_does_not_scale_with_n(self):
        """Regression test for the N+1 in latest_document_renewal_review(): the list endpoint must
        prefetch document_renewal_reviews for every row, not just the status=approved filter branch,
        or query count grows linearly with the number of applications. See
        docs/session-summaries/2026-07-08-tutee-application-admin-performance-summary.md."""

        def make_application(n):
            user = User.objects.create_user(
                username=f"phase3-tutee{n}@cpu.edu",
                email=f"phase3-tutee{n}@cpu.edu",
                password="password",
            )
            profile = UserProfile.objects.create(
                user=user, fname=f"Tutee{n}", mname="", lname="Test",
                role="Tutee", institution=self.institution,
            )
            return TuteeApplication.objects.create(
                profile=profile,
                school_id=self.upload(f"id{n}.jpg", b"id", "image/jpeg"),
                enrollment_proof=self.upload(f"rf{n}.pdf", b"rf", "application/pdf"),
                application_status="pending",
            )

        make_application(1)
        self.client.force_authenticate(user=self.admin_user)
        with CaptureQueriesContext(connection) as one_app_queries:
            self.client.get("/api/admin/tutee-applications/")

        make_application(2)
        make_application(3)
        make_application(4)
        make_application(5)
        with CaptureQueriesContext(connection) as five_app_queries:
            self.client.get("/api/admin/tutee-applications/")

        self.assertLessEqual(
            len(five_app_queries), len(one_app_queries) + 1,
            "query count scales with number of applications (N+1)",
        )

    def test_admin_can_approve_initial_tutee_application(self):
        application = TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.patch(
            f"/api/admin/tutee-applications/{application.id}/",
            {"application_status": "approved"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        application.refresh_from_db()
        self.assertEqual(application.application_status, "approved")

    def test_admin_can_approve_tutee_document_renewal_and_resets_dedup_fields(self):
        application = TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
            reviewed_at=timezone_now_minus_days(91),
            reminder_7day_sent_at=timezone_now_minus_days(3),
            reminder_1day_sent_at=timezone_now_minus_days(1),
        )
        renewal = TuteeDocumentRenewalReview.objects.create(
            application=application,
            profile=self.tutee_profile,
            school_id=self.upload("pending-id.jpg", b"pending-id", "image/jpeg"),
            enrollment_proof=self.upload("pending-rf.pdf", b"pending-rf", "application/pdf"),
            status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.patch(
            f"/api/admin/tutee-document-renewals/{renewal.id}/",
            {"application_status": "approved"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        application.refresh_from_db()
        self.assertEqual(application.document_renewal_status(), "verified")
        self.assertIsNone(application.reminder_7day_sent_at)
        self.assertIsNone(application.reminder_1day_sent_at)

    @override_settings(TUTEE_VERIFICATION_ENFORCEMENT_START_DATE='2020-01-01')
    def test_profile_status_dispatches_to_tutee_context_and_exposes_enforcement_flag(self):
        TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
            reviewed_at=timezone_now_minus_days(91),
        )
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get("/api/profile/status/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["document_renewal_status"], "due")
        self.assertTrue(response.data["tutee_verification_enforced"])

    def test_profile_status_for_tutee_with_no_application_returns_empty_context(self):
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.get("/api/profile/status/")

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.data["document_renewal_status"])
        self.assertFalse(response.data["document_renewal_required"])


@override_settings(
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    FRONTEND_URL="https://studybuddy.example",
    LOGIN_OTP_TTL_SECONDS=600,
    LOGIN_OTP_MAX_ATTEMPTS=2,
    REST_FRAMEWORK={
        "DEFAULT_AUTHENTICATION_CLASSES": (
            "rest_framework_simplejwt.authentication.JWTAuthentication",
        ),
        "DEFAULT_THROTTLE_RATES": {
            "anon": "1000/min",
            "user": "1000/min",
            "login": "1000/min",
        },
    },
)
class EmailAuthTests(APITestCase):
    def setUp(self):
        cache.clear()
        mail.outbox = []
        self.institution = PartnerInstitution.objects.create(
            institution_name="Example University",
            school_email_domain="example.edu",
            is_active=True,
        )
        self.user = User.objects.create_user(
            username="student@example.edu",
            email="student@example.edu",
            password="password",
        )
        self.profile = UserProfile.objects.create(
            user=self.user,
            fname="Email",
            mname="",
            lname="Auth",
            role="Tutee",
            institution=self.institution,
        )

    def login(self):
        return self.client.post(
            "/api/login/",
            {"email": self.user.email, "password": "password"},
            format="json",
        )

    def latest_otp_code(self):
        match = re.search(r"\b(\d{6})\b", mail.outbox[-1].body)
        self.assertIsNotNone(match)
        return match.group(1)

    def reset_link_parts(self):
        match = re.search(r"/reset-password/([^/\s]+)/([^/\s]+)", mail.outbox[-1].body)
        self.assertIsNotNone(match)
        return match.group(1), match.group(2)

    def test_login_sends_otp_challenge_instead_of_tokens_then_verify_returns_tokens(self):
        response = self.login()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["requires_2fa"])
        self.assertIn("challenge_id", response.data)
        self.assertNotIn("access", response.data)
        self.assertEqual(len(mail.outbox), 1)

        challenge = EmailOTPChallenge.objects.get(
            challenge_id=response.data["challenge_id"]
        )
        self.assertEqual(challenge.user, self.user)

        verify_response = self.client.post(
            "/api/login/verify-otp/",
            {
                "challenge_id": response.data["challenge_id"],
                "code": self.latest_otp_code(),
            },
            format="json",
        )

        self.assertEqual(verify_response.status_code, 200)
        self.assertIn("access", verify_response.data)
        self.assertIn("refresh", verify_response.data)
        self.assertEqual(verify_response.data["email"], self.user.email)
        challenge.refresh_from_db()
        self.assertIsNotNone(challenge.consumed_at)

    def test_staff_user_without_profile_can_login_as_admin(self):
        admin_user = User.objects.create_user(
            username="admin@example.edu",
            email="admin@example.edu",
            password="password",
            is_staff=True,
        )

        login_response = self.client.post(
            "/api/login/",
            {"email": admin_user.email, "password": "password"},
            format="json",
        )

        self.assertEqual(login_response.status_code, 200)
        self.assertTrue(login_response.data["requires_2fa"])

        profile = UserProfile.objects.get(user=admin_user)
        self.assertEqual(profile.role, "SuperAdmin")
        self.assertTrue(profile.profile_completed)
        self.assertTrue(profile.is_domain_exempt)

        verify_response = self.client.post(
            "/api/login/verify-otp/",
            {
                "challenge_id": login_response.data["challenge_id"],
                "code": self.latest_otp_code(),
            },
            format="json",
        )

        self.assertEqual(verify_response.status_code, 200)
        self.assertEqual(verify_response.data["role"], "SuperAdmin")
        self.assertEqual(verify_response.data["email"], admin_user.email)

    def test_staff_user_with_superadmin_profile_keeps_superadmin_on_login(self):
        admin_user = User.objects.create_user(
            username="superadmin@example.edu",
            email="superadmin@example.edu",
            password="password",
            is_staff=True,
        )
        UserProfile.objects.create(
            user=admin_user,
            fname="Super",
            mname="",
            lname="Admin",
            role="SuperAdmin",
            profile_completed=True,
            is_domain_exempt=True,
        )

        login_response = self.client.post(
            "/api/login/",
            {"email": admin_user.email, "password": "password"},
            format="json",
        )

        self.assertEqual(login_response.status_code, 200)

        verify_response = self.client.post(
            "/api/login/verify-otp/",
            {
                "challenge_id": login_response.data["challenge_id"],
                "code": self.latest_otp_code(),
            },
            format="json",
        )

        self.assertEqual(verify_response.status_code, 200)
        self.assertEqual(verify_response.data["role"], "SuperAdmin")

        profile = UserProfile.objects.get(user=admin_user)
        self.assertEqual(profile.role, "SuperAdmin")

    def test_make_superadmin_command_promotes_existing_user(self):
        target_user = User.objects.create_user(
            username="target@example.edu",
            email="target@example.edu",
            password="password",
        )
        profile = UserProfile.objects.create(
            user=target_user,
            fname="Target",
            mname="",
            lname="User",
            role="Tutee",
            profile_completed=False,
            is_domain_exempt=False,
            is_suspended=True,
        )

        output = StringIO()
        call_command("make_superadmin", target_user.email, stdout=output)

        target_user.refresh_from_db()
        profile.refresh_from_db()
        self.assertTrue(target_user.is_staff)
        self.assertEqual(profile.role, "SuperAdmin")
        self.assertTrue(profile.profile_completed)
        self.assertTrue(profile.is_domain_exempt)
        self.assertFalse(profile.is_suspended)
        self.assertIn("Promoted target@example.edu to SuperAdmin.", output.getvalue())

    def test_make_superadmin_command_prefers_login_username_when_email_is_duplicated(self):
        legacy_user = User.objects.create_user(
            username="duplicate",
            email="duplicate@example.edu",
            password="password",
        )
        legacy_profile = UserProfile.objects.create(
            user=legacy_user,
            fname="Legacy",
            mname="",
            lname="Admin",
            role="Tutee",
        )
        login_user = User.objects.create_user(
            username="duplicate@example.edu",
            email="duplicate@example.edu",
            password="password",
        )
        login_profile = UserProfile.objects.create(
            user=login_user,
            fname="Login",
            mname="",
            lname="Admin",
            role="Tutee",
        )

        output = StringIO()
        call_command("make_superadmin", login_user.email, stdout=output)

        login_user.refresh_from_db()
        legacy_profile.refresh_from_db()
        login_profile.refresh_from_db()
        self.assertTrue(login_user.is_staff)
        self.assertEqual(legacy_profile.role, "Tutee")
        self.assertEqual(login_profile.role, "SuperAdmin")
        self.assertIn(f"Promoting login username match user_id={login_user.id}.", output.getvalue())

    def test_staff_user_without_profile_gets_admin_profile_status(self):
        admin_user = User.objects.create_user(
            username="status-admin@example.edu",
            email="status-admin@example.edu",
            password="password",
            is_staff=True,
        )
        self.client.force_authenticate(user=admin_user)

        response = self.client.get("/api/profile/status/")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["profile_completed"])
        self.assertEqual(response.data["role"], "SuperAdmin")

        profile = UserProfile.objects.get(user=admin_user)
        self.assertEqual(profile.role, "SuperAdmin")
        self.assertTrue(profile.is_domain_exempt)

    @patch("studybuddy.views.generate_otp_code", side_effect=["123456", "654321"])
    def test_resend_replaces_otp_code_and_increments_resend_count(self, _generate_otp_code):
        login_response = self.login()
        old_code = self.latest_otp_code()

        resend_response = self.client.post(
            "/api/login/resend-otp/",
            {"challenge_id": login_response.data["challenge_id"]},
            format="json",
        )

        self.assertEqual(resend_response.status_code, 200)
        self.assertEqual(len(mail.outbox), 2)
        challenge = EmailOTPChallenge.objects.get(
            challenge_id=login_response.data["challenge_id"]
        )
        self.assertEqual(challenge.resend_count, 1)
        self.assertEqual(challenge.attempt_count, 0)

        old_code_response = self.client.post(
            "/api/login/verify-otp/",
            {
                "challenge_id": login_response.data["challenge_id"],
                "code": old_code,
            },
            format="json",
        )
        self.assertEqual(old_code_response.status_code, 400)

        new_code_response = self.client.post(
            "/api/login/verify-otp/",
            {
                "challenge_id": login_response.data["challenge_id"],
                "code": self.latest_otp_code(),
            },
            format="json",
        )
        self.assertEqual(new_code_response.status_code, 200)

    def test_otp_verify_locks_after_max_attempts(self):
        login_response = self.login()
        actual_code = self.latest_otp_code()
        first_wrong_code = "000000" if actual_code != "000000" else "000001"
        second_wrong_code = "111111" if actual_code != "111111" else "111112"

        first_response = self.client.post(
            "/api/login/verify-otp/",
            {
                "challenge_id": login_response.data["challenge_id"],
                "code": first_wrong_code,
            },
            format="json",
        )
        second_response = self.client.post(
            "/api/login/verify-otp/",
            {
                "challenge_id": login_response.data["challenge_id"],
                "code": second_wrong_code,
            },
            format="json",
        )

        self.assertEqual(first_response.status_code, 400)
        self.assertEqual(second_response.status_code, 429)
        challenge = EmailOTPChallenge.objects.get(
            challenge_id=login_response.data["challenge_id"]
        )
        self.assertEqual(challenge.attempt_count, 2)

    def test_password_reset_request_is_generic_and_confirm_resets_password(self):
        RefreshToken.for_user(self.user)
        self.assertTrue(OutstandingToken.objects.filter(user=self.user).exists())

        known_response = self.client.post(
            "/api/password-reset/request/",
            {"email": self.user.email},
            format="json",
        )
        unknown_response = self.client.post(
            "/api/password-reset/request/",
            {"email": "missing@example.edu"},
            format="json",
        )

        self.assertEqual(known_response.status_code, 200)
        self.assertEqual(unknown_response.status_code, 200)
        self.assertEqual(known_response.data, unknown_response.data)
        self.assertEqual(len(mail.outbox), 1)

        uid, token = self.reset_link_parts()
        confirm_response = self.client.post(
            "/api/password-reset/confirm/",
            {
                "uid": uid,
                "token": token,
                "password": "NewStudyBuddyPassword123!",
                "password_confirm": "NewStudyBuddyPassword123!",
            },
            format="json",
        )

        self.assertEqual(confirm_response.status_code, 200)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password("NewStudyBuddyPassword123!"))
        self.assertEqual(
            BlacklistedToken.objects.filter(token__user=self.user).count(),
            OutstandingToken.objects.filter(user=self.user).count(),
        )
        self.assertEqual(len(mail.outbox), 2)

    def test_password_reset_confirm_validates_password_confirmation(self):
        self.client.post(
            "/api/password-reset/request/",
            {"email": self.user.email},
            format="json",
        )
        uid, token = self.reset_link_parts()

        response = self.client.post(
            "/api/password-reset/confirm/",
            {
                "uid": uid,
                "token": token,
                "password": "NewStudyBuddyPassword123!",
                "password_confirm": "DifferentStudyBuddyPassword123!",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.user.refresh_from_db()
        self.assertFalse(self.user.check_password("NewStudyBuddyPassword123!"))


class RecommenderNeighborReuseTests(APITestCase):
    def setUp(self):
        # student 1 is our target; students 2 and 3 are potential neighbors.
        self.ratings = {
            1: {10: 5, 11: 4},
            2: {10: 4, 11: 5, 12: 3},
            3: {10: 2, 11: 1, 12: 5},
        }

    def test_compute_cf_score_uses_supplied_neighbors(self):
        from studybuddy.recommender import CF

        neighbors = CF.top_k(self.ratings, 1)

        with patch.object(CF, "top_k") as mocked_top_k:
            score = CF.compute_cf_score(
                self.ratings, 1, 12, neighbors=neighbors
            )

        mocked_top_k.assert_not_called()
        self.assertIsNotNone(score)

    def test_compute_cf_score_matches_with_and_without_neighbors(self):
        from studybuddy.recommender import CF

        neighbors = CF.top_k(self.ratings, 1)

        without = CF.compute_cf_score(self.ratings, 1, 12)
        with_neighbors = CF.compute_cf_score(self.ratings, 1, 12, neighbors=neighbors)

        self.assertEqual(without, with_neighbors)

    def test_recommend_hybrid_computes_both_neighbor_lists_once(self):
        from studybuddy.recommender import hybrid

        # three candidate tutors, but neighbors should be computed only once
        tutors = [Mock(profile_id=i, tutorsubjects_set=Mock()) for i in range(3)]
        for t in tutors:
            t.tutorsubjects_set.all.return_value = []

        student_profile = Mock(id=1, course_id=None, year_level=None)

        with patch.object(hybrid, "top_k", return_value=[]) as mocked_top_k, \
             patch.object(hybrid, "get_student_subject_codes", return_value=[]), \
             patch.object(hybrid, "compute_cbf_score", return_value=0.0), \
             patch.object(hybrid, "normalize_tutor_queryset", return_value=tutors):
            hybrid.recommend_tutors_hybrid(self.ratings, student_profile, None)

        self.assertEqual(mocked_top_k.call_count, 2)

    def test_recommend_hybrid_handles_student_with_no_ratings(self):
        from studybuddy.recommender import hybrid

        tutor = Mock(profile_id=99, tutorsubjects_set=Mock())
        tutor.tutorsubjects_set.all.return_value = []
        student_profile = Mock(id=4242, course_id=None, year_level=None)  # not in ratings

        with patch.object(hybrid, "get_student_subject_codes", return_value=[]), \
             patch.object(hybrid, "compute_cbf_score", return_value=0.5), \
             patch.object(hybrid, "normalize_tutor_queryset", return_value=[tutor]):
            results = hybrid.recommend_tutors_hybrid(self.ratings, student_profile, None)

        self.assertEqual(len(results), 1)  # did not raise


class CfPeerNeighborTests(APITestCase):
    def setUp(self):
        self.course = Course.objects.create(
            course_code="BSCS", course_name="Computer Science"
        )
        self.other_course = Course.objects.create(
            course_code="BSIT", course_name="Information Technology"
        )
        self.tutee = self._make_tutee("target", self.course)
        self.peer = self._make_tutee("peer", self.course)
        self.other_tutee = self._make_tutee("other", self.other_course)
        self.no_course_tutee = self._make_tutee("no-course", None)

    def _make_tutee(self, username, course):
        user = User.objects.create_user(
            username=f"{username}@cpu.edu", email=f"{username}@cpu.edu", password="password"
        )
        return UserProfile.objects.create(
            user=user, fname=username, mname="", lname="Tutee", role="Tutee", course=course
        )

    def test_top_k_excludes_zero_and_negative_similarity_neighbors(self):
        from studybuddy.recommender.CF import top_k

        ratings = {
            1: {10: 5, 11: 1},
            2: {10: 4, 11: 2},
            3: {10: 1, 11: 5},
            4: {10: 3, 11: 3},
        }

        neighbors = top_k(ratings, 1)
        self.assertEqual([student_id for student_id, _ in neighbors], [2])
        self.assertGreater(neighbors[0][1], 0)

    def test_peer_ids_include_only_same_course_and_null_course_is_empty(self):
        from studybuddy.recommender.CF import get_peer_student_ids

        ratings = {
            self.tutee.id: {10: 5, 11: 1},
            self.peer.id: {10: 4, 11: 2},
            self.other_tutee.id: {10: 4, 11: 2},
        }

        self.assertEqual(get_peer_student_ids(ratings, self.tutee), [self.peer.id])
        self.assertEqual(get_peer_student_ids(ratings, self.no_course_tutee), [])

    def test_peer_prediction_falls_back_to_global_per_tutor(self):
        from studybuddy.recommender.CF import compute_cf_breakdown_with_fallback

        ratings = {
            1: {10: 5, 11: 1},
            2: {10: 4, 11: 2, 12: 5},
            3: {10: 4, 11: 2, 13: 1},
        }
        peer_neighbors = [(2, 1.0)]
        global_neighbors = [(3, 1.0)]

        peer_breakdown = compute_cf_breakdown_with_fallback(
            ratings, 1, 12, peer_neighbors, global_neighbors
        )
        global_breakdown = compute_cf_breakdown_with_fallback(
            ratings, 1, 13, peer_neighbors, global_neighbors
        )
        empty_peer_breakdown = compute_cf_breakdown_with_fallback(
            ratings, 1, 13, [], global_neighbors
        )

        self.assertEqual(peer_breakdown["pool"], "peer")
        self.assertEqual(peer_breakdown["neighbors"][0]["neighbor_id"], 2)
        self.assertEqual(global_breakdown["pool"], "global")
        self.assertEqual(global_breakdown["neighbors"][0]["neighbor_id"], 3)
        self.assertEqual(empty_peer_breakdown["pool"], "global")


class CoRatedDetailTests(APITestCase):
    """co_rated_detail feeds the algorithm demo's expandable co-rated panel, which
    exists so a similarity can be checked by hand rather than taken on faith. These
    lock the curated S6/T-cast numbers documented in the algorithm demo guide."""

    # Tutor ids 1/2/6/7/8 stand in for the curated T1/T2/T6/T7/T8, tutee keys for
    # S6 and her neighbours — the same matrix as seed_data.CURATED_RATINGS.
    RATINGS = {
        6: {1: 5, 2: 3, 6: 5, 7: 2, 8: 4},   # S6 Katrina, overall avg 3.80
        2: {1: 5, 2: 3, 6: 4, 8: 4},         # S2 Gloria
        5: {3: 4, 4: 5},                     # shares nothing with S6
    }

    def test_returns_shared_tutors_with_both_students_scores(self):
        from studybuddy.recommender.CF import co_rated_detail

        detail = co_rated_detail(self.RATINGS, 6, 2)

        self.assertEqual([e["tutor_id"] for e in detail["shared"]], [1, 2, 6, 8])
        self.assertEqual([e["u_rating"] for e in detail["shared"]], [5, 3, 5, 4])
        self.assertEqual([e["v_rating"] for e in detail["shared"]], [5, 3, 4, 4])

    def test_averages_are_over_the_co_rated_set_not_all_ratings(self):
        """The trap the panel has to survive: the average Pearson uses is not the
        average the deviation term uses. S6's mean over the four tutors she shares
        with S2 is 4.25, while her CF baseline over all five of her ratings is
        3.80 — a panelist adding up the visible row must not land on 3.80."""
        from studybuddy.recommender.CF import co_rated_detail

        detail = co_rated_detail(self.RATINGS, 6, 2)
        overall_avg = sum(self.RATINGS[6].values()) / len(self.RATINGS[6])

        self.assertAlmostEqual(detail["u_avg"], 4.25)
        self.assertAlmostEqual(detail["v_avg"], 4.0)
        self.assertAlmostEqual(overall_avg, 3.8)
        self.assertNotAlmostEqual(detail["u_avg"], overall_avg)

    def test_shared_ratings_reproduce_the_reported_similarity(self):
        """The panel's whole claim is that the expanded cells derive the number on
        the collapsed row, so recomputing Pearson from those cells must agree with
        sim()."""
        from studybuddy.recommender.CF import co_rated_detail, sim

        detail = co_rated_detail(self.RATINGS, 6, 2)
        u_dev = [e["u_rating"] - detail["u_avg"] for e in detail["shared"]]
        v_dev = [e["v_rating"] - detail["v_avg"] for e in detail["shared"]]

        numerator = sum(a * b for a, b in zip(u_dev, v_dev))
        denominator = (
            math.sqrt(sum(a * a for a in u_dev)) * math.sqrt(sum(b * b for b in v_dev))
        )

        self.assertAlmostEqual(numerator / denominator, sim(self.RATINGS, 6, 2))
        self.assertAlmostEqual(sim(self.RATINGS, 6, 2), 0.853, places=3)

    def test_no_overlap_returns_empty_set_and_no_averages(self):
        from studybuddy.recommender.CF import co_rated_detail

        detail = co_rated_detail(self.RATINGS, 6, 5)

        self.assertEqual(detail["shared"], [])
        self.assertIsNone(detail["u_avg"])
        self.assertIsNone(detail["v_avg"])

    def test_unknown_student_is_tolerated(self):
        """Cold-Start tutees are absent from the rating matrix entirely."""
        from studybuddy.recommender.CF import co_rated_detail

        self.assertEqual(co_rated_detail(self.RATINGS, 999, 2)["shared"], [])


class CbfGraduatedSubjectMatchTests(APITestCase):
    """CBF graduated subject matching: a tutee requesting a specific subject
    sees exact-match tutors ranked above same-field tutors, ranked above
    unrelated tutors, instead of the old all-or-nothing subject match. See
    docs/plans/ for the approved weights (W_SPECIFIC=0.40, W_GENERAL=0.20,
    W_EXPERTISE=0.15, W_COURSE=0.10, W_YEAR=0.10, W_LEVEL=0.05)."""

    def setUp(self):
        from studybuddy.recommender.cbf import compute_cbf_breakdown

        self.compute_cbf_breakdown = compute_cbf_breakdown

        self.course = Course.objects.create(course_code="BSCS", course_name="Computer Science")

        self.requested_subject = Subjects.objects.create(
            subject_code="MATH101", subject_name="Calculus 1", department="Math",
            category="STEM",
        )
        self.same_field_subject = Subjects.objects.create(
            subject_code="MATH201", subject_name="Calculus 2", department="Math",
            category="STEM",
        )
        self.other_same_field_subject = Subjects.objects.create(
            subject_code="PHYS101", subject_name="Physics 1", department="Science",
            category="STEM",
        )
        self.unrelated_subject = Subjects.objects.create(
            subject_code="ART101", subject_name="Painting", department="Arts",
            category="Arts",
        )
        self.null_category_subject = Subjects.objects.create(
            subject_code="GEN101", subject_name="General Studies", department="Gen",
        )
        # Requested subject itself has a null category, to test the superset
        # rule (General must still count an exact match even without a category).
        self.null_category_requested_subject = Subjects.objects.create(
            subject_code="GEN201", subject_name="Study Skills", department="Gen",
        )

        student_user = User.objects.create_user(
            username="cbf-student", email="cbf-student@example.com", password="password",
        )
        self.student = UserProfile.objects.create(
            user=student_user, fname="Cbf", mname="", lname="Student", role="Tutee",
            year_level=12, course=self.course,
        )

    def _make_tutor(self, username, subjects_with_levels, teaching_level="College", year_level=12, course=None):
        user = User.objects.create_user(
            username=username, email=f"{username}@example.com", password="password",
        )
        profile = UserProfile.objects.create(
            user=user, fname=username.title(), mname="", lname="Tutor", role="Tutor",
            year_level=year_level, course=course,
        )
        tutor = Tutor.objects.create(
            profile=profile, hourly_rate=200, can_online=True, can_f2f=False,
            teaching_level=teaching_level,
        )
        for subject, level in subjects_with_levels:
            TutorSubjects.objects.create(tutor=tutor, subject=subject, expertise_level=level)
        return tutor

    def test_level_ceiling_penalizes_only_students_above_known_ceiling(self):
        cases = [
            ('High School', 14, 0),
            ('Elementary', 8, 0),
            ('College', 11, 1),
            ('', 14, 1),
        ]
        for index, (level, student_year, expected) in enumerate(cases):
            self.student.year_level = student_year
            self.student.save(update_fields=['year_level'])
            tutor = self._make_tutor(
                f'level-{index}', [(self.requested_subject, 3)], teaching_level=level,
            )
            breakdown = self.compute_cbf_breakdown(
                self.student, tutor, self.requested_subject.subject_code,
            )
            self.assertEqual(breakdown['level']['value'], expected)

    def test_dominance_exact_match_outranks_field_only_even_at_worst_case(self):
        # Exact-match tutor at minimum possible expertise (1) still must bank
        # >= 0.63 on the subject block alone.
        exact_match_tutor = self._make_tutor(
            "exact-match", [(self.requested_subject, 1)],
            course=self.course, year_level=self.student.year_level,
        )
        # Field-only tutor at maximum possible expertise (5) plus perfect
        # course/year/level must still cap out at 0.60 overall.
        field_only_tutor = self._make_tutor(
            "field-only", [(self.same_field_subject, 5)],
            course=self.course, year_level=self.student.year_level,
        )

        exact_breakdown = self.compute_cbf_breakdown(
            self.student, exact_match_tutor, self.requested_subject.subject_code,
        )
        field_breakdown = self.compute_cbf_breakdown(
            self.student, field_only_tutor, self.requested_subject.subject_code,
        )

        exact_subject_block = (
            exact_breakdown["specific"]["contribution"]
            + exact_breakdown["general"]["contribution"]
            + exact_breakdown["expertise"]["contribution"]
        )
        self.assertGreaterEqual(exact_subject_block, 0.63)
        self.assertLessEqual(field_breakdown["score"], 0.60)
        self.assertGreater(exact_breakdown["score"], field_breakdown["score"])

    def test_null_category_tutor_subjects_contribute_zero_no_exception(self):
        tutor = self._make_tutor("null-cat", [(self.null_category_subject, 5)])

        breakdown = self.compute_cbf_breakdown(
            self.student, tutor, self.requested_subject.subject_code,
        )

        self.assertEqual(breakdown["specific"]["value"], 0)
        self.assertEqual(breakdown["general"]["value"], 0)
        self.assertEqual(breakdown["expertise"]["value"], 0)

    def test_exact_match_with_null_category_requested_subject_still_scores_general(self):
        tutor = self._make_tutor(
            "null-req-exact", [(self.null_category_requested_subject, 3)],
        )

        breakdown = self.compute_cbf_breakdown(
            self.student, tutor, self.null_category_requested_subject.subject_code,
        )

        self.assertEqual(breakdown["specific"]["value"], 1)
        self.assertEqual(breakdown["general"]["value"], 1)

    def test_expertise_cascade_exact_match_uses_that_row_level(self):
        tutor = self._make_tutor("exact-expertise", [(self.requested_subject, 3)])

        breakdown = self.compute_cbf_breakdown(
            self.student, tutor, self.requested_subject.subject_code,
        )

        self.assertAlmostEqual(breakdown["expertise"]["value"], 3 / 5)

    def test_expertise_cascade_field_only_uses_mean_of_same_field_levels(self):
        tutor = self._make_tutor(
            "field-expertise",
            [(self.same_field_subject, 2), (self.other_same_field_subject, 4)],
        )

        breakdown = self.compute_cbf_breakdown(
            self.student, tutor, self.requested_subject.subject_code,
        )

        self.assertAlmostEqual(breakdown["expertise"]["value"], (2 + 4) / 2 / 5)

    def test_expertise_cascade_unrelated_tutor_is_zero(self):
        tutor = self._make_tutor("unrelated-expertise", [(self.unrelated_subject, 5)])

        breakdown = self.compute_cbf_breakdown(
            self.student, tutor, self.requested_subject.subject_code,
        )

        self.assertEqual(breakdown["expertise"]["value"], 0)

    def test_empty_requested_subject_falls_back_to_preference_list(self):
        tutor = self._make_tutor("fallback-match", [(self.requested_subject, 4)])

        breakdown = self.compute_cbf_breakdown(
            self.student,
            tutor,
            None,
            student_subjects=[self.requested_subject.subject_code],
        )

        self.assertEqual(breakdown["specific"]["value"], 1)
        self.assertEqual(breakdown["general"]["value"], 1)
        self.assertAlmostEqual(breakdown["expertise"]["value"], 4 / 5)


class DashboardRecommendationServiceTests(APITestCase):
    def setUp(self):
        cache.clear()
        self.subject = Subjects.objects.create(
            subject_code="IT101",
            subject_name="Intro to IT",
            department="IT",
        )
        self.other_subject = Subjects.objects.create(
            subject_code="BIO101",
            subject_name="Biology",
            department="Science",
        )
        self.student_user = User.objects.create_user(
            username="dash-student", email="dash@example.com", password="password",
        )
        self.student = UserProfile.objects.create(
            user=self.student_user, fname="Dash", mname="", lname="Student",
            role="Tutee", year_level=11,
        )

    def _make_tutor(self, username, subject):
        user = User.objects.create_user(
            username=username, email=f"{username}@example.com", password="password",
        )
        profile = UserProfile.objects.create(
            user=user, fname=username.title(), mname="", lname="Tutor",
            role="Tutor", year_level=12,
        )
        tutor = Tutor.objects.create(
            profile=profile, hourly_rate=200, can_online=True, can_f2f=False,
            teaching_level="SHS",
        )
        TutorSubjects.objects.create(tutor=tutor, subject=subject, expertise_level=5)
        return tutor

    def _set_preferences(self, *subjects):
        pref, _ = Preference.objects.get_or_create(user=self.student)
        pref.subjects.set([s.subject_code for s in subjects])

    def test_returns_only_tutors_teaching_preference_subjects(self):
        from studybuddy.recommender.dashboard import get_dashboard_recommendations

        match = self._make_tutor("itmatch", self.subject)
        self._make_tutor("biomatch", self.other_subject)
        self._set_preferences(self.subject)

        data = get_dashboard_recommendations(self.student)

        ids = {row["id"] for row in data}
        self.assertEqual(ids, {match.profile.id})

    def test_response_shape_matches_widget_contract(self):
        from studybuddy.recommender.dashboard import get_dashboard_recommendations

        self._make_tutor("itmatch", self.subject)
        self._set_preferences(self.subject)

        row = get_dashboard_recommendations(self.student)[0]

        self.assertEqual(
            set(row.keys()), {"id", "name", "rating", "subjects", "hourlyRate"},
        )

    def test_cold_start_returns_fallback_when_no_preferences(self):
        from studybuddy.recommender.dashboard import get_dashboard_recommendations

        for index in range(7):
            self._make_tutor(f"anytutor{index}", self.subject)
        # no preferences set

        data = get_dashboard_recommendations(self.student)

        self.assertEqual(len(data), 5)

    def test_returns_top_five_from_hybrid_algorithm_order(self):
        from studybuddy.recommender import dashboard

        tutors = [self._make_tutor(f"itmatch{index}", self.subject) for index in range(7)]
        self._set_preferences(self.subject)

        ranked = [
            {"tutor": tutor, "score": 1 - (index / 10)}
            for index, tutor in enumerate(reversed(tutors))
        ]

        with patch.object(dashboard, "recommend_tutors_hybrid", return_value=ranked):
            data = dashboard.get_dashboard_recommendations(self.student)

        self.assertEqual(len(data), 5)
        self.assertEqual(
            [row["id"] for row in data],
            [recommendation["tutor"].profile.id for recommendation in ranked[:5]],
        )

    def test_cache_hit_is_defensively_limited_to_five(self):
        from studybuddy.recommender.dashboard import (
            dashboard_recs_cache_key,
            get_dashboard_recommendations,
        )

        cache.set(
            dashboard_recs_cache_key(self.student),
            [
                {
                    "id": index,
                    "name": f"Cached Tutor {index}",
                    "rating": 5,
                    "subjects": ["Intro to IT"],
                    "hourlyRate": 200,
                }
                for index in range(8)
            ],
        )

        data = get_dashboard_recommendations(self.student)

        self.assertEqual(len(data), 5)
        self.assertEqual([row["id"] for row in data], [0, 1, 2, 3, 4])

    def test_second_call_served_from_cache_without_recompute(self):
        from studybuddy.recommender import dashboard

        self._make_tutor("itmatch", self.subject)
        self._set_preferences(self.subject)

        dashboard.get_dashboard_recommendations(self.student)  # warms cache

        with patch.object(dashboard, "recommend_tutors_hybrid") as mocked:
            dashboard.get_dashboard_recommendations(self.student)

        mocked.assert_not_called()

    def test_degrades_gracefully_when_cache_read_fails(self):
        from studybuddy.recommender import dashboard

        match = self._make_tutor("itmatch", self.subject)
        self._set_preferences(self.subject)

        with patch.object(dashboard.cache, "get", side_effect=Exception("redis down")):
            data = dashboard.get_dashboard_recommendations(self.student)

        ids = {row["id"] for row in data}
        self.assertEqual(ids, {match.profile.id})


class StudentDashboardRecommendationTests(APITestCase):
    def setUp(self):
        cache.clear()
        self.subject = Subjects.objects.create(
            subject_code="IT201", subject_name="Data Structures", department="IT",
        )
        self.other_subject = Subjects.objects.create(
            subject_code="HIST201", subject_name="History", department="Arts",
        )
        self.student_user = User.objects.create_user(
            username="dv-student", email="dv@example.com", password="password",
        )
        self.student = UserProfile.objects.create(
            user=self.student_user, fname="Dee", mname="", lname="Vee",
            role="Tutee", year_level=11,
        )
        self.client.force_authenticate(user=self.student_user)

    def _make_tutor(self, username, subject):
        user = User.objects.create_user(
            username=username, email=f"{username}@example.com", password="password",
        )
        profile = UserProfile.objects.create(
            user=user, fname=username.title(), mname="", lname="Tutor",
            role="Tutor", year_level=12,
        )
        tutor = Tutor.objects.create(
            profile=profile, hourly_rate=200, can_online=True, can_f2f=False,
            teaching_level="SHS",
        )
        TutorSubjects.objects.create(tutor=tutor, subject=subject, expertise_level=5)
        return tutor

    def test_dashboard_recommends_subject_matched_tutors(self):
        matches = [self._make_tutor(f"itone{index}", self.subject) for index in range(7)]
        self._make_tutor("histone", self.other_subject)
        pref, _ = Preference.objects.get_or_create(user=self.student)
        pref.subjects.set([self.subject.subject_code])

        response = self.client.get("/api/dashboard/")

        self.assertEqual(response.status_code, 200)
        ids = {row["id"] for row in response.data["recommendations"]}
        self.assertEqual(len(response.data["recommendations"]), 5)
        self.assertTrue(ids.issubset({tutor.profile.id for tutor in matches}))
        self.assertEqual(
            set(response.data["recommendations"][0].keys()),
            {"id", "name", "rating", "subjects", "hourlyRate"},
        )

    def test_saving_preferences_busts_dashboard_cache(self):
        from studybuddy.recommender.dashboard import dashboard_recs_cache_key

        self._make_tutor("itone", self.subject)
        pref, _ = Preference.objects.get_or_create(user=self.student)
        pref.subjects.set([self.subject.subject_code])

        # warm the cache
        self.client.get("/api/dashboard/")
        self.assertIsNotNone(cache.get(dashboard_recs_cache_key(self.student)))

        # changing preferences must clear it
        response = self.client.post(
            "/api/preferences/", {"subjects": [self.subject.subject_code]},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIsNone(cache.get(dashboard_recs_cache_key(self.student)))

    def test_submitting_rating_bumps_dashboard_cache_version(self):
        from studybuddy.recommender.dashboard import dashboard_recs_cache_key

        tutor = self._make_tutor("ratedtutor", self.subject)
        availability = TutorAvailability.objects.create(
            tutor=tutor,
            day="Mon",
            time_slot=time(9, 0),
            is_active=True,
        )
        booking = Booking.objects.create(
            student=self.student,
            tutor=tutor,
            availability=availability,
            session_date=date(2026, 6, 1),
            session_mode="Online",
            booking_request_id=uuid4(),
            status="Completed",
        )

        old_key = dashboard_recs_cache_key(self.student)

        response = self.client.post(
            f"/api/bookings/{booking.id}/rating/",
            {"rating_score": 5},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertNotEqual(old_key, dashboard_recs_cache_key(self.student))

    def test_tutor_subject_change_bumps_dashboard_cache_version(self):
        from studybuddy.recommender.dashboard import dashboard_recs_cache_key

        tutor = self._make_tutor("subjecttutor", self.subject)
        new_subject = Subjects.objects.create(
            subject_code="IT301",
            subject_name="Algorithms",
            department="IT",
        )

        old_key = dashboard_recs_cache_key(self.student)
        self.client.force_authenticate(user=tutor.profile.user)

        response = self.client.post(
            "/api/tutor/subjects/add/",
            {"subject_code": new_subject.subject_code},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotEqual(old_key, dashboard_recs_cache_key(self.student))

    def test_tutor_profile_change_bumps_dashboard_cache_version(self):
        from studybuddy.recommender.dashboard import dashboard_recs_cache_key

        tutor = self._make_tutor("profiletutor", self.subject)

        old_key = dashboard_recs_cache_key(self.student)
        self.client.force_authenticate(user=tutor.profile.user)

        response = self.client.put(
            "/api/tutor/update/",
            {
                "hourly_rate": 250,
                "teaching_level": "SHS",
                "can_online": True,
                "can_f2f": False,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotEqual(old_key, dashboard_recs_cache_key(self.student))


class DashboardLoadPerformanceTests(APITestCase):
    """Phase 1 verification: the dedicated recommendations endpoint plus
    query-count regression guards for the N+1 fixes in list_bookings and
    student_dashboard. The N+1 guard asserts the query count is constant as the
    number of bookings grows — if a per-row query crept back in, doubling the
    bookings would raise the count."""

    def setUp(self):
        cache.clear()
        self.course = Course.objects.create(
            course_code="BSIT", course_name="BS Information Technology",
        )
        self.subject = Subjects.objects.create(
            subject_code="PERF101", subject_name="Performance", department="IT",
        )
        self.student_user = User.objects.create_user(
            username="perf-student", email="perf-student@example.com", password="password",
        )
        self.student = UserProfile.objects.create(
            user=self.student_user, fname="Perf", mname="", lname="Student",
            role="Tutee", year_level=11, course=self.course,
        )
        self.tutor_user = User.objects.create_user(
            username="perf-tutor", email="perf-tutor@example.com", password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Perf", mname="", lname="Tutor",
            role="Tutor", year_level=12, course=self.course,
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile, hourly_rate=200, can_online=True,
            can_f2f=True, teaching_level="SHS",
        )
        TutorSubjects.objects.create(
            tutor=self.tutor, subject=self.subject, expertise_level=5,
        )
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor, day="Mon", time_slot=time(14, 0), is_active=True,
        )
        self.client.force_authenticate(user=self.student_user)

    def _make_completed(self, count, start_day):
        """Create `count` Completed bookings on distinct dates, each rated.

        Distinct session_dates avoid the (availability, session_date) unique
        constraint. Rating rows are created directly so they exercise the
        select_related('rating') path without bumping the recs cache version.
        """
        from datetime import timedelta
        base = date(2026, 1, 1) + timedelta(days=start_day)
        for offset in range(count):
            booking = Booking.objects.create(
                student=self.student, tutor=self.tutor, availability=self.availability,
                session_date=base + timedelta(days=offset),
                session_mode="Online", status="Completed",
            )
            Rating.objects.create(
                booking=booking, student=self.student, tutor=self.tutor, rating_score=5,
            )

    def _count_queries(self, url):
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        return len(ctx)

    def test_recommendations_endpoint_returns_only_recommendations(self):
        response = self.client.get("/api/recommendations/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("recommendations", response.data)
        self.assertNotIn("upcoming", response.data)
        self.assertNotIn("completed", response.data)

    def test_recommendations_endpoint_requires_auth(self):
        self.client.force_authenticate(user=None)
        response = self.client.get("/api/recommendations/")
        self.assertIn(response.status_code, (401, 403))

    def test_student_dashboard_has_no_n_plus_one(self):
        self._make_completed(3, start_day=0)
        self.client.get("/api/dashboard/")  # warm the (cached) recommendations
        first = self._count_queries("/api/dashboard/")
        self._make_completed(3, start_day=400)
        second = self._count_queries("/api/dashboard/")
        self.assertEqual(
            first, second,
            f"student_dashboard query count grew with booking count "
            f"({first} -> {second}); a per-row (N+1) query likely regressed.",
        )

    def test_list_bookings_has_no_n_plus_one(self):
        self._make_completed(3, start_day=0)
        first = self._count_queries("/api/bookings/")
        self._make_completed(3, start_day=400)
        second = self._count_queries("/api/bookings/")
        self.assertEqual(
            first, second,
            f"list_bookings query count grew with booking count "
            f"({first} -> {second}); a per-row (N+1) query likely regressed.",
        )

    def test_list_bookings_does_not_merge_same_group_across_dates(self):
        group_id = uuid4()
        Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 6, 6),
            session_mode="Online",
            session_group_id=group_id,
            status="Rejected",
        )
        Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 6, 10),
            session_mode="Online",
            session_group_id=group_id,
            status="Rejected",
        )

        response = self.client.get("/api/bookings/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)
        self.assertEqual(
            [str(booking["date"]) for booking in response.data],
            ["2026-06-06", "2026-06-10"],
        )

    def test_list_bookings_uses_booked_subject_for_combined_blocks(self):
        booking = Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 6, 6),
            session_mode="Online",
            status="Completed",
            subject=self.subject,
        )
        Rating.objects.create(
            booking=booking,
            student=self.student,
            tutor=self.tutor,
            rating_score=5,
        )

        response = self.client.get("/api/bookings/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data[0]["subject"], self.subject.subject_name)

    def test_list_bookings_uses_booked_subject_for_pending_request_blocks(self):
        Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 6, 6),
            session_mode="Online",
            status="Pending",
            subject=self.subject,
        )

        response = self.client.get("/api/bookings/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data[0]["subject"], self.subject.subject_name)

    def test_student_can_hide_rejected_dashboard_pill_without_hiding_for_tutor(self):
        booking = Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 6, 6),
            session_mode="Online",
            status="Rejected",
        )

        response = self.client.delete(f"/api/bookings/{booking.id}/dashboard-pill/")

        self.assertEqual(response.status_code, 200)
        booking.refresh_from_db()
        self.assertIsNotNone(booking.dashboard_hidden_by_student_at)
        self.assertIsNone(booking.dashboard_hidden_by_tutor_at)

        student_response = self.client.get("/api/bookings/")
        self.assertTrue(student_response.data[0]["dashboard_hidden_by_current_user"])

        self.client.force_authenticate(user=self.tutor_user)
        tutor_response = self.client.get("/api/bookings/")
        self.assertFalse(tutor_response.data[0]["dashboard_hidden_by_current_user"])

    def test_hide_dashboard_pill_applies_to_all_slots_in_group(self):
        group_id = uuid4()
        next_availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=time(14, 30),
            is_active=True,
        )
        first_booking = Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 6, 6),
            session_mode="Online",
            session_group_id=group_id,
            status="Cancelled",
        )
        second_booking = Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=next_availability,
            session_date=date(2026, 6, 6),
            session_mode="Online",
            session_group_id=group_id,
            status="Cancelled",
        )

        response = self.client.delete(f"/api/bookings/{first_booking.id}/dashboard-pill/")

        self.assertEqual(response.status_code, 200)
        first_booking.refresh_from_db()
        second_booking.refresh_from_db()
        self.assertIsNotNone(first_booking.dashboard_hidden_by_student_at)
        self.assertIsNotNone(second_booking.dashboard_hidden_by_student_at)
        self.assertEqual(
            sorted(response.data["hidden_booking_ids"]),
            sorted([first_booking.id, second_booking.id]),
        )

    def test_hide_dashboard_pill_rejects_active_or_completed_statuses(self):
        booking = Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 6, 6),
            session_mode="Online",
            status="Completed",
        )

        response = self.client.delete(f"/api/bookings/{booking.id}/dashboard-pill/")

        self.assertEqual(response.status_code, 400)
        booking.refresh_from_db()
        self.assertIsNone(booking.dashboard_hidden_by_student_at)

    def test_hide_dashboard_pill_rejects_unauthorized_users(self):
        booking = Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date(2026, 6, 6),
            session_mode="Online",
            status="Cancelled",
        )
        other_user = User.objects.create_user(
            username="other-student",
            email="other-student@example.com",
            password="password",
        )
        UserProfile.objects.create(
            user=other_user,
            fname="Other",
            mname="",
            lname="Student",
            role="Tutee",
            year_level=11,
            course=self.course,
        )
        self.client.force_authenticate(user=other_user)

        response = self.client.delete(f"/api/bookings/{booking.id}/dashboard-pill/")

        self.assertEqual(response.status_code, 403)
        booking.refresh_from_db()
        self.assertIsNone(booking.dashboard_hidden_by_student_at)


@override_settings(ALGORITHM_DEMO_TOOLS_ENABLED=True)
class AlgorithmDemoToolTests(APITestCase):
    """Staff-only recommendation algorithm demo endpoints — see
    docs/plans/2026-07-04-recommendation-algorithm-demo-tool.md."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu", is_active=True,
        )
        self.subject = Subjects.objects.create(
            subject_code="IT201", subject_name="Data Structures", department="IT",
        )

        self.super_user = User.objects.create_user(
            username="algo-super@cpu.edu", email="algo-super@cpu.edu", password="password",
        )
        UserProfile.objects.create(
            user=self.super_user, fname="Super", mname="", lname="Admin", role="SuperAdmin",
            institution=self.institution,
        )

        self.admin_user = User.objects.create_user(
            username="algo-admin@cpu.edu", email="algo-admin@cpu.edu", password="password",
        )
        UserProfile.objects.create(
            user=self.admin_user, fname="Regular", mname="", lname="Admin", role="Admin",
            institution=self.institution,
        )

        self.tutee_user = User.objects.create_user(
            username="algo-tutee@cpu.edu", email="algo-tutee@cpu.edu", password="password",
        )
        self.tutee = UserProfile.objects.create(
            user=self.tutee_user, fname="Maria", mname="", lname="Santos", role="Tutee",
            year_level=11, institution=self.institution,
        )
        pref, _ = Preference.objects.get_or_create(user=self.tutee)
        pref.subjects.set([self.subject])

        self.tutor_user = User.objects.create_user(
            username="algo-tutor@cpu.edu", email="algo-tutor@cpu.edu", password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Carlo", mname="", lname="Reyes", role="Tutor",
            year_level=12, institution=self.institution,
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile, hourly_rate=200, can_online=True, can_f2f=False,
            teaching_level="College",
        )
        TutorSubjects.objects.create(tutor=self.tutor, subject=self.subject, expertise_level=4)
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor, day="Mon", time_slot=time(9, 0),
        )

    def _create_rating(self, student=None, rating_score=3, session_date=None, tutor=None):
        student = student or self.tutee
        tutor = tutor or self.tutor
        availability = (
            self.availability if tutor == self.tutor
            else TutorAvailability.objects.filter(tutor=tutor).first()
        )
        session_date = session_date or date(2026, 1, 1)
        booking = Booking.objects.create(
            student=student, tutor=tutor, availability=availability,
            session_date=session_date, session_mode="Online", status="Completed",
        )
        return Rating.objects.create(
            booking=booking, student=student, tutor=tutor, rating_score=rating_score,
        )

    def test_403_when_flag_disabled_even_for_superadmin(self):
        self.client.force_authenticate(user=self.super_user)
        with override_settings(ALGORITHM_DEMO_TOOLS_ENABLED=False):
            response = self.client.get(f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}")
        self.assertEqual(response.status_code, 403)

    def test_403_for_regular_admin_even_with_flag_on(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}")
        self.assertEqual(response.status_code, 403)

    def test_tutee_search_returns_matching_profiles(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.get("/api/dev/algorithm-demo/tutees/?q=mar")
        self.assertEqual(response.status_code, 200)
        ids = {row["id"] for row in response.data["tutees"]}
        self.assertIn(self.tutee.id, ids)

    def test_tutee_search_matches_full_name(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.get("/api/dev/algorithm-demo/tutees/?q=Maria+Santos")
        self.assertEqual(response.status_code, 200)
        ids = {row["id"] for row in response.data["tutees"]}
        self.assertIn(self.tutee.id, ids)

    def test_tutee_search_does_not_n_plus_one_on_subject_lookups(self):
        # search_tutees fetches up to DEFAULT_TUTEE_SEARCH_LIMIT (500) tutees at
        # once for the frontend's client-side searchable dropdown, so a
        # per-tutee subject-preference query would silently scale to hundreds
        # of queries. Query count must stay flat as tutee count grows.
        for i in range(5):
            user = User.objects.create_user(
                username=f"algo-bulk-tutee-{i}@cpu.edu", email=f"algo-bulk-tutee-{i}@cpu.edu",
                password="password",
            )
            tutee = UserProfile.objects.create(
                user=user, fname=f"Bulk{i}", mname="", lname="Tutee", role="Tutee",
                year_level=11, institution=self.institution,
            )
            pref, _ = Preference.objects.get_or_create(user=tutee)
            pref.subjects.set([self.subject])

        self.client.force_authenticate(user=self.super_user)

        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get("/api/dev/algorithm-demo/tutees/?q=")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["tutees"]), 6)
        self.assertLess(len(ctx), 10)

    def test_recommend_requires_tutee_id(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.get("/api/dev/algorithm-demo/recommend/")
        self.assertEqual(response.status_code, 400)

    def test_recommend_404_for_unknown_tutee(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.get("/api/dev/algorithm-demo/recommend/?tutee_id=999999")
        self.assertEqual(response.status_code, 404)

    def test_recommend_matches_direct_hybrid_computation(self):
        from studybuddy.recommender.CF import build_rating_matrix
        from studybuddy.recommender.hybrid import hybrid_prediction

        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}")
        self.assertEqual(response.status_code, 200)

        rows = response.data["rows"]
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["tutor_id"], self.tutor.profile_id)

        expected = hybrid_prediction(
            build_rating_matrix(), self.tutee, self.tutor, None,
        )
        self.assertAlmostEqual(row["hybrid_score"], expected, places=6)

        cbf = row["cbf"]
        self.assertAlmostEqual(
            cbf["specific"]["contribution"]
            + cbf["general"]["contribution"]
            + cbf["expertise"]["contribution"]
            + cbf["course"]["contribution"]
            + cbf["year"]["contribution"]
            + cbf["level"]["contribution"],
            cbf["score"],
            places=6,
        )

    def test_recommend_flags_cold_start_tutee_with_no_ratings(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}")
        self.assertEqual(response.status_code, 200)

        row = response.data["rows"][0]
        self.assertTrue(row["cold_start"])
        self.assertIsNone(row["cf"]["score"])
        self.assertEqual(row["cf"]["neighbors"], [])

    def test_recommend_returns_no_preferences_reason_when_tutee_has_no_subjects(self):
        Preference.objects.filter(user=self.tutee).delete()

        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["reason"], "no_preferences")
        self.assertEqual(response.data["rows"], [])

    def test_recommend_row_includes_tutor_subjects_and_rating(self):
        self.tutor.rating_average = 4.5
        self.tutor.total_sessions = 12
        self.tutor.save()

        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}")
        self.assertEqual(response.status_code, 200)

        row = response.data["rows"][0]
        self.assertEqual(row["rating_average"], 4.5)
        self.assertEqual(row["total_sessions"], 12)
        self.assertEqual(
            row["tutor_subjects"],
            [{"code": self.subject.subject_code, "expertise_level": 4}],
        )

    def test_recommend_row_zero_rating_for_new_tutor(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}")
        self.assertEqual(response.status_code, 200)

        row = response.data["rows"][0]
        self.assertEqual(row["rating_average"], 0)
        self.assertEqual(row["total_sessions"], 0)

    def test_recommend_includes_subject_matching_tutor_from_other_institution(self):
        other_institution = PartnerInstitution.objects.create(
            institution_name="Other University", school_email_domain="other.edu", is_active=True,
        )
        other_tutor_user = User.objects.create_user(
            username="algo-tutor-other@other.edu", email="algo-tutor-other@other.edu",
            password="password",
        )
        other_tutor_profile = UserProfile.objects.create(
            user=other_tutor_user, fname="Dana", mname="", lname="Cross", role="Tutor",
            year_level=12, institution=other_institution,
        )
        other_tutor = Tutor.objects.create(
            profile=other_tutor_profile, hourly_rate=150, can_online=True, can_f2f=False,
            teaching_level="College",
        )
        TutorSubjects.objects.create(tutor=other_tutor, subject=self.subject, expertise_level=3)

        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}")
        self.assertEqual(response.status_code, 200)

        tutor_ids = {row["tutor_id"] for row in response.data["rows"]}
        self.assertIn(other_tutor.profile_id, tutor_ids)

    def test_recommend_institution_id_filter_excludes_other_institution_tutor(self):
        other_institution = PartnerInstitution.objects.create(
            institution_name="Other University", school_email_domain="other.edu", is_active=True,
        )
        other_tutor_user = User.objects.create_user(
            username="algo-tutor-scoped@other.edu", email="algo-tutor-scoped@other.edu",
            password="password",
        )
        other_tutor_profile = UserProfile.objects.create(
            user=other_tutor_user, fname="Ivy", mname="", lname="Scoped", role="Tutor",
            year_level=12, institution=other_institution,
        )
        other_tutor = Tutor.objects.create(
            profile=other_tutor_profile, hourly_rate=150, can_online=True, can_f2f=False,
            teaching_level="College",
        )
        TutorSubjects.objects.create(tutor=other_tutor, subject=self.subject, expertise_level=3)

        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(
            f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}"
            f"&institution_id={self.institution.id}"
        )
        self.assertEqual(response.status_code, 200)

        tutor_ids = {row["tutor_id"] for row in response.data["rows"]}
        self.assertIn(self.tutor.profile_id, tutor_ids)
        self.assertNotIn(other_tutor.profile_id, tutor_ids)

    def test_tutee_search_institution_id_filter_excludes_other_institution_tutee(self):
        other_institution = PartnerInstitution.objects.create(
            institution_name="Third University", school_email_domain="third.edu", is_active=True,
        )
        other_tutee_user = User.objects.create_user(
            username="algo-tutee-scoped@third.edu", email="algo-tutee-scoped@third.edu",
            password="password",
        )
        other_tutee = UserProfile.objects.create(
            user=other_tutee_user, fname="Maria", mname="", lname="Scoped", role="Tutee",
            year_level=11, institution=other_institution,
        )

        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(
            f"/api/dev/algorithm-demo/tutees/?q=Maria&institution_id={self.institution.id}"
        )
        self.assertEqual(response.status_code, 200)

        ids = {row["id"] for row in response.data["tutees"]}
        self.assertIn(self.tutee.id, ids)
        self.assertNotIn(other_tutee.id, ids)

    def test_algorithm_demo_update_rating_changes_score(self):
        rating = self._create_rating(rating_score=3)

        self.client.force_authenticate(user=self.super_user)
        response = self.client.patch(
            "/api/dev/algorithm-demo/rating/",
            {
                "student_id": self.tutee.id,
                "tutor_id": self.tutor.pk,
                "rating_score": 5,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, {"ok": True, "rating_score": 5})
        rating.refresh_from_db()
        self.assertEqual(rating.rating_score, 5)

    def test_algorithm_demo_update_rating_recomputes_tutor_average(self):
        rating = self._create_rating(rating_score=3)
        other_student_user = User.objects.create_user(
            username="algo-peer@cpu.edu", email="algo-peer@cpu.edu", password="password",
        )
        other_student = UserProfile.objects.create(
            user=other_student_user, fname="Peer", mname="", lname="Student", role="Tutee",
            year_level=11, institution=self.institution,
        )
        self._create_rating(student=other_student, rating_score=4, session_date=date(2026, 1, 2))
        self.tutor.rating_average = 3.5
        self.tutor.save(update_fields=["rating_average"])

        self.client.force_authenticate(user=self.super_user)
        response = self.client.patch(
            "/api/dev/algorithm-demo/rating/",
            {
                "student_id": self.tutee.id,
                "tutor_id": self.tutor.pk,
                "rating_score": 5,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        rating.refresh_from_db()
        self.tutor.refresh_from_db()
        self.assertEqual(rating.rating_score, 5)
        self.assertEqual(self.tutor.rating_average, 4.5)

    def test_algorithm_demo_update_rating_rejects_out_of_range_score(self):
        self._create_rating(rating_score=3)
        self.client.force_authenticate(user=self.super_user)

        low_response = self.client.patch(
            "/api/dev/algorithm-demo/rating/",
            {
                "student_id": self.tutee.id,
                "tutor_id": self.tutor.pk,
                "rating_score": 0,
            },
            format="json",
        )
        high_response = self.client.patch(
            "/api/dev/algorithm-demo/rating/",
            {
                "student_id": self.tutee.id,
                "tutor_id": self.tutor.pk,
                "rating_score": 6,
            },
            format="json",
        )

        self.assertEqual(low_response.status_code, 400)
        self.assertEqual(high_response.status_code, 400)

    def test_algorithm_demo_update_rating_404_when_no_rating_exists(self):
        self.client.force_authenticate(user=self.super_user)

        self.assertEqual(
            Rating.objects.filter(student_id=self.tutee.id, tutor_id=self.tutor.pk).count(), 0,
        )
        response = self.client.patch(
            "/api/dev/algorithm-demo/rating/",
            {
                "student_id": self.tutee.id,
                "tutor_id": self.tutor.pk,
                "rating_score": 5,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(
            Rating.objects.filter(student_id=self.tutee.id, tutor_id=self.tutor.pk).count(), 0,
        )

    def test_algorithm_demo_update_rating_requires_superadmin(self):
        self._create_rating(rating_score=3)

        self.client.force_authenticate(user=self.admin_user)
        regular_admin_response = self.client.patch(
            "/api/dev/algorithm-demo/rating/",
            {
                "student_id": self.tutee.id,
                "tutor_id": self.tutor.pk,
                "rating_score": 5,
            },
            format="json",
        )
        self.assertEqual(regular_admin_response.status_code, 403)

        self.client.force_authenticate(user=self.super_user)
        with override_settings(ALGORITHM_DEMO_TOOLS_ENABLED=False):
            disabled_response = self.client.patch(
                "/api/dev/algorithm-demo/rating/",
                {
                    "student_id": self.tutee.id,
                    "tutor_id": self.tutor.pk,
                    "rating_score": 5,
                },
                format="json",
            )
        self.assertEqual(disabled_response.status_code, 403)

    # --- what-if re-scoring + CF derivation payload -----------------------------------------
    # See docs/plans/2026-07-23-algorithm-demo-cf-clarity.md.

    def _shared_other_tutor(self):
        """A second tutor, deliberately without the demo subject so it stays out
        of the candidate pool, existing only to be co-rated. Pearson runs over the
        co-rated intersection (CF.py:37) and returns 0 for a single shared item,
        so a neighbour sharing only one tutor is filtered out before it can
        influence anything — CF needs at least two co-rated tutors to register."""
        other_user = User.objects.create_user(
            username="algo-tutor-other@cpu.edu", email="algo-tutor-other@cpu.edu",
            password="password",
        )
        other_profile = UserProfile.objects.create(
            user=other_user, fname="Nena", mname="", lname="Other", role="Tutor",
            year_level=12, institution=self.institution,
        )
        other_tutor = Tutor.objects.create(
            profile=other_profile, hourly_rate=180, can_online=True, can_f2f=False,
            teaching_level="College",
        )
        TutorAvailability.objects.create(tutor=other_tutor, day="Tue", time_slot=time(10, 0))
        return other_tutor

    def _peer_who_rated(self, rating_score):
        """A second tutee who has co-rated two tutors with the demo tutee, so the
        similarity is non-zero and the peer survives into the neighbour set."""
        peer_user = User.objects.create_user(
            username="algo-peer@cpu.edu", email="algo-peer@cpu.edu", password="password",
        )
        peer = UserProfile.objects.create(
            user=peer_user, fname="Peer", mname="", lname="Rater", role="Tutee",
            year_level=11, institution=self.institution,
        )
        other_tutor = self._shared_other_tutor()
        self._create_rating(
            student=self.tutee, rating_score=4, session_date=date(2026, 2, 2), tutor=other_tutor,
        )
        self._create_rating(
            student=peer, rating_score=5, session_date=date(2026, 2, 3), tutor=other_tutor,
        )
        self._create_rating(student=peer, rating_score=rating_score, session_date=date(2026, 2, 1))
        return peer

    def test_recommend_payload_carries_cf_derivation_terms(self):
        self._create_rating(rating_score=4)

        self.client.force_authenticate(user=self.super_user)
        response = self.client.get(
            f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}"
        )
        self.assertEqual(response.status_code, 200)

        cf = response.data["rows"][0]["cf"]
        for key in ("student_avg", "student_rating", "numerator", "denominator"):
            self.assertIn(key, cf)
        # The tutee rated this tutor herself, so the panel can compare prediction
        # against the real score — the candidate pool does not exclude rated tutors.
        self.assertEqual(cf["student_rating"], 4)

    def test_whatif_override_changes_score_without_writing(self):
        self._create_rating(rating_score=2)
        peer = self._peer_who_rated(rating_score=2)

        self.client.force_authenticate(user=self.super_user)
        baseline = self.client.get(
            f"/api/dev/algorithm-demo/recommend/?tutee_id={self.tutee.id}"
        )
        baseline_score = baseline.data["rows"][0]["hybrid_score"]

        response = self.client.post(
            "/api/dev/algorithm-demo/recommend-whatif/",
            {
                "tutee_id": self.tutee.id,
                "overrides": [
                    {"student_id": peer.id, "tutor_id": self.tutor.pk, "rating_score": 5},
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotEqual(response.data["rows"][0]["hybrid_score"], baseline_score)

        # Nothing persisted: the peer's stored rating is untouched.
        self.assertEqual(
            Rating.objects.get(student=peer, tutor=self.tutor).rating_score, 2
        )

    def test_whatif_rejects_out_of_range_override(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.post(
            "/api/dev/algorithm-demo/recommend-whatif/",
            {
                "tutee_id": self.tutee.id,
                "overrides": [
                    {"student_id": self.tutee.id, "tutor_id": self.tutor.pk, "rating_score": 9},
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400)

    def test_whatif_requires_superadmin_and_flag(self):
        self.client.force_authenticate(user=self.admin_user)
        payload = {"tutee_id": self.tutee.id, "overrides": []}
        self.assertEqual(
            self.client.post(
                "/api/dev/algorithm-demo/recommend-whatif/", payload, format="json"
            ).status_code,
            403,
        )

        self.client.force_authenticate(user=self.super_user)
        with override_settings(ALGORITHM_DEMO_TOOLS_ENABLED=False):
            disabled = self.client.post(
                "/api/dev/algorithm-demo/recommend-whatif/", payload, format="json"
            )
        self.assertEqual(disabled.status_code, 403)


class SessionCheckInTests(APITestCase):
    def setUp(self):
        self.course = Course.objects.create(
            course_code="CHECK101",
            course_name="Session Checks",
        )
        self.subject = Subjects.objects.create(
            subject_code="CHECK-SUBJ",
            subject_name="Checkpoint Subject",
            department="Checks",
        )
        self.student_user = User.objects.create_user(
            username="check-student",
            email="check-student@example.com",
            password="password",
        )
        self.student = UserProfile.objects.create(
            user=self.student_user,
            fname="Check",
            mname="",
            lname="Student",
            role="Tutee",
            year_level=11,
            course=self.course,
        )
        self.tutor_user = User.objects.create_user(
            username="check-tutor",
            email="check-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Check",
            mname="",
            lname="Tutor",
            role="Tutor",
            year_level=12,
            course=self.course,
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=200,
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        # Ongoing right now, so the default fixture matches the "session is
        # live" gate on the midpoint check-in endpoint without every test
        # needing to reason about the clock.
        now = timezone.localtime()
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=(now - timedelta(minutes=5)).time(),
            is_active=True,
        )
        self.booking = Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            subject=self.subject,
            session_date=now.date(),
            session_mode="F2F",
            preferred_location="Library",
            status="Confirmed",
        )
        self.client.force_authenticate(user=self.student_user)

    def _make_booking_with_window(self, *, start_offset_minutes, status="Confirmed"):
        now = timezone.localtime()
        availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Tue",
            time_slot=(now + timedelta(minutes=start_offset_minutes)).time(),
            is_active=True,
        )
        return Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=availability,
            session_date=(now + timedelta(minutes=start_offset_minutes)).date(),
            session_mode="F2F",
            preferred_location="Library",
            status=status,
        )

    def test_tutee_can_record_venue_confirmation(self):
        response = self.client.post(
            f"/api/bookings/{self.booking.id}/venue-confirmation/",
            {"response": "yes"},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["response"], "yes")
        self.assertTrue(
            SessionCheckIn.objects.filter(
                booking=self.booking,
                event_type=SessionCheckIn.EVENT_VENUE_CONFIRM,
                response="yes",
            ).exists()
        )

    def test_venue_confirmation_is_f2f_only(self):
        self.booking.session_mode = "Online"
        self.booking.save(update_fields=["session_mode"])

        response = self.client.post(
            f"/api/bookings/{self.booking.id}/venue-confirmation/",
            {"response": "yes"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(SessionCheckIn.objects.exists())

    def test_tutee_can_record_midpoint_check_in(self):
        response = self.client.post(
            f"/api/bookings/{self.booking.id}/midpoint-check-in/",
            {"response": "issues"},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["response"], "issues")
        self.assertTrue(
            SessionCheckIn.objects.filter(
                booking=self.booking,
                event_type=SessionCheckIn.EVENT_MIDPOINT_CHECKIN,
                response="issues",
            ).exists()
        )

    def test_duplicate_check_in_returns_existing_response(self):
        first = self.client.post(
            f"/api/bookings/{self.booking.id}/midpoint-check-in/",
            {"response": "good"},
            format="json",
        )
        second = self.client.post(
            f"/api/bookings/{self.booking.id}/midpoint-check-in/",
            {"response": "issues"},
            format="json",
        )

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.data["response"], "good")
        self.assertEqual(
            SessionCheckIn.objects.filter(
                booking=self.booking,
                event_type=SessionCheckIn.EVENT_MIDPOINT_CHECKIN,
            ).count(),
            1,
        )

    def test_midpoint_check_in_rejected_when_session_is_upcoming(self):
        booking = self._make_booking_with_window(start_offset_minutes=60)

        response = self.client.post(
            f"/api/bookings/{booking.id}/midpoint-check-in/",
            {"response": "good"},
            format="json",
        )

        self.assertEqual(response.status_code, 409)
        self.assertFalse(
            SessionCheckIn.objects.filter(booking=booking).exists()
        )

    def test_midpoint_check_in_rejected_when_session_already_ended(self):
        booking = self._make_booking_with_window(start_offset_minutes=-180)

        response = self.client.post(
            f"/api/bookings/{booking.id}/midpoint-check-in/",
            {"response": "good"},
            format="json",
        )

        self.assertEqual(response.status_code, 409)
        self.assertFalse(
            SessionCheckIn.objects.filter(booking=booking).exists()
        )

    def test_booking_detail_includes_check_in_state(self):
        SessionCheckIn.objects.create(
            booking=self.booking,
            event_type=SessionCheckIn.EVENT_VENUE_CONFIRM,
            response="no",
        )

        response = self.client.get(f"/api/bookings/{self.booking.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["check_ins"]["venue_confirm"]["response"], "no")
        self.assertIsNone(response.data["check_ins"]["midpoint_checkin"])

    def test_booking_detail_session_subject_reflects_booked_subject_not_course(self):
        response = self.client.get(f"/api/bookings/{self.booking.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.data["session"]["subject"],
            self.subject.subject_name,
        )

    def test_booking_detail_session_subject_falls_back_to_general_when_null(self):
        self.booking.subject = None
        self.booking.save(update_fields=["subject"])

        response = self.client.get(f"/api/bookings/{self.booking.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["session"]["subject"], "General")

    def test_booking_detail_uses_absolute_receipt_media_url(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        method = PaymentMethod.objects.create(
            code="CASH",
            method_name="Cash",
            is_active=True,
        )
        Payment.objects.create(
            booking=self.booking,
            amount=Decimal("200.00"),
            method=method,
            payment_status="Pending",
            receipt_image=SimpleUploadedFile(
                "receipt.jpg",
                b"receipt-bytes",
                content_type="image/jpeg",
            ),
        )

        response = self.client.get(f"/api/bookings/{self.booking.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            response.data["payment"]["receipt_image"].startswith("http://testserver/media/")
        )

    def test_other_users_cannot_record_check_in(self):
        other_user = User.objects.create_user(
            username="check-other",
            email="check-other@example.com",
            password="password",
        )
        UserProfile.objects.create(
            user=other_user,
            fname="Other",
            mname="",
            lname="User",
            role="Tutee",
            year_level=11,
        )
        self.client.force_authenticate(user=other_user)

        response = self.client.post(
            f"/api/bookings/{self.booking.id}/midpoint-check-in/",
            {"response": "good"},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(SessionCheckIn.objects.exists())


class DevLiveSessionTests(APITestCase):
    def setUp(self):
        cache.clear()
        self.course = Course.objects.create(
            course_code="DEVLIVE",
            course_name="Dev Live",
        )
        self.student_user = User.objects.create_user(
            username="dev-live-student",
            email="dev-live-student@example.com",
            password="password",
        )
        self.student = UserProfile.objects.create(
            user=self.student_user,
            fname="Dev",
            mname="",
            lname="Student",
            role="Tutee",
            year_level=11,
            course=self.course,
        )
        self.tutor_user = User.objects.create_user(
            username="dev-live-tutor",
            email="dev-live-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Dev",
            mname="",
            lname="Tutor",
            role="Tutor",
            year_level=12,
            course=self.course,
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=200,
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        TutorSubjects.objects.create(
            tutor=self.tutor,
            subject=Subjects.objects.create(
                subject_code="DEV101",
                subject_name="Dev QA",
                department="QA",
            ),
            expertise_level=5,
        )
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=time(14, 0),
            is_active=True,
        )
        self.booking = Booking.objects.create(
            student=self.student,
            tutor=self.tutor,
            availability=self.availability,
            session_date=date.today() + timedelta(days=3),
            session_mode="F2F",
            preferred_location="Library",
            status="Confirmed",
        )
        self.client.force_authenticate(user=self.student_user)

    def tearDown(self):
        cache.clear()

    @override_settings(BOOKING_DEV_TOOLS_ENABLED=False)
    def test_force_live_returns_404_when_booking_dev_tools_disabled(self):
        """The endpoint gate is BOOKING_DEV_TOOLS_ENABLED, not DEBUG -- real production has both
        off, but the demo deployment runs DEBUG=False with this flag on (see
        test_force_live_takes_effect_on_the_demo_deployment)."""
        response = self.client.post(
            f"/api/dev/bookings/{self.booking.id}/force-live/",
            {"phase": "start"},
            format="json",
        )

        self.assertEqual(response.status_code, 404)

    @override_settings(DEBUG=True)
    def test_force_live_rejects_unrelated_users(self):
        other_user = User.objects.create_user(
            username="dev-live-other",
            email="dev-live-other@example.com",
            password="password",
        )
        UserProfile.objects.create(
            user=other_user,
            fname="Other",
            mname="",
            lname="User",
            role="Tutee",
            year_level=11,
            course=self.course,
        )
        self.client.force_authenticate(user=other_user)

        response = self.client.post(
            f"/api/dev/bookings/{self.booking.id}/force-live/",
            {"phase": "start"},
            format="json",
        )

        self.assertEqual(response.status_code, 403)

    @override_settings(DEBUG=True)
    def test_force_live_updates_detail_and_list_payloads(self):
        response = self.client.post(
            f"/api/dev/bookings/{self.booking.id}/force-live/",
            {"phase": "midpoint"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["session"]["status"], "Ongoing")

        list_response = self.client.get("/api/bookings/")
        self.assertEqual(list_response.status_code, 200)
        booking_payload = next(
            item for item in list_response.data if item["id"] == self.booking.id
        )

        self.assertEqual(booking_payload["status"], "Ongoing")
        self.assertEqual(str(booking_payload["date"]), response.data["session"]["date"])
        self.assertEqual(booking_payload["startTime"], response.data["session"]["start_time"])
        self.assertEqual(booking_payload["endTime"], response.data["session"]["end_time"])

    @override_settings(DEBUG=False, BOOKING_DEV_TOOLS_ENABLED=True)
    def test_force_live_takes_effect_on_the_demo_deployment(self):
        """Reproduces the demo deployment's actual settings (DEBUG=False,
        BOOKING_DEV_TOOLS_ENABLED=True) rather than local dev's (DEBUG=True). Regression test for
        a bug where get_dev_live_override_for_bookings() independently gated on DEBUG instead of
        BOOKING_DEV_TOOLS_ENABLED, so the endpoint returned 200 but the override it just wrote was
        invisible to every read path under these settings -- the booking silently never went
        Ongoing."""
        response = self.client.post(
            f"/api/dev/bookings/{self.booking.id}/force-live/",
            {"phase": "midpoint"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["session"]["status"], "Ongoing")

    @override_settings(DEBUG=True)
    def test_tutor_can_force_and_clear_live_session(self):
        self.client.force_authenticate(user=self.tutor_user)

        force_response = self.client.post(
            f"/api/dev/bookings/{self.booking.id}/force-live/",
            {"phase": "ending"},
            format="json",
        )
        self.assertEqual(force_response.status_code, 200)
        self.assertEqual(force_response.data["session"]["status"], "Ongoing")

        clear_response = self.client.post(
            f"/api/dev/bookings/{self.booking.id}/clear-force-live/",
            format="json",
        )

        self.assertEqual(clear_response.status_code, 200)
        self.assertEqual(clear_response.data["session"]["status"], "Upcoming")


class TutorCashInTests(APITestCase):
    def setUp(self):
        self.tutor_user = User.objects.create_user(
            username="cashin-tutor",
            email="cashin-tutor@example.com",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Cash",
            mname="",
            lname="In",
            role="Tutor",
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("300.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.wallet = Wallet.objects.get(tutor=self.tutor)
        self.wallet.balance = Decimal("-50.00")
        self.wallet.save(update_fields=["balance"])
        self.client.force_authenticate(user=self.tutor_user)

    def checkout_ok(self):
        class Resp:
            status_code = 200

            @staticmethod
            def json():
                return {
                    "data": {
                        "id": "cs_test_123",
                        "attributes": {"checkout_url": "https://pm.test/checkout/cs_test_123"},
                    }
                }
        return Resp()

    def checkout_paid(self):
        class Resp:
            status_code = 200

            @staticmethod
            def json():
                return {"data": {"attributes": {"status": "paid"}}}
        return Resp()

    def make_topup(self, status_value="pending", ref="cs_test_123", amount="100.00"):
        return WalletTopUp.objects.create(
            tutor=self.tutor,
            amount=Decimal(amount),
            status=status_value,
            provider="paymongo",
            provider_reference=ref,
        )

    def test_initiate_rejects_invalid_amount(self):
        response = self.client.post(
            "/api/wallet/cash-in/", {"amount": "0"}, format="json"
        )
        self.assertEqual(response.status_code, 400)

    def test_non_tutor_gets_403(self):
        plain = User.objects.create_user(username="plain", email="p@e.com", password="x")
        self.client.force_authenticate(user=plain)
        response = self.client.post(
            "/api/wallet/cash-in/", {"amount": "100"}, format="json"
        )
        self.assertEqual(response.status_code, 403)

    @patch("studybuddy.views.requests.post")
    def test_initiate_creates_topup_and_returns_checkout_url(self, mock_post):
        mock_post.return_value = self.checkout_ok()
        response = self.client.post(
            "/api/wallet/cash-in/", {"amount": "50.00"}, format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["checkout_url"], "https://pm.test/checkout/cs_test_123")
        topup = WalletTopUp.objects.get(id=response.data["id"])
        self.assertEqual(topup.status, "pending")
        self.assertEqual(topup.amount, Decimal("50.00"))
        self.assertEqual(topup.provider_reference, "cs_test_123")
        sent = mock_post.call_args.kwargs["json"]
        self.assertEqual(
            sent["data"]["attributes"]["line_items"][0]["amount"], 5000
        )

    @patch("studybuddy.views.requests.get")
    def test_verify_credits_wallet_and_writes_transaction(self, mock_get):
        mock_get.return_value = self.checkout_paid()
        topup = self.make_topup()

        response = self.client.post(f"/api/wallet/cash-in/{topup.id}/verify/")

        self.assertEqual(response.status_code, 200)
        topup.refresh_from_db()
        self.wallet.refresh_from_db()
        self.assertEqual(topup.status, "paid")
        self.assertIsNotNone(topup.paid_at)
        self.assertEqual(self.wallet.balance, Decimal("50.00"))
        self.assertTrue(
            Transaction.objects.filter(
                wallet=self.wallet,
                transaction_type="cash_in",
                amount=Decimal("100.00"),
                reference_id=f"TOPUP-{topup.id}",
            ).exists()
        )

    @patch("studybuddy.views.requests.get")
    def test_verify_is_idempotent(self, mock_get):
        mock_get.return_value = self.checkout_paid()
        topup = self.make_topup()

        first = self.client.post(f"/api/wallet/cash-in/{topup.id}/verify/")
        second = self.client.post(f"/api/wallet/cash-in/{topup.id}/verify/")

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.wallet.refresh_from_db()
        self.assertEqual(self.wallet.balance, Decimal("50.00"))
        self.assertEqual(
            Transaction.objects.filter(reference_id=f"TOPUP-{topup.id}").count(), 1
        )

    @patch("studybuddy.views.requests.get")
    def test_verify_unpaid_returns_400_and_no_credit(self, mock_get):
        class Unpaid:
            status_code = 200

            @staticmethod
            def json():
                return {"data": {"attributes": {"status": "unpaid"}}}
        mock_get.return_value = Unpaid()
        topup = self.make_topup()

        response = self.client.post(f"/api/wallet/cash-in/{topup.id}/verify/")

        self.assertEqual(response.status_code, 400)
        self.wallet.refresh_from_db()
        self.assertEqual(self.wallet.balance, Decimal("-50.00"))
        topup.refresh_from_db()
        self.assertEqual(topup.status, "pending")


class AvatarCompressionTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="tutor_avatar",
            email="tutor_avatar@example.com",
            password="password",
        )
        self.profile = UserProfile.objects.create(
            user=self.user,
            fname="Avatar",
            mname="",
            lname="Tutor",
            role="Tutor",
            year_level=12,
        )
        Tutor.objects.create(
            profile=self.profile,
            hourly_rate=200,
            teaching_level="College",
        )
        self.client.force_authenticate(user=self.user)

    def _png_upload(self, name, size):
        from io import BytesIO
        from PIL import Image
        from django.core.files.uploadedfile import SimpleUploadedFile

        buffer = BytesIO()
        Image.new("RGB", size, (120, 80, 200)).save(buffer, format="PNG")
        buffer.seek(0)
        return SimpleUploadedFile(name, buffer.read(), content_type="image/png")

    def test_oversized_image_is_compressed_to_webp(self):
        from io import BytesIO
        from PIL import Image

        upload = self._png_upload("huge.png", (1000, 1000))
        response = self.client.post(
            "/api/tutor/profile/avatar/", {"avatar": upload}, format="multipart"
        )

        self.assertEqual(response.status_code, 200)
        self.profile.refresh_from_db()
        self.assertTrue(self.profile.profile_picture.name.endswith(".webp"))

        self.profile.profile_picture.open()
        stored = Image.open(BytesIO(self.profile.profile_picture.read()))
        self.assertEqual(stored.format, "WEBP")
        self.assertLessEqual(max(stored.size), 512)

    def test_unreadable_image_is_rejected(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        bogus = SimpleUploadedFile(
            "broken.png", b"not a real image", content_type="image/png"
        )
        response = self.client.post(
            "/api/tutor/profile/avatar/", {"avatar": bogus}, format="multipart"
        )

        self.assertEqual(response.status_code, 400)


class SupportTicketEscalationTests(APITestCase):
    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="Central Philippine University",
            school_email_domain="cpu.edu.ph",
            is_active=True,
        )
        self.other_institution = PartnerInstitution.objects.create(
            institution_name="Other University",
            school_email_domain="other.edu.ph",
            is_active=True,
        )
        self.admin_user = User.objects.create_user(
            username="support_admin",
            email="admin@cpu.edu.ph",
            password="password",
        )
        self.admin_profile = UserProfile.objects.create(
            user=self.admin_user,
            fname="Ada",
            mname="",
            lname="Admin",
            role="Admin",
            institution=self.institution,
        )
        self.super_user = User.objects.create_user(
            username="super_support",
            email="super@studybuddy.test",
            password="password",
        )
        self.super_profile = UserProfile.objects.create(
            user=self.super_user,
            fname="Sam",
            mname="",
            lname="Super",
            role="SuperAdmin",
            profile_completed=True,
            is_domain_exempt=True,
        )
        self.tutee_user = User.objects.create_user(
            username="support_tutee",
            email="tutee@cpu.edu.ph",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Tina",
            mname="",
            lname="Tutee",
            role="Tutee",
            institution=self.institution,
        )
        self.room = ChatRoom.objects.create(
            tutee=self.tutee_profile,
            tutor=self.admin_profile,
            room_type="support",
        )
        self.ticket = SupportTicket.objects.create(
            user=self.tutee_profile,
            category="Technical",
            subject="Cannot join session",
            description="The call page will not load.",
            status="In_Progress",
            assigned_agent=self.admin_profile,
            chatroom=self.room,
        )

    def test_admin_escalates_ticket_to_superadmin_queue(self):
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.post(
            f"/api/admin/support/tickets/{self.ticket.id}/escalate/",
            {"reason": "Payment records need platform access."},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.ticket.refresh_from_db()
        self.room.refresh_from_db()
        self.assertEqual(self.ticket.status, "Escalated")
        self.assertEqual(self.ticket.escalation_reason, "Payment records need platform access.")
        self.assertEqual(self.ticket.escalated_by, self.admin_profile)
        self.assertIsNotNone(self.ticket.escalated_at)
        self.assertIsNone(self.ticket.assigned_agent)
        self.assertIsNone(self.room.tutor)
        self.assertTrue(
            Message.objects.filter(
                room=self.room,
                sender=None,
                content="This support ticket has been escalated to SuperAdmin support.",
            ).exists()
        )

        self.client.force_authenticate(user=self.super_user)
        super_list = self.client.get("/api/admin/support/tickets/")
        escalated_ticket = next(ticket for ticket in super_list.data if ticket["id"] == self.ticket.id)
        self.assertEqual(escalated_ticket["status"], "Escalated")
        self.assertEqual(escalated_ticket["escalation_reason"], "Payment records need platform access.")

    def test_escalation_requires_reason_and_institution_scope(self):
        self.client.force_authenticate(user=self.admin_user)

        missing_reason = self.client.post(
            f"/api/admin/support/tickets/{self.ticket.id}/escalate/",
            {"reason": "   "},
            format="json",
        )

        self.assertEqual(missing_reason.status_code, 400)
        self.ticket.refresh_from_db()
        self.assertEqual(self.ticket.status, "In_Progress")

        other_admin_user = User.objects.create_user(
            username="other_admin",
            email="admin@other.edu.ph",
            password="password",
        )
        UserProfile.objects.create(
            user=other_admin_user,
            fname="Omar",
            mname="",
            lname="Admin",
            role="Admin",
            institution=self.other_institution,
        )
        self.client.force_authenticate(user=other_admin_user)

        cross_institution = self.client.post(
            f"/api/admin/support/tickets/{self.ticket.id}/escalate/",
            {"reason": "Needs platform review."},
            format="json",
        )

        self.assertEqual(cross_institution.status_code, 404)
        self.ticket.refresh_from_db()
        self.assertEqual(self.ticket.status, "In_Progress")

    def test_superadmin_claims_and_resolves_escalated_ticket(self):
        self.client.force_authenticate(user=self.admin_user)
        self.client.post(
            f"/api/admin/support/tickets/{self.ticket.id}/escalate/",
            {"reason": "Needs platform review."},
            format="json",
        )

        admin_resolve = self.client.post(f"/api/admin/support/tickets/{self.ticket.id}/resolve/")
        self.assertEqual(admin_resolve.status_code, 403)

        self.client.force_authenticate(user=self.super_user)
        claim = self.client.post(f"/api/admin/support/tickets/{self.ticket.id}/claim/")
        self.assertEqual(claim.status_code, 200)
        self.ticket.refresh_from_db()
        self.room.refresh_from_db()
        self.assertEqual(self.ticket.status, "Escalated")
        self.assertEqual(self.ticket.assigned_agent, self.super_profile)
        self.assertEqual(self.room.tutor, self.super_profile)

        resolve = self.client.post(f"/api/admin/support/tickets/{self.ticket.id}/resolve/")
        self.assertEqual(resolve.status_code, 200)
        self.ticket.refresh_from_db()
        self.assertEqual(self.ticket.status, "Resolved")


class CrossInstitutionMatchingTests(APITestCase):
    """Matching is unscoped by institution — a tutee from one school can match a tutor from
    any other, or a tutor with no institution at all. See
    docs/plans/2026-07-17-remove-institution-matching-constraint.md."""

    def setUp(self):
        cache.clear()

        self.inst_a = PartnerInstitution.objects.create(
            institution_name="Central Philippine University",
            school_email_domain="cpu.edu.ph",
            is_active=True,
            contact_person="Registrar",
        )
        self.inst_b = PartnerInstitution.objects.create(
            institution_name="Western Visayas University",
            school_email_domain="wvu.edu.ph",
            is_active=True,
            contact_person="Registrar",
        )
        self.subject = Subjects.objects.create(
            subject_code="SCOPE101",
            subject_name="Scoping Test Subject",
            department="Test",
        )

        # Tutee from inst_a
        tutee_user = User.objects.create_user(
            username="scope_tutee_a", email="tutee@cpu.edu.ph", password="pass"
        )
        self.tutee = UserProfile.objects.create(
            user=tutee_user, fname="Ana", mname="", lname="Cruz",
            role="Tutee", institution=self.inst_a,
        )

        # Tutor from inst_a
        tutor_a_user = User.objects.create_user(
            username="scope_tutor_a", email="tutor_a@cpu.edu.ph", password="pass"
        )
        tutor_a_profile = UserProfile.objects.create(
            user=tutor_a_user, fname="Ben", mname="", lname="Santos",
            role="Tutor", institution=self.inst_a,
        )
        self.tutor_a = Tutor.objects.create(
            profile=tutor_a_profile, hourly_rate=250,
            can_online=True, can_f2f=True, teaching_level="SHS",
        )
        TutorSubjects.objects.create(
            tutor=self.tutor_a, subject=self.subject, expertise_level=5
        )

        # Tutor from inst_b — a different institution than the tutee
        tutor_b_user = User.objects.create_user(
            username="scope_tutor_b", email="tutor_b@wvu.edu.ph", password="pass"
        )
        tutor_b_profile = UserProfile.objects.create(
            user=tutor_b_user, fname="Cal", mname="", lname="Ramos",
            role="Tutor", institution=self.inst_b,
        )
        self.tutor_b = Tutor.objects.create(
            profile=tutor_b_profile, hourly_rate=250,
            can_online=True, can_f2f=True, teaching_level="SHS",
        )
        TutorSubjects.objects.create(
            tutor=self.tutor_b, subject=self.subject, expertise_level=5
        )

        # Tutor with no institution at all
        tutor_null_user = User.objects.create_user(
            username="scope_tutor_null", email="tutor_null@example.com", password="pass"
        )
        tutor_null_profile = UserProfile.objects.create(
            user=tutor_null_user, fname="Dan", mname="", lname="Lee",
            role="Tutor", institution=None,
        )
        self.tutor_null = Tutor.objects.create(
            profile=tutor_null_profile, hourly_rate=250,
            can_online=True, can_f2f=True, teaching_level="SHS",
        )
        TutorSubjects.objects.create(
            tutor=self.tutor_null, subject=self.subject, expertise_level=5
        )

        # recommend-tutors only surfaces approved tutors — approve all three so the
        # inclusion assertions below are meaningful rather than vacuously true.
        for tutor in (self.tutor_a, self.tutor_b, self.tutor_null):
            self._approve(tutor)

    def _approve(self, tutor):
        from django.core.files.uploadedfile import SimpleUploadedFile

        TutorApplication.objects.create(
            profile=tutor.profile,
            school_id=SimpleUploadedFile("id.jpg", b"id", content_type="image/jpeg"),
            enrollment_proof=SimpleUploadedFile("rf.pdf", b"rf", content_type="application/pdf"),
            application_status="approved",
        )

    def _set_preferences(self, *subjects):
        pref, _ = Preference.objects.get_or_create(user=self.tutee)
        pref.subjects.set([s.subject_code for s in subjects])

    def test_recommend_includes_same_institution_tutor(self):
        self.client.force_authenticate(user=self.tutee.user)
        resp = self.client.post(
            "/api/recommend-tutors/",
            {"subject": "SCOPE101", "date": timezone.localdate().isoformat()},
            format="json",
        )
        self.assertEqual(resp.status_code, 200)
        ids = {r["id"] for r in resp.data["tutors"]}
        self.assertIn(self.tutor_a.profile.id, ids)

    def test_recommend_includes_other_institution_tutor(self):
        self.client.force_authenticate(user=self.tutee.user)
        resp = self.client.post(
            "/api/recommend-tutors/",
            {"subject": "SCOPE101", "date": timezone.localdate().isoformat()},
            format="json",
        )
        self.assertEqual(resp.status_code, 200)
        ids = {r["id"] for r in resp.data["tutors"]}
        self.assertIn(self.tutor_b.profile.id, ids)

    def test_recommend_tutee_no_institution_still_matches(self):
        no_inst_user = User.objects.create_user(
            username="scope_no_inst2", email="noinst2@test.com", password="pass"
        )
        UserProfile.objects.create(
            user=no_inst_user, fname="No", mname="", lname="Inst",
            role="Tutee", institution=None,
        )
        self.client.force_authenticate(user=no_inst_user)
        resp = self.client.post(
            "/api/recommend-tutors/",
            {"subject": "SCOPE101", "date": timezone.localdate().isoformat()},
            format="json",
        )
        self.assertEqual(resp.status_code, 200)
        ids = {r["id"] for r in resp.data["tutors"]}
        self.assertIn(self.tutor_a.profile.id, ids)

    def test_recommend_includes_null_institution_tutor(self):
        self.client.force_authenticate(user=self.tutee.user)
        resp = self.client.post(
            "/api/recommend-tutors/",
            {"subject": "SCOPE101", "date": timezone.localdate().isoformat()},
            format="json",
        )
        self.assertEqual(resp.status_code, 200)
        ids = {r["id"] for r in resp.data["tutors"]}
        self.assertIn(self.tutor_null.profile.id, ids)

    def test_dashboard_widget_includes_cross_institution_tutors(self):
        from studybuddy.recommender import dashboard
        self._set_preferences(self.subject)
        data = dashboard.get_dashboard_recommendations(self.tutee)
        ids = {row["id"] for row in data}
        self.assertIn(self.tutor_b.profile.id, ids)
        self.assertIn(self.tutor_null.profile.id, ids)

    def test_dashboard_fallback_includes_cross_institution_tutors(self):
        from studybuddy.recommender import dashboard
        # tutee has no preferences set, so get_student_subject_codes returns []
        # and the code falls through to _fallback
        data = dashboard.get_dashboard_recommendations(self.tutee)
        ids = {row["id"] for row in data}
        self.assertIn(self.tutor_b.profile.id, ids)
        self.assertIn(self.tutor_null.profile.id, ids)

    def test_search_tutors_includes_cross_institution_tutors(self):
        self.client.force_authenticate(user=self.tutee.user)
        resp = self.client.get("/api/search-tutors/", {"subject": "SCOPE101"})
        self.assertEqual(resp.status_code, 200)
        ids = {r["profile_id"] for r in resp.data}
        self.assertIn(self.tutor_a.profile.id, ids)
        self.assertIn(self.tutor_b.profile.id, ids)
        self.assertIn(self.tutor_null.profile.id, ids)

    def test_search_tutors_requires_authentication(self):
        resp = self.client.get("/api/search-tutors/", {"subject": "SCOPE101"})
        self.assertEqual(resp.status_code, 401)


@override_settings(VERIFICATION_DEV_TOOLS_ENABLED=True)
class VerificationDevToolsTests(APITestCase):
    """Self-service verification dev tools — state jumper, enforcement override, gating.
    See docs/plans/2026-07-02-verification-dev-tools.md."""

    def setUp(self):
        from . import _verification_dev

        _verification_dev.enforcement_override_clear()
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu", is_active=True,
        )
        self.tutee = self._make_profile("devtools-tutee@cpu.edu", "Tutee")
        self.tutor = self._make_profile("devtools-tutor@cpu.edu", "Tutor")

    def tearDown(self):
        from . import _verification_dev

        _verification_dev.enforcement_override_clear()

    def _make_profile(self, email, role):
        user = User.objects.create_user(username=email, email=email, password="password")
        return UserProfile.objects.create(
            user=user, fname="Dev", mname="", lname=role, role=role, institution=self.institution,
        )

    # --- Helper: set_verification_state produces each state -----------------------------------

    def _assert_state(self, profile, state, expected_app_status, expected_renewal_status):
        from . import _verification_dev

        _verification_dev.set_verification_state(profile, state)
        related_name = "tutee_application" if profile.role == "Tutee" else "tutor_application"
        # Re-fetch the profile so a just-deleted OneToOne reverse accessor re-evaluates cleanly.
        profile = UserProfile.objects.get(pk=profile.pk)
        app = getattr(profile, related_name, None)
        if state == "not_submitted":
            self.assertIsNone(app)
            return
        app.refresh_from_db()
        self.assertEqual(app.application_status, expected_app_status)
        self.assertEqual(app.document_renewal_status(), expected_renewal_status)

    def test_set_state_all_seven_states_tutee(self):
        cases = [
            ("initial_pending", "pending", None),
            ("initial_rejected", "rejected", None),
            ("verified", "approved", "verified"),
            ("renewal_due", "approved", "due"),
            ("renewal_pending", "approved", "pending"),
            ("renewal_rejected", "approved", "rejected"),
            ("not_submitted", None, None),
        ]
        for state, app_status, renewal_status in cases:
            with self.subTest(state=state):
                self._assert_state(self.tutee, state, app_status, renewal_status)

    def test_set_state_works_for_tutor(self):
        self._assert_state(self.tutor, "renewal_due", "approved", "due")
        self._assert_state(self.tutor, "verified", "approved", "verified")

    # --- Enforcement override -----------------------------------------------------------------

    @override_settings(TUTEE_VERIFICATION_ENFORCEMENT_START_DATE=None)
    def test_enforcement_override_flips_gate(self):
        from .views import tutee_verification_enforced
        from . import _verification_dev

        self.assertFalse(tutee_verification_enforced())  # default: unset env
        _verification_dev.enforcement_override_set(True)
        self.assertTrue(tutee_verification_enforced())
        _verification_dev.enforcement_override_set(False)
        self.assertFalse(tutee_verification_enforced())
        _verification_dev.enforcement_override_clear()
        self.assertFalse(tutee_verification_enforced())

    # --- Endpoints ----------------------------------------------------------------------------

    def test_set_state_endpoint_updates_state(self):
        self.client.force_authenticate(user=self.tutee.user)
        resp = self.client.post("/api/dev/verification/set-state/", {"state": "renewal_due"})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["document_renewal_status"], "due")

    def test_set_state_endpoint_rejects_bad_state(self):
        self.client.force_authenticate(user=self.tutee.user)
        resp = self.client.post("/api/dev/verification/set-state/", {"state": "bogus"})
        self.assertEqual(resp.status_code, 400)

    def test_enforcement_endpoint_sets_override(self):
        self.client.force_authenticate(user=self.tutee.user)
        resp = self.client.post("/api/dev/verification/enforcement/", {"mode": "on"})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.data["tutee_verification_enforced"])
        self.assertTrue(resp.data["enforcement_override"])
        self.assertTrue(
            PlatformActivity.objects.filter(
                activity_type="admin_action",
                message__contains="set tutee verification enforcement override to 'on'",
                institution=self.institution,
            ).exists()
        )

    @override_settings(VERIFICATION_DEV_TOOLS_ENABLED=False)
    def test_endpoints_403_when_flag_disabled(self):
        self.client.force_authenticate(user=self.tutee.user)
        self.assertEqual(self.client.get("/api/dev/verification/").status_code, 403)
        self.assertEqual(
            self.client.post("/api/dev/verification/set-state/", {"state": "verified"}).status_code,
            403,
        )

    def test_readout_endpoint_requires_auth(self):
        self.assertEqual(self.client.get("/api/dev/verification/").status_code, 401)


class VerificationEmailWiringTests(APITestCase):
    """Phase 4 email wiring — see docs/plans/2026-07-01-tutee-verification-phase4-email-devtools.md.

    Asserts the right email function is called with the right role_label/status, not the actual
    SMTP send (that's email_utils's own concern and already synchronous/best-effort)."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu", is_active=True,
        )
        self.admin_user = User.objects.create_user(
            username="email-admin@cpu.edu", email="email-admin@cpu.edu", password="password",
        )
        UserProfile.objects.create(
            user=self.admin_user, fname="Admin", mname="", lname="User", role="SuperAdmin",
            institution=self.institution,
        )
        self.tutor_profile = self._make_profile("email-tutor@cpu.edu", "Tutor")
        self.tutee_profile = self._make_profile("email-tutee@cpu.edu", "Tutee")

    def _make_profile(self, email, role):
        user = User.objects.create_user(username=email, email=email, password="password")
        return UserProfile.objects.create(
            user=user, fname="Email", mname="", lname=role, role=role, institution=self.institution,
        )

    def upload(self, name, content, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, content, content_type=content_type)

    def test_tutee_initial_submission_sends_received_email_with_tutee_label(self):
        self.client.force_authenticate(user=self.tutee_profile.user)
        with patch("studybuddy.views.send_application_received_email") as mock_send:
            response = self.client.post(
                "/api/tutee-application/resubmit/",
                {
                    "school_id": self.upload("id.jpg", b"id", "image/jpeg"),
                    "enrollment_proof": self.upload("rf.pdf", b"rf", "application/pdf"),
                },
                format="multipart",
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once_with(self.tutee_profile, role_label='tutee')

    def test_tutee_resubmission_sends_received_email_with_tutee_label(self):
        TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("old-id.jpg", b"old-id", "image/jpeg"),
            enrollment_proof=self.upload("old-rf.pdf", b"old-rf", "application/pdf"),
            application_status="rejected",
            rejection_reason="Blurry photo.",
        )
        self.client.force_authenticate(user=self.tutee_profile.user)
        with patch("studybuddy.views.send_application_received_email") as mock_send:
            response = self.client.post(
                "/api/tutee-application/resubmit/",
                {
                    "school_id": self.upload("new-id.jpg", b"new-id", "image/jpeg"),
                    "enrollment_proof": self.upload("new-rf.pdf", b"new-rf", "application/pdf"),
                },
                format="multipart",
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once_with(self.tutee_profile, role_label='tutee')

    def test_admin_approve_tutor_application_sends_approved_email_with_tutor_label(self):
        application = TutorApplication.objects.create(
            profile=self.tutor_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)
        with patch("studybuddy.admin_views.enqueue_application_approved_email") as mock_send:
            response = self.client.patch(
                f"/api/admin/tutor-applications/{application.id}/",
                {"application_status": "approved"},
                format="json",
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once_with(self.tutor_profile, role_label='tutor')

    def test_admin_reject_tutee_application_sends_rejected_email_with_tutee_label(self):
        application = TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)
        with patch("studybuddy.admin_views.enqueue_application_rejected_email") as mock_send:
            response = self.client.patch(
                f"/api/admin/tutee-applications/{application.id}/",
                {"application_status": "rejected", "rejection_reason": "Illegible ID."},
                format="json",
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once_with(self.tutee_profile, "Illegible ID.", role_label='tutee')

    def test_admin_approve_tutor_renewal_sends_renewal_result_email(self):
        application = TutorApplication.objects.create(
            profile=self.tutor_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
        )
        renewal = TutorDocumentRenewalReview.objects.create(
            application=application,
            profile=self.tutor_profile,
            school_id=self.upload("r-id.jpg", b"r-id", "image/jpeg"),
            enrollment_proof=self.upload("r-rf.pdf", b"r-rf", "application/pdf"),
            status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)
        with patch("studybuddy.admin_views.enqueue_document_renewal_result_email") as mock_send:
            response = self.client.patch(
                f"/api/admin/tutor-document-renewals/{renewal.id}/",
                {"application_status": "approved"},
                format="json",
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once_with(self.tutor_profile, 'tutor', 'approved', '')

    def test_admin_reject_tutee_renewal_sends_renewal_result_email(self):
        application = TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
        )
        renewal = TuteeDocumentRenewalReview.objects.create(
            application=application,
            profile=self.tutee_profile,
            school_id=self.upload("r-id.jpg", b"r-id", "image/jpeg"),
            enrollment_proof=self.upload("r-rf.pdf", b"r-rf", "application/pdf"),
            status="pending",
        )
        self.client.force_authenticate(user=self.admin_user)
        with patch("studybuddy.admin_views.enqueue_document_renewal_result_email") as mock_send:
            response = self.client.patch(
                f"/api/admin/tutee-document-renewals/{renewal.id}/",
                {"application_status": "rejected", "rejection_reason": "Expired ID."},
                format="json",
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once_with(self.tutee_profile, 'tutee', 'rejected', 'Expired ID.')


class EmailUtilsRoleLabelTests(APITestCase):
    """email_utils role_label generalization — default stays byte-identical to the original
    tutor-worded text; role_label='tutee' swaps the wording.

    Patches send_mail directly (not just EMAIL_BACKEND) because _get_smtp_connection() builds its
    own dedicated SMTP connection straight from settings.EMAIL_HOST_* whenever those are configured
    (they are, in this dev .env) — bypassing EMAIL_BACKEND overrides entirely. Asserting via
    mail.outbox would silently attempt real outbound SMTP connections during the test run."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu", is_active=True,
        )
        user = User.objects.create_user(
            username="wording@cpu.edu", email="wording@cpu.edu", password="password",
        )
        self.profile = UserProfile.objects.create(
            user=user, fname="Word", mname="", lname="Ing", role="Tutor", institution=self.institution,
        )

    def test_received_email_defaults_to_tutor_wording(self):
        from .email_utils import send_application_received_email

        with patch("studybuddy.email_utils.send_mail") as mock_send:
            send_application_received_email(self.profile)
        subject, message = mock_send.call_args[0][0], mock_send.call_args[0][1]
        self.assertEqual(subject, "Your StudyBuddy Tutor Application")
        self.assertIn("applying to be a tutor on StudyBuddy", message)

    def test_received_email_tutee_label_swaps_wording(self):
        from .email_utils import send_application_received_email

        with patch("studybuddy.email_utils.send_mail") as mock_send:
            send_application_received_email(self.profile, role_label='tutee')
        subject, message = mock_send.call_args[0][0], mock_send.call_args[0][1]
        self.assertEqual(subject, "Your StudyBuddy Tutee Application")
        self.assertIn("applying to be a tutee on StudyBuddy", message)

    def test_renewal_result_email_approved(self):
        from .email_utils import send_document_renewal_result_email

        with patch("studybuddy.email_utils.send_mail") as mock_send:
            send_document_renewal_result_email(self.profile, 'tutee', 'approved')
        subject, message = mock_send.call_args[0][0], mock_send.call_args[0][1]
        self.assertIn("Approved", subject)
        self.assertIn("tutee verification is current again", message)

    def test_renewal_result_email_rejected_includes_reason(self):
        from .email_utils import send_document_renewal_result_email

        with patch("studybuddy.email_utils.send_mail") as mock_send:
            send_document_renewal_result_email(self.profile, 'tutor', 'rejected', reason="Blurry scan.")
        message = mock_send.call_args[0][1]
        self.assertIn("Blurry scan.", message)


class EmailDeliveryDisabledTests(APITestCase):
    """EMAIL_DELIVERY_DISABLED is the single kill switch for outbound mail. It used to be implied
    by IS_DEMO_DEPLOYMENT (Render's free tier blocked outbound SMTP); the demo now runs on a paid
    instance, so delivery defaults to on and the switch is an explicit opt-out.
    See docs/plans/2026-08-05-enable-demo-email-delivery.md."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu", is_active=True,
        )
        self.user = User.objects.create_user(
            username="killswitch@cpu.edu", email="killswitch@cpu.edu", password="password",
        )
        self.profile = UserProfile.objects.create(
            user=self.user, fname="Kill", mname="", lname="Switch", role="Tutor",
            institution=self.institution,
        )

    @override_settings(EMAIL_DELIVERY_DISABLED=True)
    def test_application_email_suppressed_when_disabled(self):
        from .email_utils import send_application_received_email

        with patch("studybuddy.email_utils.send_mail") as mock_send:
            send_application_received_email(self.profile)
        mock_send.assert_not_called()

    @override_settings(EMAIL_DELIVERY_DISABLED=False)
    def test_application_email_delivered_when_enabled(self):
        from .email_utils import send_application_received_email

        with patch("studybuddy.email_utils.send_mail") as mock_send:
            send_application_received_email(self.profile)
        mock_send.assert_called_once()

    @override_settings(EMAIL_DELIVERY_DISABLED=True)
    def test_login_otp_suppressed_when_disabled(self):
        from . import mailer

        with patch("studybuddy.mailer._deliver") as mock_deliver:
            mailer.send_login_otp(self.user, "123456")
        mock_deliver.assert_not_called()

    @override_settings(EMAIL_DELIVERY_DISABLED=False)
    def test_login_otp_delivered_when_enabled(self):
        from . import mailer

        with patch("studybuddy.mailer._deliver") as mock_deliver:
            mailer.send_login_otp(self.user, "123456")
        mock_deliver.assert_called_once()
        self.assertEqual(mock_deliver.call_args.kwargs["recipient"], self.user.email)


class RenewalReminderTests(APITestCase):
    """Opportunistic 7-day/1-day renewal reminders, triggered from get_document_review_context via
    profile_status. See docs/plans/2026-07-01-tutee-verification-phase4-email-devtools.md."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu", is_active=True,
        )
        self.tutor_profile = self._make_profile("reminder-tutor@cpu.edu", "Tutor")
        self.tutee_profile = self._make_profile("reminder-tutee@cpu.edu", "Tutee")

    def _make_profile(self, email, role):
        user = User.objects.create_user(username=email, email=email, password="password")
        return UserProfile.objects.create(
            user=user, fname="Rem", mname="", lname=role, role=role, institution=self.institution,
        )

    def upload(self, name, content, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, content, content_type=content_type)

    def _approved_application(self, profile, reviewed_at):
        model = TutorApplication if profile.role == "Tutor" else TuteeApplication
        return model.objects.create(
            profile=profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
            reviewed_at=reviewed_at,
        )

    def test_1day_window_enqueues_reminder_and_stamps_dedup_field(self):
        interval = TutorApplication.DOCUMENT_RENEWAL_INTERVAL_DAYS
        reviewed_at = timezone.now() - timedelta(days=interval) + timedelta(hours=12)
        application = self._approved_application(self.tutor_profile, reviewed_at)

        with patch("studybuddy.views.mailer.enqueue_document_renewal_reminder") as mock_enqueue:
            from .views import get_document_review_context

            get_document_review_context(application, 'tutor')

        mock_enqueue.assert_called_once()
        args = mock_enqueue.call_args[0]
        self.assertEqual(args[0], self.tutor_profile)
        self.assertEqual(args[1], 'tutor')
        self.assertEqual(args[2], 1)
        application.refresh_from_db()
        self.assertIsNotNone(application.reminder_1day_sent_at)
        self.assertIsNone(application.reminder_7day_sent_at)

    def test_7day_window_enqueues_reminder_when_1day_window_not_reached(self):
        interval = TutorApplication.DOCUMENT_RENEWAL_INTERVAL_DAYS
        reviewed_at = timezone.now() - timedelta(days=interval) + timedelta(days=4)
        application = self._approved_application(self.tutee_profile, reviewed_at)

        with patch("studybuddy.views.mailer.enqueue_document_renewal_reminder") as mock_enqueue:
            from .views import get_document_review_context

            get_document_review_context(application, 'tutee')

        mock_enqueue.assert_called_once()
        args = mock_enqueue.call_args[0]
        self.assertEqual(args[1], 'tutee')
        self.assertEqual(args[2], 7)
        application.refresh_from_db()
        self.assertIsNotNone(application.reminder_7day_sent_at)

    def test_no_reenqueue_when_dedup_field_already_set(self):
        interval = TutorApplication.DOCUMENT_RENEWAL_INTERVAL_DAYS
        reviewed_at = timezone.now() - timedelta(days=interval) + timedelta(hours=12)
        application = self._approved_application(self.tutor_profile, reviewed_at)
        application.reminder_1day_sent_at = timezone.now()
        application.save(update_fields=['reminder_1day_sent_at'])

        with patch("studybuddy.views.mailer.enqueue_document_renewal_reminder") as mock_enqueue:
            from .views import get_document_review_context

            get_document_review_context(application, 'tutor')

        mock_enqueue.assert_not_called()

    def test_no_reminder_outside_verified_status(self):
        interval = TutorApplication.DOCUMENT_RENEWAL_INTERVAL_DAYS
        due_reviewed_at = timezone.now() - timedelta(days=interval + 1)
        application = self._approved_application(self.tutor_profile, due_reviewed_at)
        self.assertEqual(application.document_renewal_status(), 'due')

        with patch("studybuddy.views.mailer.enqueue_document_renewal_reminder") as mock_enqueue:
            from .views import get_document_review_context

            get_document_review_context(application, 'tutor')

        mock_enqueue.assert_not_called()

    def test_no_reminder_when_not_yet_in_any_window(self):
        application = self._approved_application(self.tutor_profile, timezone.now())
        self.assertEqual(application.document_renewal_status(), 'verified')

        with patch("studybuddy.views.mailer.enqueue_document_renewal_reminder") as mock_enqueue:
            from .views import get_document_review_context

            get_document_review_context(application, 'tutor')

        mock_enqueue.assert_not_called()


class VerificationDevToolsAdminEndpointTests(APITestCase):
    """AdminUserVerificationDevToolsView — see docs/plans/2026-07-01-tutee-verification-phase4-email-devtools.md."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu", is_active=True,
        )
        self.super_user = User.objects.create_user(
            username="devtools-super@cpu.edu", email="devtools-super@cpu.edu", password="password",
        )
        self.super_profile = UserProfile.objects.create(
            user=self.super_user, fname="Super", mname="", lname="Admin", role="SuperAdmin",
            institution=self.institution,
        )
        self.admin_user = User.objects.create_user(
            username="devtools-admin@cpu.edu", email="devtools-admin@cpu.edu", password="password",
        )
        UserProfile.objects.create(
            user=self.admin_user, fname="Regular", mname="", lname="Admin", role="Admin",
            institution=self.institution,
        )
        self.tutee_user = User.objects.create_user(
            username="devtools-tutee@cpu.edu", email="devtools-tutee@cpu.edu", password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user, fname="Dev", mname="", lname="Tutee", role="Tutee",
            institution=self.institution,
        )

    def upload(self, name, content, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, content, content_type=content_type)

    @override_settings(VERIFICATION_DEV_TOOLS_ENABLED=True)
    def test_403_for_regular_admin_even_with_flag_on(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.post(
            f"/api/admin/users/{self.tutee_profile.id}/verification-dev-tools/",
            {"action": "send_received"},
        )
        self.assertEqual(response.status_code, 403)

    @override_settings(VERIFICATION_DEV_TOOLS_ENABLED=False)
    def test_403_for_superadmin_when_flag_off(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.post(
            f"/api/admin/users/{self.tutee_profile.id}/verification-dev-tools/",
            {"action": "send_received"},
        )
        self.assertEqual(response.status_code, 403)

    @override_settings(VERIFICATION_DEV_TOOLS_ENABLED=True)
    def test_400_for_invalid_action(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.post(
            f"/api/admin/users/{self.tutee_profile.id}/verification-dev-tools/",
            {"action": "bogus"},
        )
        self.assertEqual(response.status_code, 400)

    @override_settings(VERIFICATION_DEV_TOOLS_ENABLED=True)
    def test_send_approved_calls_email_function(self):
        TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
        )
        self.client.force_authenticate(user=self.super_user)
        with patch("studybuddy.admin_views.send_application_approved_email") as mock_send:
            response = self.client.post(
                f"/api/admin/users/{self.tutee_profile.id}/verification-dev-tools/",
                {"action": "send_approved"},
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once_with(self.tutee_profile, role_label='tutee')

    @override_settings(VERIFICATION_DEV_TOOLS_ENABLED=True)
    def test_force_expire_flips_renewal_status_to_due(self):
        application = TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
        )
        self.client.force_authenticate(user=self.super_user)
        response = self.client.post(
            f"/api/admin/users/{self.tutee_profile.id}/verification-dev-tools/",
            {"action": "force_expire"},
        )
        self.assertEqual(response.status_code, 200)
        application.refresh_from_db()
        self.assertEqual(application.document_renewal_status(), 'due')

    @override_settings(VERIFICATION_DEV_TOOLS_ENABLED=True)
    def test_404_for_send_action_when_no_application(self):
        self.client.force_authenticate(user=self.super_user)
        response = self.client.post(
            f"/api/admin/users/{self.tutee_profile.id}/verification-dev-tools/",
            {"action": "send_received"},
        )
        self.assertEqual(response.status_code, 404)

    @override_settings(VERIFICATION_DEV_TOOLS_ENABLED=True)
    def test_400_for_send_rejected_when_no_rejection_reason(self):
        TuteeApplication.objects.create(
            profile=self.tutee_profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="pending",
        )
        self.client.force_authenticate(user=self.super_user)
        response = self.client.post(
            f"/api/admin/users/{self.tutee_profile.id}/verification-dev-tools/",
            {"action": "send_rejected"},
        )
        self.assertEqual(response.status_code, 400)


class TutorDetailVerifiedBadgeTests(APITestCase):
    """Public tutor-detail page's is_verified field — binary, hidden unless approved + verified.
    See docs/plans/2026-07-01-tutee-verification-phase4-email-devtools.md (Section 6)."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu", is_active=True,
        )
        self.viewer_user = User.objects.create_user(
            username="badge-viewer@cpu.edu", email="badge-viewer@cpu.edu", password="password",
        )
        UserProfile.objects.create(
            user=self.viewer_user, fname="Viewer", mname="", lname="Tutee", role="Tutee",
            institution=self.institution,
        )

    def upload(self, name, content, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, content, content_type=content_type)

    def _make_tutor(self, email):
        user = User.objects.create_user(username=email, email=email, password="password")
        profile = UserProfile.objects.create(
            user=user, fname="Badge", mname="", lname="Tutor", role="Tutor", institution=self.institution,
        )
        Tutor.objects.create(profile=profile)
        return profile

    def test_is_verified_false_when_no_application(self):
        profile = self._make_tutor("badge-noapp@cpu.edu")
        self.client.force_authenticate(user=self.viewer_user)
        response = self.client.get(f"/api/tutors/{profile.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["is_verified"])

    def test_is_verified_true_when_approved_and_current(self):
        profile = self._make_tutor("badge-verified@cpu.edu")
        TutorApplication.objects.create(
            profile=profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
        )
        self.client.force_authenticate(user=self.viewer_user)
        response = self.client.get(f"/api/tutors/{profile.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["is_verified"])

    def test_is_verified_false_when_pending(self):
        profile = self._make_tutor("badge-pending@cpu.edu")
        TutorApplication.objects.create(
            profile=profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="pending",
        )
        self.client.force_authenticate(user=self.viewer_user)
        response = self.client.get(f"/api/tutors/{profile.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["is_verified"])

    def test_is_verified_false_when_renewal_due(self):
        interval = TutorApplication.DOCUMENT_RENEWAL_INTERVAL_DAYS
        profile = self._make_tutor("badge-renewaldue@cpu.edu")
        TutorApplication.objects.create(
            profile=profile,
            school_id=self.upload("id.jpg", b"id", "image/jpeg"),
            enrollment_proof=self.upload("rf.pdf", b"rf", "application/pdf"),
            application_status="approved",
            reviewed_at=timezone.now() - timedelta(days=interval + 1),
        )
        self.client.force_authenticate(user=self.viewer_user)
        response = self.client.get(f"/api/tutors/{profile.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["is_verified"])

    def test_includes_institution_name(self):
        profile = self._make_tutor("badge-institution@cpu.edu")
        self.client.force_authenticate(user=self.viewer_user)
        response = self.client.get(f"/api/tutors/{profile.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["institution_name"], "CPU")


class TutorOnboardingSearchVisibilityTests(APITestCase):
    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu.ph", is_active=True,
        )
        self.course = Course.objects.create(
            course_code="ONBOARD", course_name="Onboarding Course",
        )
        self.subject = Subjects.objects.create(
            subject_code="ONBOARD101", subject_name="Onboarding Search", department="Education",
            category=self.course.course_code,
        )
        tutee_user = User.objects.create_user(
            username="onboarding-tutee@cpu.edu.ph", email="onboarding-tutee@cpu.edu.ph",
            password="password",
        )
        self.tutee = UserProfile.objects.create(
            user=tutee_user, fname="Search", mname="", lname="Tutee", role="Tutee",
            institution=self.institution, course=self.course,
        )

    def upload(self, name, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, b"document", content_type=content_type)

    def make_tutor(self, label, application_status=None, renewal_due=False):
        user = User.objects.create_user(
            username=f"{label}@cpu.edu.ph", email=f"{label}@cpu.edu.ph", password="password",
        )
        profile = UserProfile.objects.create(
            user=user, fname=label.title(), mname="", lname="Tutor", role="Tutor",
            institution=self.institution,
        )
        tutor = Tutor.objects.create(
            profile=profile, hourly_rate=200, can_online=True, teaching_level="College",
        )
        TutorSubjects.objects.create(tutor=tutor, subject=self.subject, expertise_level=3)
        if application_status:
            reviewed_at = None
            if renewal_due:
                reviewed_at = timezone.now() - timedelta(
                    days=TutorApplication.DOCUMENT_RENEWAL_INTERVAL_DAYS + 1
                )
            TutorApplication.objects.create(
                profile=profile,
                school_id=self.upload(f"{label}-id.jpg", "image/jpeg"),
                enrollment_proof=self.upload(f"{label}-proof.pdf", "application/pdf"),
                application_status=application_status,
                reviewed_at=reviewed_at,
            )
        return tutor

    def test_recommendations_only_include_approved_initial_applications(self):
        no_application = self.make_tutor("no-application")
        pending = self.make_tutor("pending", "pending")
        rejected = self.make_tutor("rejected", "rejected")
        approved = self.make_tutor("approved", "approved")
        renewal_due = self.make_tutor("renewal-due", "approved", renewal_due=True)
        self.client.force_authenticate(user=self.tutee.user)

        response = self.client.post(
            "/api/recommend-tutors/",
            {"subject": self.subject.subject_code, "date": timezone.localdate().isoformat()},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        tutor_ids = {row["id"] for row in response.data["tutors"]}
        self.assertNotIn(no_application.profile_id, tutor_ids)
        self.assertNotIn(pending.profile_id, tutor_ids)
        self.assertNotIn(rejected.profile_id, tutor_ids)
        self.assertIn(approved.profile_id, tutor_ids)
        self.assertIn(renewal_due.profile_id, tutor_ids)


class TutorOnboardingStateTests(APITestCase):
    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu.ph", is_active=True,
        )
        self.tutor_user = User.objects.create_user(
            username="state-tutor@cpu.edu.ph", email="state-tutor@cpu.edu.ph", password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="State", mname="", lname="Tutor", role="Tutor",
            institution=self.institution,
        )
        self.tutor = Tutor.objects.create(profile=self.tutor_profile)
        self.tutee_user = User.objects.create_user(
            username="state-tutee@cpu.edu.ph", email="state-tutee@cpu.edu.ph", password="password",
        )
        UserProfile.objects.create(
            user=self.tutee_user, fname="State", mname="", lname="Tutee", role="Tutee",
            institution=self.institution,
        )

    def upload(self, name, content_type):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(name, b"document", content_type=content_type)

    def test_profile_status_exposes_initial_onboarding_state(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.get("/api/profile/status/")

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.data["tutor_onboarding_skipped_at"])
        self.assertFalse(response.data["tutor_onboarding_complete"])
        self.assertEqual(response.data["tutor_subject_count"], 0)
        self.assertFalse(response.data["tutor_subjects_completed"])

    def test_tutor_can_skip_initial_verification(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post("/api/tutor/onboarding/skip-verification/")

        self.assertEqual(response.status_code, 200)
        self.tutor_profile.refresh_from_db()
        self.assertIsNotNone(self.tutor_profile.tutor_onboarding_skipped_at)
        self.assertTrue(response.data["tutor_onboarding_complete"])

    def test_non_tutor_cannot_skip_tutor_verification(self):
        self.client.force_authenticate(user=self.tutee_user)

        response = self.client.post("/api/tutor/onboarding/skip-verification/")

        self.assertEqual(response.status_code, 403)

    def test_initial_submission_creates_pending_application(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            "/api/tutor-application/submit/",
            {
                "school_id": self.upload("id.jpg", "image/jpeg"),
                "enrollment_proof": self.upload("proof.pdf", "application/pdf"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        application = TutorApplication.objects.get(profile=self.tutor_profile)
        self.assertEqual(application.application_status, "pending")
        self.tutor_profile.refresh_from_db()
        self.assertIsNone(self.tutor_profile.tutor_onboarding_skipped_at)

    def test_initial_submission_attaches_pre_application_subject_proposals(self):
        subject = Subjects.objects.create(
            subject_code="BEFORE-VERIFY",
            subject_name="Before Verify",
            department="Education",
            status="pending",
            proposed_by_tutor=self.tutor,
        )
        TutorSubjects.objects.create(tutor=self.tutor, subject=subject, expertise_level=3)
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.post(
            "/api/tutor-application/submit/",
            {
                "school_id": self.upload("id.jpg", "image/jpeg"),
                "enrollment_proof": self.upload("proof.pdf", "application/pdf"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        subject.refresh_from_db()
        self.assertEqual(subject.proposed_application.profile, self.tutor_profile)

    @override_settings(LOGIN_OTP_DISABLED=True)
    def test_pending_tutor_application_does_not_block_login(self):
        TutorApplication.objects.create(
            profile=self.tutor_profile,
            school_id="tutor_applications/school_ids/id.jpg",
            enrollment_proof="tutor_applications/enrollment_proofs/proof.pdf",
            application_status="pending",
        )

        response = self.client.post(
            "/api/login/",
            {"email": self.tutor_user.email, "password": "password"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)

    def test_tutor_registration_creates_account_without_application(self):
        self.client.force_authenticate(user=None)

        response = self.client.post(
            "/api/register/",
            {
                "fname": "New",
                "mname": "",
                "lname": "Tutor",
                "email": "new-tutor@cpu.edu.ph",
                "password": "StrongPassword123!",
                "role": "Tutor",
                "institution_id": self.institution.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        profile = UserProfile.objects.get(user__email="new-tutor@cpu.edu.ph")
        self.assertTrue(Tutor.objects.filter(profile=profile).exists())
        self.assertFalse(TutorApplication.objects.filter(profile=profile).exists())


class TutorSubjectProposalTests(APITestCase):
    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu.ph", is_active=True,
        )
        self.catalog_subject = Subjects.objects.create(
            subject_code="MATH101", subject_name="College Algebra", department="Mathematics",
            category="Mathematics & Data Sciences",
        )
        self.course = Course.objects.create(course_code="PROPOSAL", course_name="Proposal Course")
        self.tutor_user = User.objects.create_user(
            username="proposal-tutor@cpu.edu.ph", email="proposal-tutor@cpu.edu.ph",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Proposal", mname="", lname="Tutor", role="Tutor",
            institution=self.institution, course=self.course,
        )
        self.tutor = Tutor.objects.create(profile=self.tutor_profile)
        self.application = TutorApplication.objects.create(
            profile=self.tutor_profile,
            school_id="tutor_applications/school_ids/id.jpg",
            enrollment_proof="tutor_applications/enrollment_proofs/proof.pdf",
        )
        self.client.force_authenticate(user=self.tutor_user)

    def propose(self, subject_name="Educational Data Visualization",
                category="Mathematics & Data Sciences"):
        return self.client.post(
            "/api/tutor/subjects/propose/",
            {
                "subject_name": subject_name,
                "category": category,
                "description": "A proposed specialty.",
            },
            format="json",
        )

    def test_proposal_creates_pending_subject_and_tutor_link(self):
        with patch("studybuddy.views.subject_is_recognized_for_profile") as recognition_check:
            response = self.propose()

        self.assertEqual(response.status_code, 201)
        recognition_check.assert_not_called()
        subject = Subjects.objects.get(subject_code="EDUCATIONAL-DATA-VIS")
        self.assertEqual(subject.status, "pending")
        self.assertEqual(subject.proposed_by_tutor, self.tutor)
        self.assertEqual(subject.proposed_application, self.application)
        self.assertTrue(TutorSubjects.objects.filter(tutor=self.tutor, subject=subject).exists())

    def test_proposal_rejects_unknown_category(self):
        response = self.propose(category="Invented Category")

        self.assertEqual(response.status_code, 400)

    def test_proposal_persists_keywords_when_provided(self):
        with patch("studybuddy.views.subject_is_recognized_for_profile"):
            response = self.client.post(
                "/api/tutor/subjects/propose/",
                {
                    "subject_name": "Quantum Computing",
                    "category": "Technology & Computer Science",
                    "description": "A proposed specialty.",
                    "keywords": "quantum, qubits",
                },
                format="json",
            )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["keywords"], "quantum, qubits")
        subject = Subjects.objects.get(subject_code="QUANTUM-COMPUTING")
        self.assertEqual(subject.keywords, "quantum, qubits")

    def test_proposal_defaults_keywords_to_blank(self):
        with patch("studybuddy.views.subject_is_recognized_for_profile"):
            response = self.propose()

        self.assertEqual(response.status_code, 201)
        subject = Subjects.objects.get(subject_code=response.data["subject_code"])
        self.assertEqual(subject.keywords, "")

    def test_proposal_enforces_eight_subject_cap_across_statuses(self):
        for index in range(8):
            subject = Subjects.objects.create(
                subject_code=f"CAP{index}", subject_name=f"Cap Subject {index}",
                department="Mathematics", status="pending" if index % 2 else "approved",
            )
            TutorSubjects.objects.create(tutor=self.tutor, subject=subject, expertise_level=3)

        response = self.propose()

        self.assertEqual(response.status_code, 400)
        self.assertIn("8 subjects", response.data["error"])

    def test_generated_codes_dedupe_collisions(self):
        first = self.propose(subject_name="Special Topics")
        second = self.propose(subject_name="Special Topics")

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)
        self.assertTrue(Subjects.objects.filter(subject_code="SPECIAL-TOPICS").exists())
        self.assertTrue(Subjects.objects.filter(subject_code="SPECIAL-TOPICS-2").exists())

    def test_pending_subject_is_hidden_from_catalog(self):
        proposal = self.propose()
        self.assertEqual(proposal.status_code, 201)

        response = self.client.get("/api/subjects/", {"catalog_scope": "all"})

        self.assertEqual(response.status_code, 200)
        codes = {row["subject_code"] for row in response.data}
        self.assertIn(self.catalog_subject.subject_code, codes)
        self.assertNotIn(proposal.data["subject_code"], codes)

    def test_catalog_search_is_server_filtered(self):
        response = self.client.get("/api/subjects/", {"search": "College Algebra"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [row["subject_code"] for row in response.data],
            [self.catalog_subject.subject_code],
        )

    def test_selection_is_not_scoped_to_the_requesting_users_course(self):
        # Course-based subject gating was retired: a BSCS-course tutor (self.course above is
        # unrelated to any taxonomy category) can still select/search a subject from a wholly
        # different field, e.g. Hobbies & Arts, proving selection is catalog-wide now.
        unrelated_subject = Subjects.objects.create(
            subject_code="guitar-101", subject_name="Guitar Basics", department="Instruments",
            category="Hobbies & Arts",
        )

        response = self.client.get("/api/subjects/", {"search": "Guitar"})

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            unrelated_subject.subject_code,
            [row["subject_code"] for row in response.data],
        )


class AdminProposedSubjectReviewTests(APITestCase):
    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu.ph", is_active=True,
        )
        self.admin_user = User.objects.create_user(
            username="proposal-admin@cpu.edu.ph", email="proposal-admin@cpu.edu.ph",
            password="password",
        )
        UserProfile.objects.create(
            user=self.admin_user, fname="Proposal", mname="", lname="Admin", role="SuperAdmin",
            institution=self.institution,
        )
        tutor_user = User.objects.create_user(
            username="review-tutor@cpu.edu.ph",
            email="review-tutor@cpu.edu.ph",
            password="password",
        )
        tutor_profile = UserProfile.objects.create(
            user=tutor_user, fname="Review", mname="", lname="Tutor", role="Tutor",
            institution=self.institution,
        )
        self.tutor = Tutor.objects.create(profile=tutor_profile)
        self.application = TutorApplication.objects.create(
            profile=tutor_profile,
            school_id="tutor_applications/school_ids/id.jpg",
            enrollment_proof="tutor_applications/enrollment_proofs/proof.pdf",
            application_status="rejected",
        )
        self.client.force_authenticate(user=self.admin_user)

    def make_proposal(self, code, keywords="", tutor_note=""):
        subject = Subjects.objects.create(
            subject_code=code, subject_name=f"Proposal {code}", department="Mathematics",
            status="pending", proposed_by_tutor=self.tutor,
            proposed_application=self.application, keywords=keywords,
        )
        TutorSubjects.objects.create(
            tutor=self.tutor, subject=subject, expertise_level=3, description=tutor_note,
        )
        return subject

    def make_selected_subject(self, code, name, category="Natural Sciences"):
        """An approved catalog subject the tutor picked, as opposed to proposed."""
        subject = Subjects.objects.create(
            subject_code=code, subject_name=name, department="Sciences",
            category=category, status="approved",
        )
        TutorSubjects.objects.create(tutor=self.tutor, subject=subject, expertise_level=4)
        return subject

    def test_application_detail_includes_pending_proposed_subjects(self):
        proposal = self.make_proposal("DETAIL-PROP", keywords="coding, cs")

        response = self.client.get(f"/api/admin/tutor-applications/{self.application.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.data["proposed_subjects"][0]["subject_code"],
            proposal.subject_code,
        )
        self.assertEqual(response.data["proposed_subjects"][0]["keywords"], "coding, cs")

    def test_approve_subject_keeps_tutor_link_regardless_of_application_status(self):
        proposal = self.make_proposal("APPROVE-PROP")

        response = self.client.patch(
            f"/api/admin/tutor-applications/{self.application.id}/subjects/"
            f"{proposal.subject_code}/",
            {"action": "approve"}, format="json",
        )

        self.assertEqual(response.status_code, 200)
        proposal.refresh_from_db()
        self.assertEqual(proposal.status, "approved")
        self.assertTrue(TutorSubjects.objects.filter(tutor=self.tutor, subject=proposal).exists())
        self.application.refresh_from_db()
        self.assertEqual(self.application.application_status, "rejected")

    def test_update_action_writes_catalog_description_not_the_tutor_note(self):
        """The admin form edits Subjects.description (global catalog copy, which
        the subject pickers search). The tutor's own note is evidence, not copy,
        and must survive the edit untouched."""
        proposal = self.make_proposal("UPDATE-PROP", tutor_note="I tutored this for two terms.")

        response = self.client.patch(
            f"/api/admin/tutor-applications/{self.application.id}/subjects/"
            f"{proposal.subject_code}/",
            {
                "action": "update",
                "subject_name": "Updated Proposal Name",
                "category": "Mathematics & Data Sciences",
                "keywords": "algebra, math",
                "catalog_description": "Refined by admin.",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        proposal.refresh_from_db()
        self.assertEqual(proposal.subject_name, "Updated Proposal Name")
        self.assertEqual(proposal.category, "Mathematics & Data Sciences")
        self.assertEqual(proposal.keywords, "algebra, math")
        self.assertEqual(proposal.status, "pending")
        self.assertEqual(proposal.description, "Refined by admin.")
        link = TutorSubjects.objects.get(tutor=self.tutor, subject=proposal)
        self.assertEqual(link.description, "I tutored this for two terms.")
        self.assertEqual(response.data["tutor_note"], "I tutored this for two terms.")

    def test_update_action_leaves_catalog_description_alone_when_omitted(self):
        proposal = self.make_proposal("UPDATE-NO-DESC")
        proposal.description = "Existing catalog copy."
        proposal.save(update_fields=["description"])

        response = self.client.patch(
            f"/api/admin/tutor-applications/{self.application.id}/subjects/"
            f"{proposal.subject_code}/",
            {
                "action": "update",
                "subject_name": "Renamed Only",
                "category": "Mathematics & Data Sciences",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        proposal.refresh_from_db()
        self.assertEqual(proposal.description, "Existing catalog copy.")

    def test_application_detail_separates_catalog_description_from_tutor_note(self):
        proposal = self.make_proposal("SPLIT-PROP", tutor_note="My own note.")
        proposal.description = "Catalog copy."
        proposal.save(update_fields=["description"])

        response = self.client.get(f"/api/admin/tutor-applications/{self.application.id}/")

        self.assertEqual(response.status_code, 200)
        row = response.data["proposed_subjects"][0]
        self.assertEqual(row["catalog_description"], "Catalog copy.")
        self.assertEqual(row["tutor_note"], "My own note.")

    def test_application_detail_includes_selected_catalog_subjects(self):
        self.make_proposal("SELECTED-PROP")
        selected = self.make_selected_subject("BIO-101", "General Biology")

        response = self.client.get(f"/api/admin/tutor-applications/{self.application.id}/")

        self.assertEqual(response.status_code, 200)
        codes = [row["subject_code"] for row in response.data["selected_subjects"]]
        self.assertEqual(codes, [selected.subject_code])
        self.assertEqual(response.data["selected_subjects"][0]["subject_name"], "General Biology")
        # Pending proposals are the complement and must not leak into this list.
        self.assertNotIn("SELECTED-PROP", codes)

    def test_selected_subjects_is_empty_when_the_tutor_only_proposed(self):
        self.make_proposal("ONLY-PROP")

        response = self.client.get(f"/api/admin/tutor-applications/{self.application.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["selected_subjects"], [])

    def test_update_action_rejects_unknown_category(self):
        proposal = self.make_proposal("UPDATE-BAD-CAT")

        response = self.client.patch(
            f"/api/admin/tutor-applications/{self.application.id}/subjects/"
            f"{proposal.subject_code}/",
            {
                "action": "update",
                "subject_name": "Still Fine",
                "category": "Invented Category",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)

    def test_reject_subject_deletes_subject_and_tutor_link(self):
        proposal = self.make_proposal("REJECT-PROP")

        response = self.client.patch(
            f"/api/admin/tutor-applications/{self.application.id}/subjects/"
            f"{proposal.subject_code}/",
            {"action": "reject"}, format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Subjects.objects.filter(subject_code=proposal.subject_code).exists())
        self.assertFalse(
            TutorSubjects.objects.filter(
                tutor=self.tutor,
                subject_id=proposal.subject_code,
            ).exists()
        )


class AdminEndpointsRequireSuperAdminTests(APITestCase):
    """Ticket: Admin role removed -> admin endpoints are SuperAdmin-only.

    Representative admin endpoints must reject a Tutor (403), an unauthenticated
    caller (401), and a directly-created role='Admin' profile (simulating an
    unmigrated row, 403), while a SuperAdmin gets 200."""

    ADMIN_ENDPOINTS = (
        "/api/admin/stats/",
        "/api/admin/users/",
        "/api/admin/withdrawals/",
        "/api/admin/analytics/",
    )

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu.ph", is_active=True,
        )
        self.super_user = User.objects.create_user(
            username="perm-super@cpu.edu.ph", email="perm-super@cpu.edu.ph", password="password",
        )
        UserProfile.objects.create(
            user=self.super_user, fname="Perm", mname="", lname="Super", role="SuperAdmin",
            profile_completed=True, is_domain_exempt=True,
        )
        self.tutor_user = User.objects.create_user(
            username="perm-tutor@cpu.edu.ph", email="perm-tutor@cpu.edu.ph", password="password",
        )
        UserProfile.objects.create(
            user=self.tutor_user, fname="Perm", mname="", lname="Tutor", role="Tutor",
            institution=self.institution, profile_completed=True,
        )
        self.legacy_admin_user = User.objects.create_user(
            username="perm-admin@cpu.edu.ph", email="perm-admin@cpu.edu.ph", password="password",
        )
        # Directly stamp the removed 'Admin' role to simulate an unmigrated row.
        UserProfile.objects.create(
            user=self.legacy_admin_user, fname="Perm", mname="", lname="Admin", role="Admin",
            institution=self.institution, profile_completed=True,
        )

    def test_unauthenticated_gets_401(self):
        for url in self.ADMIN_ENDPOINTS:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 401, url)

    def test_tutor_gets_403(self):
        self.client.force_authenticate(user=self.tutor_user)
        for url in self.ADMIN_ENDPOINTS:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 403, url)

    def test_legacy_admin_role_gets_403(self):
        self.client.force_authenticate(user=self.legacy_admin_user)
        for url in self.ADMIN_ENDPOINTS:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 403, url)

    def test_superadmin_gets_200(self):
        self.client.force_authenticate(user=self.super_user)
        for url in self.ADMIN_ENDPOINTS:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)


class AdminToSuperAdminMigrationTests(APITestCase):
    """Ticket: 0072 data migration converts every role='Admin' row to 'SuperAdmin'."""

    def _forward_function(self):
        import importlib

        module = importlib.import_module(
            "studybuddy.migrations.0072_convert_admin_to_superadmin"
        )
        return module.convert_admin_to_superadmin

    def test_migration_converts_admin_rows_to_superadmin(self):
        from django.apps import apps as django_apps

        admin_user = User.objects.create_user(
            username="mig-admin@cpu.edu.ph", email="mig-admin@cpu.edu.ph", password="password",
        )
        # Directly stamp the removed 'Admin' role, bypassing model validation.
        admin_profile = UserProfile.objects.create(
            user=admin_user, fname="Mig", mname="", lname="Admin", role="Admin",
        )
        tutee_user = User.objects.create_user(
            username="mig-tutee@cpu.edu.ph", email="mig-tutee@cpu.edu.ph", password="password",
        )
        tutee_profile = UserProfile.objects.create(
            user=tutee_user, fname="Mig", mname="", lname="Tutee", role="Tutee",
        )

        self._forward_function()(django_apps, None)

        admin_profile.refresh_from_db()
        tutee_profile.refresh_from_db()
        self.assertEqual(admin_profile.role, "SuperAdmin")
        self.assertEqual(tutee_profile.role, "Tutee")
        self.assertFalse(UserProfile.objects.filter(role="Admin").exists())


class SubjectTaxonomyModuleTests(APITestCase):
    """subject_taxonomy.py is the single source of truth for the Preply-style catalog
    (see docs/plans/2026-07-16-subjects-taxonomy-reseed.md). Its slugs must fit the
    Subjects.subject_code column and stay unique; every subject must carry a real
    taxonomy category."""

    def test_slugs_are_unique_and_fit_the_subject_code_column(self):
        from .models import Subjects
        from .subject_taxonomy import SUBJECTS, slugify_subject_code

        max_length = Subjects._meta.get_field('subject_code').max_length
        slugs = [slugify_subject_code(name) for name, _category, _subgroup in SUBJECTS]

        self.assertEqual(len(slugs), len(set(slugs)), "subject slugs must be unique")
        for slug in slugs:
            self.assertLessEqual(len(slug), max_length)
            self.assertTrue(slug, "slug must not be empty")

    def test_every_subject_has_a_valid_category(self):
        from .subject_taxonomy import CATEGORIES, SUBJECTS

        for name, category, _subgroup in SUBJECTS:
            self.assertIn(category, CATEGORIES, f"{name} has an unrecognized category")


class AdminCourseCatalogTaxonomyTests(APITestCase):
    """Admin subject-catalog CRUD on the taxonomy shape: auto-generated slugs, category
    validation, and category-based filtering (replacing the retired course filter)."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu.ph", is_active=True,
        )
        self.admin_user = User.objects.create_user(
            username="catalog-admin@cpu.edu.ph", email="catalog-admin@cpu.edu.ph",
            password="password",
        )
        UserProfile.objects.create(
            user=self.admin_user, fname="Catalog", mname="", lname="Admin", role="SuperAdmin",
            institution=self.institution,
        )
        self.client.force_authenticate(user=self.admin_user)

    def test_create_without_code_generates_a_slug(self):
        response = self.client.post(
            "/api/admin/course-catalog/",
            {
                "subject_name": "Organic Chemistry",
                "category": "Natural Sciences",
                "department": "Chemistry",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["subject_code"], "organic-chemistry")
        self.assertTrue(Subjects.objects.filter(subject_code="organic-chemistry").exists())

    def test_create_rejects_a_category_outside_the_taxonomy(self):
        response = self.client.post(
            "/api/admin/course-catalog/",
            {"subject_name": "Mystery Subject", "category": "Not A Real Category"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(Subjects.objects.filter(subject_name="Mystery Subject").exists())

    def test_list_filters_by_category(self):
        Subjects.objects.create(
            subject_code="calculus", subject_name="Calculus", category="Mathematics & Data Sciences",
        )
        Subjects.objects.create(
            subject_code="guitar", subject_name="Guitar", category="Hobbies & Arts",
        )

        response = self.client.get(
            "/api/admin/course-catalog/", {"category": "Hobbies & Arts"},
        )

        self.assertEqual(response.status_code, 200)
        codes = [row["subject_code"] for row in response.data]
        self.assertEqual(codes, ["guitar"])


class CuratedPersonaCbfOrderingTests(APITestCase):
    """Locks in the seeded demo's headline claim (docs/plans/2026-07-16-subjects-taxonomy-
    reseed.md, curated persona S1/T1-T4): a Python-seeking BSCS tutee ranks an exact-match,
    high-expertise, same-course tutor first, then the weaker exact match, then the
    same-category-only match, then the unrelated tutor — independent of the full seed
    command, by constructing the same shape directly."""

    def setUp(self):
        from .recommender.cbf import compute_cbf_score

        self.compute_cbf_score = compute_cbf_score

        self.bscs = Course.objects.create(course_code="BSCS", course_name="Computer Science")
        self.bsba = Course.objects.create(course_code="BSBA", course_name="Business Administration")

        self.python = Subjects.objects.create(
            subject_code="python", subject_name="Python", category="Technology & Computer Science",
        )
        self.data_structures = Subjects.objects.create(
            subject_code="data-structures", subject_name="Data Structures",
            category="Technology & Computer Science",
        )
        self.cpp = Subjects.objects.create(
            subject_code="cpp", subject_name="C++", category="Technology & Computer Science",
        )
        self.financial_accounting = Subjects.objects.create(
            subject_code="financial-accounting", subject_name="Financial Accounting",
            category="Business, Finance & Economics",
        )

        s1_user = User.objects.create_user(
            username="s1.felipe@cpu.edu.ph", email="s1.felipe@cpu.edu.ph", password="password",
        )
        self.s1 = UserProfile.objects.create(
            user=s1_user, fname="Felipe", lname="Fernandez", role="Tutee",
            year_level=14, course=self.bscs,
        )

        self.t1 = self._make_tutor(
            "t1.marisol", self.bscs, 16, [(self.python, 5), (self.data_structures, 4)],
        )
        self.t4 = self._make_tutor("t4.domingo", self.bscs, 16, [(self.python, 3)])
        self.t2 = self._make_tutor("t2.benigno", self.bscs, 15, [(self.cpp, 3)])
        self.t3 = self._make_tutor(
            "t3.corazon", self.bsba, 15, [(self.financial_accounting, 5)],
        )

    def _make_tutor(self, username, course, year_level, subjects_with_levels):
        user = User.objects.create_user(
            username=f"{username}@cpu.edu.ph", email=f"{username}@cpu.edu.ph", password="password",
        )
        profile = UserProfile.objects.create(
            user=user, fname=username, lname="Tutor", role="Tutor",
            year_level=year_level, course=course,
        )
        tutor = Tutor.objects.create(
            profile=profile, hourly_rate=200, can_online=True, can_f2f=False,
            teaching_level="College",
        )
        for subject, level in subjects_with_levels:
            TutorSubjects.objects.create(tutor=tutor, subject=subject, expertise_level=level)
        return tutor

    def test_s1_ranking_is_t1_then_t4_then_t2_then_t3(self):
        scores = {
            key: self.compute_cbf_score(self.s1, tutor, self.python.subject_code)
            for key, tutor in [("T1", self.t1), ("T4", self.t4), ("T2", self.t2), ("T3", self.t3)]
        }

        self.assertGreater(scores["T1"], scores["T4"])
        self.assertGreater(scores["T4"], scores["T2"])
        self.assertGreater(scores["T2"], scores["T3"])


class PendingProposedSubjectRemainsSelectableTests(APITestCase):
    """Regression test for a bug found while retiring course-based subject gating: a tutor's
    own pending proposed subject must stay visible in their subject-selection queryset, even
    though visible_subject_queryset_for_profile is approved-only (see
    subject_recognition.subject_selection_queryset_for_profile and its include_current
    handling in SubjectListView.get_queryset)."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU", school_email_domain="cpu.edu.ph", is_active=True,
        )
        self.tutor_user = User.objects.create_user(
            username="pending-subj-tutor@cpu.edu.ph", email="pending-subj-tutor@cpu.edu.ph",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user, fname="Pending", mname="", lname="Tutor", role="Tutor",
            institution=self.institution,
        )
        self.tutor = Tutor.objects.create(profile=self.tutor_profile)
        self.pending_subject = Subjects.objects.create(
            subject_code="pending-subj", subject_name="Pending Subject",
            category="Hobbies & Arts", status="pending",
        )
        TutorSubjects.objects.create(
            tutor=self.tutor, subject=self.pending_subject, expertise_level=3,
        )

    def test_selection_queryset_includes_own_pending_subject(self):
        from .subject_recognition import subject_selection_queryset_for_profile

        queryset, _recognized = subject_selection_queryset_for_profile(
            self.tutor_profile, include_current=True,
        )

        self.assertIn(self.pending_subject, queryset)

    def test_subjects_endpoint_includes_own_pending_subject_when_include_current(self):
        self.client.force_authenticate(user=self.tutor_user)

        response = self.client.get("/api/subjects/", {"include_current": "1"})

        self.assertEqual(response.status_code, 200)
        codes = [row["subject_code"] for row in response.data]
        self.assertIn(self.pending_subject.subject_code, codes)


def _make_tiebreak_profile(username, role):
    user = User.objects.create_user(
        username=username,
        email=f"{username}@studybuddy.test",
        password="password",
    )
    return UserProfile.objects.create(
        user=user,
        fname=username,
        mname="",
        lname=role,
        role=role,
        profile_completed=True,
    )

class UpcomingWeekLoadTests(APITestCase):
    """Upcoming Week Load counts a tutor's booked sessions in the next 7 days
    (see docs/adr/0009-tie-breaker-upcoming-week-load.md). Deliberately distinct
    from Tutor.accepted_session_load(), which has no date bound."""

    def setUp(self):
        self.tutee_profile = _make_tiebreak_profile("load-tutee", "Tutee")
        self.tutor = Tutor.objects.create(profile=_make_tiebreak_profile("load-tutor", "Tutor"))
        self.availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day="Mon",
            time_slot=time(9, 0),
            is_active=True,
        )
        self.today = timezone.localdate()

    def _book(self, day_offset, status="Confirmed"):
        return Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=self.availability,
            session_date=self.today + timedelta(days=day_offset),
            session_mode="Online",
            status=status,
        )

    def _load(self):
        from studybuddy.recommender.workload import get_upcoming_week_loads

        return get_upcoming_week_loads([self.tutor.profile_id]).get(self.tutor.profile_id, 0)

    def test_counts_sessions_inside_the_window(self):
        self._book(0)
        self._book(3)
        self._book(6)

        self.assertEqual(self._load(), 3)

    def test_today_is_inside_the_window(self):
        self._book(0)

        self.assertEqual(self._load(), 1)

    def test_day_seven_is_outside_the_window(self):
        self._book(7)

        self.assertEqual(self._load(), 0)

    def test_past_sessions_are_excluded(self):
        self._book(-1)

        self.assertEqual(self._load(), 0)

    def test_awaiting_payment_verification_counts(self):
        self._book(2, status="Awaiting Payment Verification")

        self.assertEqual(self._load(), 1)

    def test_cancelled_rejected_and_completed_are_excluded(self):
        self._book(1, status="Cancelled")
        self._book(2, status="Rejected")
        self._book(3, status="Completed")

        self.assertEqual(self._load(), 0)

    def test_sessions_in_one_group_count_individually(self):
        # Unlike Accepted Session Load, a multi-session package is not collapsed:
        # each dated occurrence inside the window is real burden that week.
        group = uuid4()
        for offset in (1, 2, 3):
            booking = self._book(offset)
            booking.session_group_id = group
            booking.save()

        self.assertEqual(self._load(), 3)

    def test_tutor_with_no_bookings_is_absent_from_the_map(self):
        from studybuddy.recommender.workload import get_upcoming_week_loads

        loads = get_upcoming_week_loads([self.tutor.profile_id])

        self.assertEqual(loads.get(self.tutor.profile_id, 0), 0)

    def test_empty_tutor_id_list_makes_no_query(self):
        from studybuddy.recommender.workload import get_upcoming_week_loads

        with self.assertNumQueries(0):
            self.assertEqual(get_upcoming_week_loads([]), {})

    def test_loads_for_many_tutors_take_one_query(self):
        from studybuddy.recommender.workload import get_upcoming_week_loads

        other = Tutor.objects.create(profile=_make_tiebreak_profile("load-tutor-2", "Tutor"))
        self._book(1)

        with self.assertNumQueries(1):
            loads = get_upcoming_week_loads([self.tutor.profile_id, other.profile_id])

        self.assertEqual(loads.get(self.tutor.profile_id, 0), 1)
        self.assertEqual(loads.get(other.profile_id, 0), 0)


class TieBreakerOrderingTests(APITestCase):
    """The Tie Breaker: equal Hybrid Score quantized to 3dp ranks by Upcoming
    Week Load ascending, then profile_id ascending."""

    def _rank(self, scored, loads):
        """scored: list of (profile_id, hybrid_score). Returns ranked profile_ids."""
        from studybuddy.recommender import hybrid

        tutors = []
        for profile_id, _score in scored:
            tutor = Mock(profile_id=profile_id, tutorsubjects_set=Mock())
            tutor.tutorsubjects_set.all.return_value = []
            tutors.append(tutor)

        score_by_id = dict(scored)
        student_profile = Mock(id=1, course_id=None, year_level=None)

        def fake_cbf(_student, tutor, *_args, **_kwargs):
            # CF is coerced to 0 for this student, so hybrid = CBF_WEIGHT * cbf.
            return score_by_id[tutor.profile_id] / hybrid.CBF_WEIGHT

        with patch.object(hybrid, "get_student_subject_codes", return_value=[]), \
             patch.object(hybrid, "compute_cbf_score", side_effect=fake_cbf), \
             patch.object(hybrid, "get_upcoming_week_loads", return_value=loads), \
             patch.object(hybrid, "normalize_tutor_queryset", return_value=tutors):
            results = hybrid.recommend_tutors_hybrid({}, student_profile, None)

        return [row["tutor"].profile_id for row in results]

    def test_equal_scores_rank_by_fewer_upcoming_sessions(self):
        order = self._rank([(1, 0.5), (2, 0.5)], {1: 7, 2: 2})

        self.assertEqual(order, [2, 1])

    def test_a_real_score_gap_is_never_overridden_by_load(self):
        # 0.520 vs 0.500 is well outside the 3dp tie band: score wins despite
        # the higher-scoring tutor being far busier.
        order = self._rank([(1, 0.520), (2, 0.500)], {1: 9, 2: 0})

        self.assertEqual(order, [1, 2])

    def test_scores_differing_below_the_quantum_are_treated_as_tied(self):
        # 0.5004 and 0.5001 both quantize to 0.500, so load decides.
        order = self._rank([(1, 0.5004), (2, 0.5001)], {1: 5, 2: 1})

        self.assertEqual(order, [2, 1])

    def test_scores_differing_above_the_quantum_are_not_tied(self):
        # 0.5014 -> 0.501 and 0.5001 -> 0.500: different buckets, score decides
        # even though the winner is the busier tutor.
        order = self._rank([(1, 0.5001), (2, 0.5014)], {1: 0, 2: 8})

        self.assertEqual(order, [2, 1])

    def test_equal_score_and_equal_load_fall_back_to_profile_id(self):
        order = self._rank([(3, 0.5), (1, 0.5), (2, 0.5)], {1: 4, 2: 4, 3: 4})

        self.assertEqual(order, [1, 2, 3])

    def test_tutor_missing_from_the_load_map_is_treated_as_zero(self):
        order = self._rank([(1, 0.5), (2, 0.5)], {1: 3})

        self.assertEqual(order, [2, 1])

    def test_upcoming_week_load_is_attached_to_each_recommendation(self):
        from studybuddy.recommender import hybrid

        tutor = Mock(profile_id=1, tutorsubjects_set=Mock())
        tutor.tutorsubjects_set.all.return_value = []
        student_profile = Mock(id=1, course_id=None, year_level=None)

        with patch.object(hybrid, "get_student_subject_codes", return_value=[]), \
             patch.object(hybrid, "compute_cbf_score", return_value=0.5), \
             patch.object(hybrid, "get_upcoming_week_loads", return_value={1: 6}), \
             patch.object(hybrid, "normalize_tutor_queryset", return_value=[tutor]):
            results = hybrid.recommend_tutors_hybrid({}, student_profile, None)

        self.assertEqual(results[0]["upcoming_week_load"], 6)


class DemoTieGroupTests(APITestCase):
    """The demo tool tags tie groups so the panel can show which rows the Tie
    Breaker ordered — see _mark_tie_groups."""

    def _mark(self, scores):
        from studybuddy.recommender.demo import _mark_tie_groups

        rows = [{"hybrid_score": score} for score in scores]
        _mark_tie_groups(rows)
        return [row["tie_group_id"] for row in rows]

    def test_unique_scores_are_not_tagged(self):
        self.assertEqual(self._mark([0.9, 0.8, 0.7]), [None, None, None])

    def test_adjacent_equal_scores_share_a_group(self):
        self.assertEqual(self._mark([0.9, 0.845, 0.845, 0.7]), [None, 0, 0, None])

    def test_separate_ties_get_separate_groups(self):
        self.assertEqual(
            self._mark([0.9, 0.9, 0.845, 0.7, 0.7]),
            [0, 0, None, 1, 1],
        )

    def test_scores_equal_only_below_the_quantum_still_tie(self):
        self.assertEqual(self._mark([0.8452, 0.8449]), [0, 0])

    def test_scores_differing_above_the_quantum_do_not_tie(self):
        self.assertEqual(self._mark([0.846, 0.845]), [None, None])

    def test_empty_rows_are_handled(self):
        self.assertEqual(self._mark([]), [])


class LateCancellationSupportTicketTests(APITestCase):
    """Cancelling inside the Grace Cutoff opens a Late_Cancellation SupportTicket
    (backend/studybuddy/views.py:cancel_booking) which a SuperAdmin resolves via
    admin_resolve_ticket. See docs/architecture/booking-flow.md."""

    def setUp(self):
        self.institution = PartnerInstitution.objects.create(
            institution_name="CPU",
            school_email_domain="cpu.edu.ph",
            is_active=True,
        )
        self.super_user = User.objects.create_user(
            username="lc-superadmin",
            email="lc-superadmin@studybuddy.test",
            password="password",
        )
        self.super_profile = UserProfile.objects.create(
            user=self.super_user,
            fname="Super",
            mname="",
            lname="Admin",
            role="SuperAdmin",
            profile_completed=True,
            is_domain_exempt=True,
        )
        self.tutee_user = User.objects.create_user(
            username="lc-tutee@cpu.edu.ph",
            email="lc-tutee@cpu.edu.ph",
            password="password",
        )
        self.tutee_profile = UserProfile.objects.create(
            user=self.tutee_user,
            fname="Lc",
            mname="",
            lname="Tutee",
            role="Tutee",
            institution=self.institution,
        )
        self.tutor_user = User.objects.create_user(
            username="lc-tutor@cpu.edu.ph",
            email="lc-tutor@cpu.edu.ph",
            password="password",
        )
        self.tutor_profile = UserProfile.objects.create(
            user=self.tutor_user,
            fname="Lc",
            mname="",
            lname="Tutor",
            role="Tutor",
            institution=self.institution,
        )
        self.tutor = Tutor.objects.create(
            profile=self.tutor_profile,
            hourly_rate=Decimal("280.00"),
            can_online=True,
            can_f2f=True,
            teaching_level="SHS",
        )
        self.wallet, _ = Wallet.objects.get_or_create(tutor=self.tutor)
        self.wallet.balance = Decimal("500.00")
        self.wallet.save(update_fields=["balance"])

    def create_confirmed_booking(self, *, hours_from_now):
        """A Confirmed booking whose session starts `hours_from_now` from now."""
        session_start = timezone.localtime(timezone.now()) + timedelta(hours=hours_from_now)
        availability = TutorAvailability.objects.create(
            tutor=self.tutor,
            day=WEEKDAY_MAP[session_start.weekday()],
            time_slot=session_start.time(),
            is_active=True,
        )
        group_id = uuid4()
        return Booking.objects.create(
            student=self.tutee_profile,
            tutor=self.tutor,
            availability=availability,
            session_date=session_start.date(),
            session_mode="Online",
            booking_request_id=group_id,
            session_group_id=group_id,
            status="Confirmed",
        )

    def cancel_as(self, user, booking, *, reason="Something came up"):
        self.client.force_authenticate(user=user)
        return self.client.post(
            f"/api/bookings/{booking.id}/cancel/",
            {"reason": reason},
            format="json",
        )

    def test_cancellation_before_grace_cutoff_does_not_open_a_ticket(self):
        booking = self.create_confirmed_booking(hours_from_now=GRACE_CUTOFF_HOURS + 1)

        response = self.cancel_as(self.tutee_user, booking)

        self.assertEqual(response.status_code, 200)
        booking.refresh_from_db()
        self.assertEqual(booking.status, "Cancelled")
        self.assertFalse(SupportTicket.objects.filter(booking=booking).exists())

    def test_tutee_late_cancellation_opens_ticket_penalizing_the_tutee(self):
        booking = self.create_confirmed_booking(hours_from_now=GRACE_CUTOFF_HOURS - 1)

        response = self.cancel_as(self.tutee_user, booking, reason="Family emergency")

        self.assertEqual(response.status_code, 200)
        ticket = SupportTicket.objects.get(booking=booking)
        self.assertEqual(ticket.category, "Late_Cancellation")
        self.assertEqual(ticket.status, "Open")
        self.assertTrue(ticket.reported_by_system)
        self.assertEqual(ticket.user_id, self.tutor_profile.id)
        self.assertEqual(ticket.penalized_user_id, self.tutee_profile.id)
        self.assertIn("Family emergency", ticket.description)
        self.assertIsNone(ticket.resolution_verdict)

    def test_tutor_late_cancellation_opens_ticket_penalizing_the_tutor(self):
        booking = self.create_confirmed_booking(hours_from_now=GRACE_CUTOFF_HOURS - 1)

        response = self.cancel_as(self.tutor_user, booking, reason="Sudden illness")

        self.assertEqual(response.status_code, 200)
        ticket = SupportTicket.objects.get(booking=booking)
        self.assertEqual(ticket.user_id, self.tutee_profile.id)
        self.assertEqual(ticket.penalized_user_id, self.tutor_profile.id)

    def test_superadmin_counted_verdict_deducts_tutor_wallet(self):
        booking = self.create_confirmed_booking(hours_from_now=GRACE_CUTOFF_HOURS - 1)
        self.cancel_as(self.tutor_user, booking, reason="Sudden illness")
        ticket = SupportTicket.objects.get(booking=booking)

        self.client.force_authenticate(user=self.super_user)
        response = self.client.post(
            f"/api/admin/support/tickets/{ticket.id}/resolve/",
            {"verdict": "counted"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        ticket.refresh_from_db()
        self.assertEqual(ticket.status, "Resolved")
        self.assertEqual(ticket.resolution_verdict, "counted")
        self.wallet.refresh_from_db()
        self.assertEqual(self.wallet.balance, Decimal("500.00") - COUNTED_STRIKE_WALLET_DEDUCTION)
        self.assertTrue(
            Transaction.objects.filter(
                wallet=self.wallet,
                transaction_type="counted_strike",
                reference_id=f"LT-{ticket.id}",
            ).exists()
        )
        self.assertEqual(response.data["monthly_counted_strikes"], 1)

    def test_superadmin_excused_verdict_leaves_wallet_untouched(self):
        booking = self.create_confirmed_booking(hours_from_now=GRACE_CUTOFF_HOURS - 1)
        self.cancel_as(self.tutor_user, booking, reason="Sudden illness")
        ticket = SupportTicket.objects.get(booking=booking)

        self.client.force_authenticate(user=self.super_user)
        response = self.client.post(
            f"/api/admin/support/tickets/{ticket.id}/resolve/",
            {"verdict": "excused"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        ticket.refresh_from_db()
        self.assertEqual(ticket.resolution_verdict, "excused")
        self.wallet.refresh_from_db()
        self.assertEqual(self.wallet.balance, Decimal("500.00"))
        self.assertFalse(Transaction.objects.filter(wallet=self.wallet).exists())

    def test_ticket_cannot_be_resolved_twice(self):
        booking = self.create_confirmed_booking(hours_from_now=GRACE_CUTOFF_HOURS - 1)
        self.cancel_as(self.tutor_user, booking, reason="Sudden illness")
        ticket = SupportTicket.objects.get(booking=booking)

        self.client.force_authenticate(user=self.super_user)
        self.client.post(
            f"/api/admin/support/tickets/{ticket.id}/resolve/",
            {"verdict": "excused"},
            format="json",
        )
        second_response = self.client.post(
            f"/api/admin/support/tickets/{ticket.id}/resolve/",
            {"verdict": "counted"},
            format="json",
        )

        self.assertEqual(second_response.status_code, 400)
